from datetime import date, datetime, timedelta, timezone
from openpyxl import load_workbook
import os
from decimal import Decimal, InvalidOperation
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, Response, session, send_file
import os
from pathlib import Path
from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, inspect, text
import csv
import io
import os
import shutil
import json
import requests
import re
from pypdf import PdfReader
from pathlib import Path
from werkzeug.utils import secure_filename
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
)
from bank_import import import_bank_statement
from services.dashboard_service import DashboardService
from services.member_intelligence_service import MemberIntelligenceService
from services.loan_intelligence_service import LoanIntelligenceService
from services.welfare_intelligence_service import WelfareIntelligenceService
from services.meeting_intelligence_service import MeetingIntelligenceService
from services.financial_intelligence_service import FinancialIntelligenceService
from io import BytesIO
from flask import send_file
from flask import send_from_directory
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'change-this-secret-key')
ALLOWED_LOGO_EXTENSIONS = {
    'png',
    'jpg',
    'jpeg',
    'webp',
}

if os.environ.get('LOGO_UPLOAD_FOLDER'):
    # Use a configured persistent folder when available.
    LOGO_UPLOAD_FOLDER = Path(
        os.environ['LOGO_UPLOAD_FOLDER']
    )

elif os.environ.get('RENDER'):
    # Writable temporary storage on Render.
    # Files stored here will not survive every restart or redeploy.
    LOGO_UPLOAD_FOLDER = Path('/tmp/logos')

else:
    # Local development storage.
    LOGO_UPLOAD_FOLDER = (
        Path(app.root_path)
        / 'static'
        / 'logos'
    )

LOGO_UPLOAD_FOLDER.mkdir(
    parents=True,
    exist_ok=True
)

app.config['MAX_CONTENT_LENGTH'] = (
    5 * 1024 * 1024
)

# Local development uses SQLite. Render/production uses PostgreSQL through DATABASE_URL.
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Render provides postgresql:// URLs. Use the psycopg v3 SQLAlchemy driver.
    if database_url.startswith('postgresql://'):
        database_url = database_url.replace('postgresql://', 'postgresql+psycopg://', 1)
    elif database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = database_url or 'sqlite:///village_banking.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

INTEREST_RATE = Decimal('0.15')
PAYMENT_METHODS = ['Bank Transfer', 'Mobile Money', 'Cash']
ABSENCE_FINE_AMOUNT = Decimal('50.00')
LATE_ATTENDANCE_FINE_AMOUNT = Decimal('20.00')
LATE_CONTRIBUTION_FINE_AMOUNT = Decimal('10.00')
CLIENT_NAME = 'Higher Achievers'
PRODUCER_NAME = 'SL Consulting Limited'
BACKUP_RETENTION = 30
ACCUMULATED_SURPLUS_ACCOUNT_CODE = "3000"

FINE_CATEGORIES = [
    'Late Meeting Attendance',
    'Absence from Meeting',
    'Late Contribution',
    'Late Loan Repayment',
    'Rule Violation',
    'Other Penalty',
]
FINE_STATUSES = ['Unpaid', 'Partially Paid', 'Paid', 'Waived']

WELFARE_CATEGORIES = [
    'Funeral Support',
    'Hospitalization',
    'Birth Support',
    'School Emergency',
    'House Fire',
    'Other Emergency',
]
WELFARE_STATUSES = ['Requested', 'Reviewed', 'Approved', 'Paid', 'Rejected']
ATTENDANCE_STATUSES = ['Present', 'Absent', 'Late', 'Excused']
NOTIFICATION_CHANNELS = ['SMS', 'WhatsApp']
NOTIFICATION_TYPES = ['Contribution Reminder', 'Loan Repayment Reminder', 'Meeting Reminder', 'Welfare Notification', 'Share-Out Notification', 'General Notice']
ACCOUNT_TYPES = ['Asset', 'Liability', 'Equity', 'Income', 'Expense']

ROLES = ['Administrator', 'Chairperson', 'Treasurer', 'Secretary', 'Auditor', 'Data Clerk']
ROLE_PERMISSIONS = {
    'Administrator': [
    'dashboard',
    'members',
    'contributions',
    'loans',
    'repayments',
    'distributions',
    'meetings',
    'attendance',
    'reports',
    'statements',
    'shareout',
    'fines',
    'welfare',
    'users',
    'audit',
    'backups',
    'notifications',
    'accounting',
    'exports',
    'settings'
],
    'Chairperson': ['dashboard', 'loans', 'distributions', 'meetings', 'attendance', 'reports', 'statements', 'shareout', 'fines', 'welfare', 'audit', 'notifications', 'accounting', 'exports'],
    'Treasurer': ['dashboard', 'contributions', 'loans', 'repayments', 'distributions', 'reports', 'statements', 'shareout', 'fines', 'welfare', 'notifications', 'accounting', 'exports'],
    'Secretary': ['dashboard', 'members', 'meetings', 'attendance', 'reports', 'statements', 'welfare', 'notifications', 'accounting', 'exports'],
    'Auditor': ['dashboard', 'attendance', 'reports', 'statements', 'shareout', 'fines', 'welfare', 'audit', 'notifications', 'accounting', 'exports'],
    'Data Clerk': ['dashboard', 'members', 'contributions', 'notifications'],
}

def get_pdf_logo_path(setting):
    """
    Return the best available logo for PDF documents.

    Priority:
    1. Report logo
    2. System logo
    3. Default application logo
    """

    logo_filename = None

    if setting:
        logo_filename = (
            setting.report_logo
            or setting.logo
        )

    if logo_filename:
        uploaded_path = (
            LOGO_UPLOAD_FOLDER
            / Path(logo_filename).name
        )

        if uploaded_path.exists():
            return str(uploaded_path)

    default_path = (
        Path(app.root_path)
        / 'static'
        / 'higher-achievers-logo.jpeg'
    )

    if default_path.exists():
        return str(default_path)

    return None

def pdf_colour(value, fallback):
    """
    Safely convert a stored hexadecimal colour into a ReportLab colour.
    """

    try:
        return colors.HexColor(
            value or fallback
        )
    except Exception:
        return colors.HexColor(fallback)

def build_pdf_branding(setting, styles):
    """
    Build reusable organisation branding for PDF documents.
    """

    organization_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    registration_number = (
        setting.registration_number
        if setting and setting.registration_number
        else ''
    )

    organization_address = (
        setting.organization_address
        if setting and setting.organization_address
        else ''
    )

    organization_phone = (
        setting.organization_phone
        if setting and setting.organization_phone
        else ''
    )

    organization_email = (
        setting.organization_email
        if setting and setting.organization_email
        else ''
    )

    motto = (
        setting.motto
        if setting and setting.motto
        else ''
    )

    primary_colour = pdf_colour(
        setting.primary_color if setting else None,
        '#0D6EFD'
    )

    secondary_colour = pdf_colour(
        setting.secondary_color if setting else None,
        '#198754'
    )

    logo_path = get_pdf_logo_path(setting)

    organisation_style = ParagraphStyle(
        'PdfOrganisationName',
        parent=styles['Title'],
        fontSize=16,
        leading=19,
        textColor=primary_colour,
        spaceAfter=3,
    )

    contact_style = ParagraphStyle(
        'PdfContactDetails',
        parent=styles['Normal'],
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor('#555555'),
    )

    motto_style = ParagraphStyle(
        'PdfMotto',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=secondary_colour,
        italic=True,
    )

    organisation_parts = [
        Paragraph(
            organization_name,
            organisation_style
        )
    ]

    if motto:
        organisation_parts.append(
            Paragraph(
                motto,
                motto_style
            )
        )

    contact_parts = []

    if registration_number:
        contact_parts.append(
            f'Registration No: {registration_number}'
        )

    if organization_address:
        contact_parts.append(
            organization_address
        )

    if organization_phone:
        contact_parts.append(
            organization_phone
        )

    if organization_email:
        contact_parts.append(
            organization_email
        )

    if contact_parts:
        organisation_parts.append(
            Paragraph(
                ' | '.join(contact_parts),
                contact_style
            )
        )

    organisation_content = []

    if logo_path:
        try:
            logo = Image(
                logo_path,
                width=24 * mm,
                height=24 * mm,
                kind='proportional',
            )

            organisation_content.append(logo)

        except Exception:
            organisation_content.append('')

    else:
        organisation_content.append('')

    organisation_content.append(
        organisation_parts
    )

    header_table = Table(
        [organisation_content],
        colWidths=[
            30 * mm,
            135 * mm,
        ],
    )

    header_table.setStyle(
        TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            (
                'LINEBELOW',
                (0, 0),
                (-1, -1),
                1.2,
                primary_colour,
            ),
        ])
    )

    return header_table

def draw_pdf_footer(canvas, document, setting):
    canvas.saveState()

    organization_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    product_name = (
        setting.product_name
        if setting and setting.product_name
        else 'SL Village Banking Pro'
    )

    developer_name = (
        setting.developer_name
        if setting and setting.developer_name
        else 'SL Consulting Limited'
    )

    primary_colour = pdf_colour(
        setting.primary_color if setting else None,
        '#0D6EFD'
    )

    page_width, _ = document.pagesize

    canvas.setStrokeColor(primary_colour)
    canvas.setLineWidth(0.6)
    canvas.line(
        document.leftMargin,
        13 * mm,
        page_width - document.rightMargin,
        13 * mm,
    )

    canvas.setFont('Helvetica', 7)
    canvas.setFillColor(
        colors.HexColor('#555555')
    )

    canvas.drawString(
        document.leftMargin,
        8 * mm,
        f'{organization_name} | {product_name}'
    )

    canvas.drawRightString(
        page_width - document.rightMargin,
        8 * mm,
        (
            f'Page {canvas.getPageNumber()} | '
            f'Produced by {developer_name}'
        )
    )

    canvas.restoreState()

def format_zambian_phone(phone):
    if not phone:
        return ''

    phone = str(phone).strip().replace(' ', '').replace('-', '')

    if phone.startswith('+260'):
        return phone

    if phone.startswith('260'):
        return '+' + phone

    if phone.startswith('0'):
        return '+260' + phone[1:]

    return '+260' + phone
def allowed_logo_file(filename):
    return (
        '.' in filename
        and filename.rsplit('.', 1)[1].lower()
        in ALLOWED_LOGO_EXTENSIONS
    )


def save_logo_file(uploaded_file, logo_type):
    if not uploaded_file or not uploaded_file.filename:
        return None

    if not allowed_logo_file(uploaded_file.filename):
        raise ValueError(
            'Only PNG, JPG, JPEG and WEBP '
            'images are allowed.'
        )

    original_name = secure_filename(
        uploaded_file.filename
    )

    extension = original_name.rsplit(
        '.',
        1
    )[1].lower()

    filename = (
        f'{logo_type}.{extension}'
    )

    destination = (
        LOGO_UPLOAD_FOLDER / filename
    )

    for existing_file in (
        LOGO_UPLOAD_FOLDER.glob(
            f'{logo_type}.*'
        )
    ):
        if existing_file.is_file():
            existing_file.unlink()

    uploaded_file.save(destination)

    return filename


def delete_logo_file(filename):
    if not filename:
        return

    filepath = (
        LOGO_UPLOAD_FOLDER
        / Path(filename).name
    )

    if filepath.exists() and filepath.is_file():
        filepath.unlink()

def send_sms_via_africas_talking(phone, message):
    setting = get_settings()

    username = setting.sms_username
    api_key = setting.sms_api_key
    sender_id = setting.sms_sender_id

    if not username or not api_key:
        return False, "Africa's Talking username or API key is missing."

    formatted_phone = format_zambian_phone(phone)

    payload = {
        'username': username,
        'to': formatted_phone,
        'message': message,
    }

    if sender_id and sender_id.strip():
        payload['from'] = sender_id.strip()

    response = requests.post(
        'https://api.africastalking.com/version1/messaging',
        headers={
            'apiKey': api_key,
            'Accept': 'application/json',
            'Content-Type': 'application/x-www-form-urlencoded',
        },
        data=payload,
        timeout=20
    )

    if response.status_code not in [200, 201]:
        return False, response.text

    try:
        data = response.json()

        recipients = (
            data.get('SMSMessageData', {})
                .get('Recipients', [])
        )

        if not recipients:
            return False, response.text

        recipient = recipients[0]

        status = str(
            recipient.get('status')
            or recipient.get('statusCode')
            or ''
        ).lower()

        if status == 'success' or status in ['100', '101', '102']:
            return True, response.text

        return False, response.text

    except Exception:
        return False, response.text
    
def ensure_month_end_columns():
    columns = {
        'members_processed': 'INTEGER DEFAULT 0',
        'loans_processed': 'INTEGER DEFAULT 0',
        'reversed': 'BOOLEAN DEFAULT FALSE',
        'reversed_by': 'VARCHAR(120)',
        'reversed_on': 'DATE',
        'reversal_reason': 'VARCHAR(250)',
    }

    for column, definition in columns.items():
        try:
            db.session.execute(
                db.text(f'ALTER TABLE month_end_process ADD COLUMN IF NOT EXISTS {column} {definition}')
            )
            db.session.commit()
        except Exception:
            db.session.rollback()

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(30), default='Administrator')
    password_hash = db.Column(db.String(255), nullable=False)
    active = db.Column(db.Boolean, default=True)


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer)
    username = db.Column(db.String(50))
    full_name = db.Column(db.String(120))
    role = db.Column(db.String(30))
    action = db.Column(db.String(80), nullable=False)
    entity = db.Column(db.String(80))
    entity_id = db.Column(db.String(40))
    details = db.Column(db.Text)
    ip_address = db.Column(db.String(60))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class BackupRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    file_size = db.Column(db.Integer, default=0)
    backup_type = db.Column(db.String(30), default='Manual')
    created_by = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    notes = db.Column(db.Text)


class NotificationLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    channel = db.Column(db.String(30), nullable=False)
    notification_type = db.Column(db.String(80), nullable=False)
    recipient_type = db.Column(db.String(80))
    phone = db.Column(db.String(40))
    subject = db.Column(db.String(160))
    message = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(30), default='Prepared')
    created_by = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    sent_at = db.Column(db.DateTime)
    provider_reference = db.Column(db.String(120))
    member_id = db.Column(
        db.Integer,
        db.ForeignKey('member.id'),
        nullable=True
    )
    member = db.relationship('Member', backref='notification_logs')

class Member(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_no = db.Column(db.String(20), unique=True, nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    phone = db.Column(db.String(30))
    national_id = db.Column(db.String(50))
    group_name = db.Column(db.String(80))
    status = db.Column(db.String(20), default='Active')
    created_at = db.Column(db.Date, default=date.today)
    member_type = db.Column(db.String(50), default='Ordinary Member')
    committee_position = db.Column(db.String(100))

class Contribution(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(
        db.Integer,
        db.ForeignKey('member.id'),
        nullable=False
    )
    month = db.Column(db.String(7), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    method = db.Column(db.String(30), nullable=False)
    reference = db.Column(db.String(80))
    paid_on = db.Column(db.Date, default=date.today)

    source_type = db.Column(
        db.String(30),
        default='Normal',
        nullable=False
    )

    source_id = db.Column(db.Integer)

    member = db.relationship(
        'Member',
        backref='contributions'
    )

class Loan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    principal = db.Column(db.Numeric(12, 2), nullable=False)
    interest_rate = db.Column(db.Numeric(5, 2), default=INTEREST_RATE)
    issued_on = db.Column(db.Date, default=date.today)
    due_on = db.Column(db.Date, nullable=False)
    purpose = db.Column(db.String(200))
    status = db.Column(db.String(20), default='Applied')  # Applied, Reviewed, Approved, Disbursed, Partially Paid, Paid, Rejected
    approved_by = db.Column(db.String(120), default='')
    reviewed_by = db.Column(db.String(120))
    disbursed_by = db.Column(db.String(120))
    reviewed_on = db.Column(db.Date)
    approved_on = db.Column(db.Date)
    disbursed_on = db.Column(db.Date)
    rejected_on = db.Column(db.Date)
    rejection_reason = db.Column(db.String(250))
    member = db.relationship('Member', backref='loans')
    loan_no = db.Column(db.String(30), unique=True)

    disbursed_by = db.Column(db.String(120))
    disbursed_on = db.Column(db.Date)

    disbursement_method = db.Column(db.String(50))
    disbursement_reference = db.Column(db.String(100))
    source_type = db.Column(
    db.String(30),
    default='Normal',
    nullable=False
)

    source_id = db.Column(db.Integer)
    opening_interest = db.Column(
    db.Numeric(12,2),
    default=0,
    nullable=False
    )

    @property
    def interest_amount(self):
        return money(self.principal * self.interest_rate)

    @property
    def total_due(self):
        return money(self.principal + self.interest_amount)

    @property
    def total_paid(self):
        return money(sum((r.amount for r in self.repayments), Decimal('0.00')))

    @property
    def balance(self):
        return money(self.total_due - self.total_paid)

    @property
    def overdue(self):
        return self.status in ['Disbursed', 'Partially Paid'] and self.due_on < date.today() and self.balance > 0

class LoanGuarantor(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    guarantor_member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)

    guaranteed_amount = db.Column(db.Numeric(12, 2), default=0)
    notes = db.Column(db.String(250))
    created_at = db.Column(db.Date, default=date.today)

    loan = db.relationship('Loan', backref='guarantors')
    guarantor = db.relationship('Member', foreign_keys=[guarantor_member_id])

class Repayment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    method = db.Column(db.String(30), nullable=False)
    reference = db.Column(db.String(80))
    paid_on = db.Column(db.Date, default=date.today)
    loan = db.relationship('Loan', backref='repayments')

class Distribution(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    method = db.Column(db.String(30), nullable=False)
    reference = db.Column(db.String(80))
    authorized_by = db.Column(db.String(120))
    paid_on = db.Column(db.Date, default=date.today)
    member = db.relationship('Member', backref='distributions')


class WelfareContribution(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    month = db.Column(db.String(7), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    method = db.Column(db.String(30), nullable=False)
    reference = db.Column(db.String(80))
    paid_on = db.Column(db.Date, default=date.today)
    member = db.relationship('Member', backref='welfare_contributions')
    source_type = db.Column(
    db.String(30),
    default='Normal',
    nullable=False
)

    source_id = db.Column(db.Integer)

class WelfareClaim(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    category = db.Column(db.String(60), nullable=False)
    amount_requested = db.Column(db.Numeric(12, 2), nullable=False)
    amount_approved = db.Column(db.Numeric(12, 2), default=0)
    status = db.Column(db.String(20), default='Requested')
    reason = db.Column(db.String(250))
    requested_on = db.Column(db.Date, default=date.today)
    reviewed_by = db.Column(db.String(120))
    reviewed_on = db.Column(db.Date)
    approved_by = db.Column(db.String(120))
    approved_on = db.Column(db.Date)
    paid_by = db.Column(db.String(120))
    paid_on = db.Column(db.Date)
    payment_method = db.Column(db.String(30))
    reference = db.Column(db.String(80))
    rejection_reason = db.Column(db.String(250))
    member = db.relationship('Member', backref='welfare_claims')

    @property
    def payable_amount(self):
        return money(self.amount_approved or 0)



class FinePenalty(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    category = db.Column(db.String(60), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    reason = db.Column(db.String(250))
    fine_date = db.Column(db.Date, default=date.today)
    status = db.Column(db.String(20), default='Unpaid')
    recorded_by = db.Column(db.String(120))
    waived_by = db.Column(db.String(120))
    waived_on = db.Column(db.Date)
    member = db.relationship('Member', backref='fines')
    source_type = db.Column(
    db.String(30),
    default='Normal',
    nullable=False
)

    source_id = db.Column(db.Integer)

    @property
    def total_paid(self):
        return money(sum((p.amount for p in self.payments), Decimal('0.00')))

    @property
    def balance(self):
        return money(self.amount - self.total_paid)
    
class FinePayment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fine_id = db.Column(db.Integer, db.ForeignKey('fine_penalty.id'), nullable=False)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    method = db.Column(db.String(30), nullable=False)
    reference = db.Column(db.String(80))
    paid_on = db.Column(db.Date, default=date.today)
    fine = db.relationship('FinePenalty', backref='payments')

class Meeting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    meeting_type = db.Column(db.String(40), nullable=False)
    meeting_date = db.Column(db.Date, nullable=False)
    agenda = db.Column(db.Text)
    resolutions = db.Column(db.Text)
    attendance_count = db.Column(db.Integer, default=0)


class MeetingAttendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    meeting_id = db.Column(db.Integer, db.ForeignKey('meeting.id'), nullable=False)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    status = db.Column(db.String(20), default='Present')
    remarks = db.Column(db.String(250))
    recorded_by = db.Column(db.String(120))
    recorded_at = db.Column(db.DateTime, default=datetime.utcnow)
    fine_generated = db.Column(db.Boolean, default=False)
    meeting = db.relationship('Meeting', backref='attendance_records')
    member = db.relationship('Member', backref='attendance_records')


class Account(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    account_type = db.Column(db.String(30), nullable=False)
    normal_balance = db.Column(db.String(10), default='Debit')
    active = db.Column(db.Boolean, default=True)

class FinancialYear(db.Model):
    __tablename__ = 'financial_years'

    id = db.Column(db.Integer, primary_key=True)

    year = db.Column(
        db.Integer,
        unique=True,
        nullable=False,
        index=True
    )

    start_date = db.Column(
        db.Date,
        nullable=False
    )

    end_date = db.Column(
        db.Date,
        nullable=False
    )

    status = db.Column(
        db.String(20),
        nullable=False,
        default='Open'
    )

    opened_on = db.Column(
        db.DateTime,
        nullable=False,
        default=lambda: datetime.now(timezone.utc)
    )

    opened_by = db.Column(
        db.String(120)
    )

    closed_on = db.Column(
        db.DateTime
    )

    closed_by = db.Column(
        db.String(120)
    )

    closing_reference = db.Column(
        db.String(50)
    )

    is_locked = db.Column(
        db.Boolean,
        nullable=False,
        default=False
    )

    created_on = db.Column(
        db.DateTime,
        nullable=False,
        default=lambda: datetime.now(timezone.utc)
    )

    updated_on = db.Column(
        db.DateTime,
        nullable=False,
        default=lambda: datetime.now(timezone.utc),
        onupdate=lambda: datetime.now(timezone.utc)
    )

    def __repr__(self):
        return (
            f"<FinancialYear "
            f"{self.year} - {self.status}>"
        )

class JournalEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    entry_date = db.Column(db.Date, default=date.today, nullable=False)
    description = db.Column(db.String(250), nullable=False)
    reference = db.Column(db.String(100))
    source_type = db.Column(db.String(60))
    source_id = db.Column(db.String(60))
    posted_by = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    @property
    def total_debit(self):
        return money(sum((line.debit for line in self.lines), Decimal('0.00')))

    @property
    def total_credit(self):
        return money(sum((line.credit for line in self.lines), Decimal('0.00')))

    @property
    def balanced(self):
        return self.total_debit == self.total_credit


class JournalLine(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    journal_entry_id = db.Column(db.Integer, db.ForeignKey('journal_entry.id'), nullable=False)
    account_id = db.Column(db.Integer, db.ForeignKey('account.id'), nullable=False)
    debit = db.Column(db.Numeric(12, 2), default=0)
    credit = db.Column(db.Numeric(12, 2), default=0)
    memo = db.Column(db.String(250))
    entry = db.relationship('JournalEntry', backref='lines')
    account = db.relationship('Account', backref='journal_lines')

class SavingsInterest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    month = db.Column(db.String(7), nullable=False)  # YYYY-MM
    opening_balance = db.Column(db.Numeric(12, 2), nullable=False)
    interest_rate = db.Column(db.Numeric(5, 2), default=15.00)
    interest_amount = db.Column(db.Numeric(12, 2), nullable=False)
    closing_balance = db.Column(db.Numeric(12, 2), nullable=False)
    processed_on = db.Column(db.Date, default=date.today)

    member = db.relationship('Member', backref='savings_interest')


class LoanInterest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    member_id = db.Column(db.Integer, db.ForeignKey('member.id'), nullable=False)
    month = db.Column(db.String(7), nullable=False)  # YYYY-MM
    opening_balance = db.Column(db.Numeric(12, 2), nullable=False)
    interest_rate = db.Column(db.Numeric(5, 2), default=15.00)
    interest_amount = db.Column(db.Numeric(12, 2), nullable=False)
    closing_balance = db.Column(db.Numeric(12, 2), nullable=False)
    processed_on = db.Column(db.Date, default=date.today)

    loan = db.relationship('Loan', backref='loan_interest_entries')
    member = db.relationship('Member', backref='loan_interest_entries')
    source_type = db.Column(
    db.String(30),
    default='Month End',
    nullable=False
)

    source_id = db.Column(db.Integer)

class MonthEndProcess(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    month = db.Column(db.String(7), nullable=False)  # YYYY-MM

    savings_interest_total = db.Column(db.Numeric(12, 2), default=0)
    loan_interest_total = db.Column(db.Numeric(12, 2), default=0)

    members_processed = db.Column(db.Integer, default=0)
    loans_processed = db.Column(db.Integer, default=0)

    processed_by = db.Column(db.String(120))
    processed_on = db.Column(db.Date, default=date.today)

    reversed = db.Column(db.Boolean, default=False)
    reversed_by = db.Column(db.String(120))
    reversed_on = db.Column(db.Date)
    reversal_reason = db.Column(db.String(250))

class SystemSetting(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    organisation_name = db.Column(
    'organization_name',
    db.String(200),
    default='Your Organisation Name'
)
    contribution_amount = db.Column(db.Numeric(12, 2), default=100)
    savings_interest_rate = db.Column(db.Numeric(5, 2), default=15)
    loan_interest_rate = db.Column(db.Numeric(5, 2), default=15)
    welfare_contribution_amount = db.Column(db.Numeric(12, 2), default=0)

    updated_on = db.Column(db.DateTime, default=datetime.utcnow)
    organization_address = db.Column(db.String(250))
    organization_phone = db.Column(db.String(50))
    organization_email = db.Column(db.String(120))
    registration_number = db.Column(db.String(100))

    sms_provider = db.Column(db.String(50), default='Manual')
    sms_api_key = db.Column(db.String(500))
    sms_sender_id = db.Column(db.String(100))
    whatsapp_enabled = db.Column(db.Boolean, default=False)
    sms_username = db.Column(db.String(100))

        # Branding and product identity
    short_name = db.Column(db.String(80), default='Village Banking')
    motto = db.Column(
        db.String(200),
        default='Empowering Community Finance'
    )
    developer_name = db.Column(
        db.String(150),
        default='SL Consulting Limited'
    )
    product_name = db.Column(
        db.String(150),
        default='SL Village Banking Pro'
    )
    product_version = db.Column(
        db.String(50),
        default='1.0.0'
    )
    logo = db.Column(db.String(255))
    
    report_logo = db.Column(db.String(255))
    watermark_logo = db.Column(db.String(255))
    favicon = db.Column(db.String(255))

    # Theme and visual customization
    theme_name = db.Column(
        db.String(80),
        default='Professional Blue'
    )

    primary_color = db.Column(
        db.String(20),
        default='#0D6EFD'
    )

    secondary_color = db.Column(
        db.String(20),
        default='#198754'
    )

    accent_color = db.Column(
        db.String(20),
        default='#6F42C1'
    )

    sidebar_color = db.Column(
        db.String(20),
        default='#17324D'
    )

    sidebar_text_color = db.Column(
        db.String(20),
        default='#FFFFFF'
    )

    navbar_color = db.Column(
        db.String(20),
        default='#FFFFFF'
    )

    button_color = db.Column(
        db.String(20),
        default='#0D6EFD'
    )

    background_color = db.Column(
        db.String(20),
        default='#F4F7FB'
    )

    card_color = db.Column(
        db.String(20),
        default='#FFFFFF'
    )

    table_header_color = db.Column(
        db.String(20),
        default='#EAF2F8'
    )

    success_color = db.Column(
        db.String(20),
        default='#198754'
    )

    warning_color = db.Column(
        db.String(20),
        default='#FFC107'
    )

    danger_color = db.Column(
        db.String(20),
        default='#DC3545'
    )

    # Additional contact information
    phone = db.Column(db.String(50))
    email = db.Column(db.String(120))
    website = db.Column(db.String(150))
    postal_address = db.Column(db.Text)
    physical_address = db.Column(db.Text)

    # Currency configuration
    currency = db.Column(db.String(20), default='ZMW')
    currency_symbol = db.Column(db.String(10), default='K')
    decimal_places = db.Column(db.Integer, default=2)

    # Financial defaults
    default_interest_rate = db.Column(
        db.Numeric(10, 2),
        default=15.00
    )
    default_loan_term = db.Column(db.Integer, default=6)
    penalty_rate = db.Column(
        db.Numeric(10, 2),
        default=0.00
    )
    share_out_month = db.Column(
        db.String(20),
        default='December'
    )

    # Meeting defaults
    committee_meeting_frequency = db.Column(
        db.String(50),
        default='Monthly'
    )
    member_meeting_frequency = db.Column(
        db.String(50),
        default='Quarterly'
    )

    # Dashboard options
    enable_ai_advisor = db.Column(db.Boolean, default=True)
    enable_notifications = db.Column(db.Boolean, default=True)
    enable_dashboard_charts = db.Column(
        db.Boolean,
        default=True
    )

    created_at = db.Column(
        db.DateTime,
        default=datetime.utcnow
    )

class CashBookEntry(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    entry_date = db.Column(db.Date, default=date.today, nullable=False)
    entry_type = db.Column(db.String(20), nullable=False)  # In or Out
    category = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(250))
    amount = db.Column(db.Numeric(12, 2), nullable=False)

    method = db.Column(db.String(50))
    reference = db.Column(db.String(100))

    source_type = db.Column(db.String(100))
    source_id = db.Column(db.Integer)

    created_by = db.Column(db.String(120))
    created_at = db.Column(db.Date, default=date.today)

class BankStatementLine(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    statement_date = db.Column(db.Date, nullable=False)
    description = db.Column(db.String(250))
    reference = db.Column(db.String(100))
    amount = db.Column(db.Numeric(12, 2), nullable=False)

    entry_type = db.Column(db.String(20), nullable=False)  # In or Out

    reconciled = db.Column(db.Boolean, default=False)
    cash_book_entry_id = db.Column(db.Integer, db.ForeignKey('cash_book_entry.id'))
    cash_book_entry = db.relationship('CashBookEntry')

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.String(120))

class BankStatementImportBatch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    file_name = db.Column(db.String(250))
    bank_name = db.Column(db.String(80), default='FNB')
    imported_on = db.Column(db.DateTime, default=datetime.utcnow)
    imported_by = db.Column(db.String(120))
    line_count = db.Column(db.Integer, default=0)

    import_batch_id = db.Column(db.Integer, db.ForeignKey('bank_statement_import_batch.id'))
    import_batch = db.relationship('BankStatementImportBatch')

class BankReconciliation(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    reconciliation_date = db.Column(db.Date, nullable=False)

    bank_statement_balance = db.Column(db.Numeric(12, 2), default=0)
    outstanding_deposits = db.Column(db.Numeric(12, 2), default=0)
    unpresented_cheques = db.Column(db.Numeric(12, 2), default=0)
    adjusted_bank_balance = db.Column(db.Numeric(12, 2), default=0)
    cash_book_balance = db.Column(db.Numeric(12, 2), default=0)
    difference = db.Column(db.Numeric(12, 2), default=0)

    status = db.Column(db.String(30), default='Completed')

    prepared_by = db.Column(db.String(120))
    prepared_on = db.Column(db.DateTime, default=datetime.utcnow)

    notes = db.Column(db.Text)

class ShareOutCycle(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    start_month = db.Column(
        db.String(7),
        nullable=False
    )

    end_month = db.Column(
        db.String(7),
        nullable=False
    )

    expenses = db.Column(
        db.Numeric(12, 2),
        nullable=False,
        default=0
    )

    other_income = db.Column(
        db.Numeric(12, 2),
        nullable=False,
        default=0
    )

    total_contributions = db.Column(
        db.Numeric(12, 2),
        nullable=False,
        default=0
    )

    shareout_fund = db.Column(
        db.Numeric(12, 2),
        nullable=False,
        default=0
    )

    total_net_payable = db.Column(
        db.Numeric(12, 2),
        nullable=False,
        default=0
    )

    eligible_members = db.Column(
        db.Integer,
        nullable=False,
        default=0
    )

    members_requiring_review = db.Column(
        db.Integer,
        nullable=False,
        default=0
    )

    readiness_score = db.Column(
        db.Integer,
        nullable=False,
        default=0
    )

    status = db.Column(
        db.String(30),
        nullable=False,
        default='Draft'
    )

    approved_by = db.Column(
        db.String(120)
    )

    approved_on = db.Column(
        db.DateTime
    )

    approval_comments = db.Column(
        db.Text
    )

    locked_by = db.Column(
        db.String(120)
    )

    locked_on = db.Column(
        db.DateTime
    )

    lock_comments = db.Column(
        db.Text
    )

    created_by = db.Column(
        db.String(120)
    )

    created_at = db.Column(
        db.DateTime,
        nullable=False,
        default=datetime.utcnow
    )

    updated_at = db.Column(
        db.DateTime,
        nullable=False,
        default=datetime.utcnow,
        onupdate=datetime.utcnow
    )

    __table_args__ = (
        db.UniqueConstraint(
            'start_month',
            'end_month',
            name='uq_shareout_cycle_period'
        ),
    )
class OpeningBalanceBatch(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    batch_no = db.Column(
        db.String(30),
        unique=True,
        nullable=False
    )

    effective_date = db.Column(
        db.Date,
        nullable=False
    )

    description = db.Column(db.String(250))

    status = db.Column(
        db.String(20),
        default='Draft',
        nullable=False
    )
    # Draft, Validated, Approved, Posted, Locked, Reversed

    total_members = db.Column(
        db.Integer,
        default=0,
        nullable=False
    )

    total_savings = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    total_loan_principal = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    total_loan_interest = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    total_fines = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    total_welfare = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    imported_by = db.Column(db.String(120))
    approved_by = db.Column(db.String(120))
    posted_by = db.Column(db.String(120))
    reversed_by = db.Column(db.String(120))

    imported_on = db.Column(
        db.DateTime,
        default=datetime.utcnow,
        nullable=False
    )

    approved_by = db.Column(db.String(120))
    approved_on = db.Column(db.DateTime)
    posted_by = db.Column(db.String(120))
    posted_on = db.Column(db.DateTime)
    locked_on = db.Column(db.DateTime)
    reversed_on = db.Column(db.DateTime)
    posting_reference = db.Column(db.String(80))

    reversal_reason = db.Column(db.String(250))
    
    is_locked = db.Column(
    db.Boolean,
    default=False,
    nullable=False
)


class OpeningBalance(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    batch_id = db.Column(
        db.Integer,
        db.ForeignKey('opening_balance_batch.id'),
        nullable=False
    )

    member_id = db.Column(
        db.Integer,
        db.ForeignKey('member.id'),
        nullable=False
    )

    savings_balance = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    loan_principal = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    loan_interest = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    fine_balance = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    welfare_balance = db.Column(
        db.Numeric(14, 2),
        default=0,
        nullable=False
    )

    loan_due_date = db.Column(db.Date)
    loan_interest_rate = db.Column(db.Numeric(5, 2))

    remarks = db.Column(db.String(250))

    validation_status = db.Column(
        db.String(20),
        default='Valid',
        nullable=False
    )

    validation_message = db.Column(db.String(250))

    is_posted = db.Column(
        db.Boolean,
        default=False,
        nullable=False
    )
    posted_on = db.Column(db.DateTime)
    posted_by = db.Column(db.String(120))
    created_loan_id = db.Column(db.Integer)
    created_fine_id = db.Column(db.Integer)
    posted_contribution_id = db.Column(db.Integer)
    posted_loan_id = db.Column(db.Integer)
    posted_loan_interest_id = db.Column(db.Integer)
    posted_fine_id = db.Column(db.Integer)
    posted_welfare_id = db.Column(db.Integer)

    member = db.relationship(
        'Member',
        backref='opening_balances'
    )

    batch = db.relationship(
        'OpeningBalanceBatch',
        backref=db.backref(
            'balances',
            lazy=True,
            cascade='all, delete-orphan'
        )
    )

    __table_args__ = (
        db.UniqueConstraint(
            'batch_id',
            'member_id',
            name='uq_opening_balance_batch_member'
        ),
    )

def build_financial_year_closing_preview(financial_year):
    """
    Prepare the year-end closing summary without posting anything.
    """

    validation = run_financial_year_preclose_checks(financial_year)
    balances = ledger_balances(
    financial_year.start_date,
    financial_year.end_date
)

    income_accounts = []
    expense_accounts = []

    total_income = Decimal("0.00")
    total_expenses = Decimal("0.00")

    for item in balances:
        account = item["account"]
        balance = money(item.get("balance", Decimal("0.00")))

        if account.account_type == "Income" and balance != Decimal("0.00"):
            total_income += balance

            income_accounts.append({
                "code": account.code,
                "name": account.name,
                "balance": balance
            })

        elif account.account_type == "Expense" and balance != Decimal("0.00"):
            total_expenses += balance

            expense_accounts.append({
                "code": account.code,
                "name": account.name,
                "balance": balance
            })

    current_surplus = money(
        total_income - total_expenses
    )

    return {
        "validation": validation,
        "income_accounts": income_accounts,
        "expense_accounts": expense_accounts,
        "total_income": money(total_income),
        "total_expenses": money(total_expenses),
        "current_surplus": current_surplus,
        "is_ready": validation["is_ready"]
    }

def financial_year_balances(financial_year):
    return ledger_balances(
        financial_year.start_date,
        financial_year.end_date
    )

def current_financial_year():
    """
    Return the currently open financial year.

    If more than one year is open because of bad data,
    return the most recent one.
    """
    years = FinancialYear.query.all()

    print("All financial years:", years)

    return (
        FinancialYear.query
        .filter_by(status='Open')
        .order_by(FinancialYear.year.desc())
        .first()
    )
def run_financial_year_preclose_checks(financial_year):
    """
    Run year-end validation checks for the supplied financial year.

    Returns:
        {
            "checks": [...],
            "passed_count": 0,
            "failed_count": 0,
            "warning_count": 0,
            "is_ready": False
        }
    """

    checks = []

    if not financial_year:
        checks.append({
            "code": "FINANCIAL_YEAR",
            "name": "Financial Year",
            "status": "Failed",
            "message": "No open financial year was found.",
            "critical": True
        })

        return {
            "checks": checks,
            "passed_count": 0,
            "failed_count": 1,
            "warning_count": 0,
            "is_ready": False
        }

    # ---------------------------------------------------------
    # Check 1: Financial year status
    # ---------------------------------------------------------

    if financial_year.status == "Open" and not financial_year.is_locked:
        checks.append({
            "code": "YEAR_STATUS",
            "name": "Financial Year Status",
            "status": "Passed",
            "message": (
                f"Financial year {financial_year.year} is open "
                "and available for year-end processing."
            ),
            "critical": True
        })
    else:
        checks.append({
            "code": "YEAR_STATUS",
            "name": "Financial Year Status",
            "status": "Failed",
            "message": (
                f"Financial year {financial_year.year} is not open "
                "or is already locked."
            ),
            "critical": True
        })

    # ---------------------------------------------------------
    # Check 2: Trial Balance
    # ---------------------------------------------------------

    balances = ledger_balances(
    financial_year.start_date,
    financial_year.end_date
)

    total_debits = Decimal("0.00")
    total_credits = Decimal("0.00")

    for item in balances:
        balance = money(item.get("balance", Decimal("0.00")))
        account = item["account"]

        if account.normal_balance == "Debit":
            if balance >= 0:
                total_debits += balance
            else:
                total_credits += abs(balance)

        elif account.normal_balance == "Credit":
            if balance >= 0:
                total_credits += balance
            else:
                total_debits += abs(balance)

    trial_balance_difference = money(
        total_debits - total_credits
    )

    if trial_balance_difference == Decimal("0.00"):
        checks.append({
            "code": "TRIAL_BALANCE",
            "name": "Trial Balance",
            "status": "Passed",
            "message": "The Trial Balance is balanced.",
            "critical": True
        })
    else:
        checks.append({
            "code": "TRIAL_BALANCE",
            "name": "Trial Balance",
            "status": "Failed",
            "message": (
                "The Trial Balance has a difference of "
                f"{kwacha(trial_balance_difference)}."
            ),
            "critical": True
        })

    # ---------------------------------------------------------
    # Check 3: Balance Sheet
    # ---------------------------------------------------------

    balance_map = {
        item["account"].code: money(item["balance"])
        for item in balances
    }

    cash_on_hand = money(
        balance_map.get("1000", Decimal("0.00"))
    )

    bank_account = money(
        balance_map.get("1010", Decimal("0.00"))
    )

    mobile_money = money(
        balance_map.get("1020", Decimal("0.00"))
    )

    loans_receivable = money(
        balance_map.get("1100", Decimal("0.00"))
    )

    loan_interest_receivable = money(
        balance_map.get("1110", Decimal("0.00"))
    )

    fines_receivable = money(
        balance_map.get("1120", Decimal("0.00"))
    )

    member_savings = money(
        balance_map.get("2000", Decimal("0.00"))
    )

    welfare_fund = money(
        balance_map.get("2010", Decimal("0.00"))
    )

    accumulated_surplus = money(
        balance_map.get("3000", Decimal("0.00"))
    )

    opening_balance_equity = money(
        balance_map.get("3900", Decimal("0.00"))
    )

    income = Decimal("0.00")
    expenses = Decimal("0.00")

    for item in balances:
        account = item["account"]

        if account.account_type == "Income":
            income += money(item["balance"])

        elif account.account_type == "Expense":
            expenses += money(item["balance"])

    current_surplus = money(income - expenses)

    total_assets = money(
        cash_on_hand
        + bank_account
        + mobile_money
        + loans_receivable
        + loan_interest_receivable
        + fines_receivable
    )

    total_liabilities = money(
        member_savings
        + welfare_fund
    )

    total_equity = money(
        accumulated_surplus
        + opening_balance_equity
        + current_surplus
    )

    balance_sheet_difference = money(
        total_assets
        - total_liabilities
        - total_equity
    )

    if balance_sheet_difference == Decimal("0.00"):
        checks.append({
            "code": "BALANCE_SHEET",
            "name": "Balance Sheet",
            "status": "Passed",
            "message": "The Balance Sheet is balanced.",
            "critical": True
        })
    else:
        checks.append({
            "code": "BALANCE_SHEET",
            "name": "Balance Sheet",
            "status": "Failed",
            "message": (
                "The Balance Sheet has a difference of "
                f"{kwacha(balance_sheet_difference)}."
            ),
            "critical": True
        })

    # ---------------------------------------------------------
    # Check 4: Opening Balance batches
    # ---------------------------------------------------------

    pending_opening_batches = (
        OpeningBalanceBatch.query
        .filter(
            OpeningBalanceBatch.status.notin_(
                ["Posted", "Locked", "Reversed"]
            )
        )
        .count()
    )

    if pending_opening_batches == 0:
        checks.append({
            "code": "OPENING_BALANCES",
            "name": "Opening Balances",
            "status": "Passed",
            "message": "There are no incomplete opening-balance batches.",
            "critical": True
        })
    else:
        checks.append({
            "code": "OPENING_BALANCES",
            "name": "Opening Balances",
            "status": "Failed",
            "message": (
                f"{pending_opening_batches} opening-balance batch(es) "
                "have not been fully posted or resolved."
            ),
            "critical": True
        })

    # ---------------------------------------------------------
    # Check 5: Financial-year dates
    # ---------------------------------------------------------

    if financial_year.start_date <= financial_year.end_date:
        checks.append({
            "code": "YEAR_DATES",
            "name": "Financial Year Dates",
            "status": "Passed",
            "message": (
                f"The period runs from "
                f"{financial_year.start_date.strftime('%d %b %Y')} "
                f"to {financial_year.end_date.strftime('%d %b %Y')}."
            ),
            "critical": True
        })
    else:
        checks.append({
            "code": "YEAR_DATES",
            "name": "Financial Year Dates",
            "status": "Failed",
            "message": "The financial-year date range is invalid.",
            "critical": True
        })

    passed_count = sum(
        1 for check in checks
        if check["status"] == "Passed"
    )

    failed_count = sum(
        1 for check in checks
        if check["status"] == "Failed"
    )

    warning_count = sum(
        1 for check in checks
        if check["status"] == "Warning"
    )

    critical_failures = any(
        check["status"] == "Failed"
        and check.get("critical", False)
        for check in checks
    )

    return {
        "checks": checks,
        "passed_count": passed_count,
        "failed_count": failed_count,
        "warning_count": warning_count,
        "is_ready": not critical_failures
    }

def financial_year_for_date(transaction_date):
    """
    Return the financial year containing the supplied date.
    """
    if not transaction_date:
        return None

    return (
        FinancialYear.query
        .filter(
            FinancialYear.start_date <= transaction_date,
            FinancialYear.end_date >= transaction_date
        )
        .first()
    )

def is_financial_year_locked(transaction_date):
    """
    Return True when the financial year containing the date is locked.
    """
    financial_year = financial_year_for_date(transaction_date)

    if not financial_year:
        return False

    return bool(
        financial_year.is_locked
        or financial_year.status == 'Closed'
    )

def ensure_financial_year_columns():
    """
    Create and maintain the FinancialYear table.
    Safe to run every time the application starts.
    """

    inspector = db.inspect(db.engine)

    # Create table if it does not exist
    if 'financial_years' not in inspector.get_table_names():
        FinancialYear.__table__.create(db.engine)
        inspector = db.inspect(db.engine)

    existing_columns = {
        c['name']
        for c in inspector.get_columns('financial_years')
    }

    with db.engine.begin() as conn:

        if 'opened_by' not in existing_columns:
            conn.execute(text(
                "ALTER TABLE financial_years "
                "ADD COLUMN opened_by VARCHAR(120)"
            ))

        if 'closed_on' not in existing_columns:
            conn.execute(text(
                "ALTER TABLE financial_years "
                "ADD COLUMN closed_on TIMESTAMP"
            ))

        if 'closed_by' not in existing_columns:
            conn.execute(text(
                "ALTER TABLE financial_years "
                "ADD COLUMN closed_by VARCHAR(120)"
            ))

        if 'closing_reference' not in existing_columns:
            conn.execute(text(
                "ALTER TABLE financial_years "
                "ADD COLUMN closing_reference VARCHAR(50)"
            ))

        if 'is_locked' not in existing_columns:
            conn.execute(text(
                "ALTER TABLE financial_years "
                "ADD COLUMN is_locked BOOLEAN DEFAULT FALSE"
            ))

        if 'created_on' not in existing_columns:
            conn.execute(text(
                "ALTER TABLE financial_years "
                "ADD COLUMN created_on TIMESTAMP"
            ))

        if 'updated_on' not in existing_columns:
            conn.execute(text(
                "ALTER TABLE financial_years "
                "ADD COLUMN updated_on TIMESTAMP"
            ))

    # ------------------------------------------------------------------
    # Automatically create the current financial year if none exists
    # ------------------------------------------------------------------

    current_year = date.today().year

    existing_year = FinancialYear.query.filter_by(
        year=current_year
    ).first()

    if not existing_year:
        now = datetime.now(timezone.utc)

        financial_year = FinancialYear(
            year=current_year,
            start_date=date(current_year, 1, 1),
            end_date=date(current_year, 12, 31),
            status='Open',
            opened_on=now,
            opened_by='System',
            is_locked=False,
            created_on=now,
            updated_on=now
        )

        db.session.add(financial_year)
        db.session.commit()    

def opening_balance_decimal(value):
    """Safely convert imported values to Decimal."""
    if value is None:
        return Decimal("0.00")

    text = str(value).strip().replace(",", "")

    if text == "":
        return Decimal("0.00")

    try:
        amount = Decimal(text)

        if amount < 0:
            raise ValueError("Negative balances are not allowed.")

        return amount.quantize(Decimal("0.01"))

    except (InvalidOperation, ValueError):
        raise ValueError(f"Invalid amount: {value}")


def normalise_opening_balance_heading(value):
    """Convert spreadsheet headings into predictable names."""
    if value is None:
        return ""

    return (
        str(value)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )


def read_opening_balance_file(file_storage):
    """
    Read an uploaded XLSX or CSV opening-balance file.
    """

    filename = secure_filename(file_storage.filename or "")
    extension = os.path.splitext(filename)[1].lower()

    if extension not in {".xlsx", ".csv"}:
        raise ValueError(
            "Only Excel (.xlsx) and CSV (.csv) files are allowed."
        )

    rows = []

    # ---------------------------------------------------------
    # Excel file
    # ---------------------------------------------------------
    if extension == ".xlsx":

        file_storage.stream.seek(0)

        workbook = load_workbook(
            file_storage.stream,
            read_only=True,
            data_only=True
        )

        worksheet = workbook.active

        all_rows = list(
            worksheet.iter_rows(values_only=True)
        )

        if not all_rows:
            raise ValueError(
                "The uploaded Excel file is empty."
            )

        heading_row_index = None
        headings = None

        # Search the first 15 rows for the actual heading row.
        for index, values in enumerate(all_rows[:15]):

            normalised_values = [
                normalise_opening_balance_heading(value)
                for value in values
            ]

            if "member_no" in normalised_values:
                heading_row_index = index
                headings = normalised_values
                break

        if heading_row_index is None:
            raise ValueError(
                "Could not find the opening balance headings. "
                "Please use the downloaded Excel template."
            )

        for row_number, values in enumerate(
            all_rows[heading_row_index + 1:],
            start=heading_row_index + 2
        ):
            row = dict(zip(headings, values))

            if any(
                value is not None
                and str(value).strip() != ""
                for value in values
            ):
                row["_row_number"] = row_number
                rows.append(row)

        workbook.close()

    # ---------------------------------------------------------
    # CSV file
    # ---------------------------------------------------------
    elif extension == ".csv":

        file_storage.stream.seek(0)

        content = file_storage.stream.read().decode(
            "utf-8-sig"
        ).splitlines()

        reader = csv.DictReader(content)

        if not reader.fieldnames:
            raise ValueError(
                "The uploaded CSV file has no headings."
            )

        for row_number, raw_row in enumerate(reader, start=2):

            row = {
                normalise_opening_balance_heading(key): value
                for key, value in raw_row.items()
            }

            if any(
                value is not None
                and str(value).strip() != ""
                for value in row.values()
            ):
                row["_row_number"] = row_number
                rows.append(row)

    if not rows:
        raise ValueError(
            "The uploaded file contains no balance records."
        )

    return rows 

def generate_opening_balance_batch_no():
    year = date.today().year

    last_batch = OpeningBalanceBatch.query.filter(
        OpeningBalanceBatch.batch_no.like(f'OB-{year}-%')
    ).order_by(
        OpeningBalanceBatch.id.desc()
    ).first()

    next_number = 1

    if last_batch and last_batch.batch_no:
        try:
            next_number = int(
                last_batch.batch_no.split('-')[-1]
            ) + 1
        except (ValueError, IndexError):
            next_number = last_batch.id + 1

    return f'OB-{year}-{next_number:03d}'

def build_financial_year_closing_lines(financial_year):
    """
    Build the closing journal lines for the selected financial year.

    Income accounts are closed against accumulated surplus.
    Expense accounts are closed against accumulated surplus.
    """

    balances = ledger_balances(
        financial_year.start_date,
        financial_year.end_date
    )

    accumulated_surplus_account = Account.query.filter_by(
        code=ACCUMULATED_SURPLUS_ACCOUNT_CODE
    ).first()

    if not accumulated_surplus_account:
        raise ValueError(
            f"Accumulated Surplus account "
            f"{ACCUMULATED_SURPLUS_ACCOUNT_CODE} is missing "
            f"from the Chart of Accounts."
        )

    lines = []

    total_income = Decimal("0.00")
    total_expenses = Decimal("0.00")

    for item in balances:
        account = item["account"]
        balance = money(item["balance"])

        if balance == Decimal("0.00"):
            continue

        if account.account_type == "Income":
            total_income += balance

            if balance > 0:
                lines.append({
                    "account": account,
                    "debit": balance,
                    "credit": Decimal("0.00"),
                    "memo": (
                        f"Close {account.code} - {account.name} "
                        f"for FY{financial_year.year}"
                    )
                })
            else:
                lines.append({
                    "account": account,
                    "debit": Decimal("0.00"),
                    "credit": abs(balance),
                    "memo": (
                        f"Close abnormal balance on "
                        f"{account.code} - {account.name} "
                        f"for FY{financial_year.year}"
                    )
                })

        elif account.account_type == "Expense":
            total_expenses += balance

            if balance > 0:
                lines.append({
                    "account": account,
                    "debit": Decimal("0.00"),
                    "credit": balance,
                    "memo": (
                        f"Close {account.code} - {account.name} "
                        f"for FY{financial_year.year}"
                    )
                })
            else:
                lines.append({
                    "account": account,
                    "debit": abs(balance),
                    "credit": Decimal("0.00"),
                    "memo": (
                        f"Close abnormal balance on "
                        f"{account.code} - {account.name} "
                        f"for FY{financial_year.year}"
                    )
                })

    current_surplus = money(total_income - total_expenses)

    if current_surplus > Decimal("0.00"):
        lines.append({
            "account": accumulated_surplus_account,
            "debit": Decimal("0.00"),
            "credit": current_surplus,
            "memo": (
                f"Transfer FY{financial_year.year} surplus "
                f"to accumulated surplus"
            )
        })

    elif current_surplus < Decimal("0.00"):
        lines.append({
            "account": accumulated_surplus_account,
            "debit": abs(current_surplus),
            "credit": Decimal("0.00"),
            "memo": (
                f"Transfer FY{financial_year.year} deficit "
                f"to accumulated surplus"
            )
        })

    debit_total = money(sum(
        (money(line.get("debit", 0)) for line in lines),
        Decimal("0.00")
    ))

    credit_total = money(sum(
        (money(line.get("credit", 0)) for line in lines),
        Decimal("0.00")
    ))

    if debit_total != credit_total:
        raise ValueError(
            "The proposed financial-year closing journal is not balanced. "
            f"Debits: {debit_total}; Credits: {credit_total}."
        )

    return {
        "lines": lines,
        "total_income": money(total_income),
        "total_expenses": money(total_expenses),
        "current_surplus": current_surplus,
        "debit_total": debit_total,
        "credit_total": credit_total
    }

def generate_financial_year_closing_reference(financial_year):
    """
    Generate a unique closing reference such as FY2026-0001.
    """

    prefix = f"FY{financial_year.year}-"

    latest_entry = (
        JournalEntry.query
        .filter(JournalEntry.reference.like(f"{prefix}%"))
        .order_by(JournalEntry.id.desc())
        .first()
    )

    sequence = 1

    if latest_entry and latest_entry.reference:
        try:
            sequence = int(latest_entry.reference.split("-")[-1]) + 1
        except (ValueError, IndexError):
            sequence = 1

    reference = f"{prefix}{sequence:04d}"

    while JournalEntry.query.filter_by(reference=reference).first():
        sequence += 1
        reference = f"{prefix}{sequence:04d}"

    return reference

def ensure_opening_balance_columns():
    inspector = db.inspect(db.engine)
    table_names = inspector.get_table_names()

    column_definitions = {
        'contribution': {
            'source_type': "VARCHAR(30) DEFAULT 'Normal'",
            'source_id': 'INTEGER'
        },
        'loan': {
            'source_type': "VARCHAR(30) DEFAULT 'Normal'",
            'source_id': 'INTEGER'
        },
        'fine_penalty': {
            'source_type': "VARCHAR(30) DEFAULT 'Normal'",
            'source_id': 'INTEGER'
        },
        'welfare_contribution': {
            'source_type': "VARCHAR(30) DEFAULT 'Normal'",
            'source_id': 'INTEGER'
        },
        'loan_interest': {
            'source_type': "VARCHAR(30) DEFAULT 'Month End'",
            'source_id': 'INTEGER'
        }
    }

    for table_name, required_columns in column_definitions.items():
        if table_name not in table_names:
            continue

        existing_columns = {
            column['name']
            for column in inspector.get_columns(table_name)
        }

        for column_name, column_sql in required_columns.items():
            if column_name not in existing_columns:
                db.session.execute(
                    db.text(
                        f'ALTER TABLE {table_name} '
                        f'ADD COLUMN {column_name} {column_sql}'
                    )
                )

        # ---------------------------------------------------------
    # OpeningBalanceBatch table
    # ---------------------------------------------------------

    if 'opening_balance_batch' in table_names:

        existing_columns = {
            column['name']
            for column in inspector.get_columns('opening_balance_batch')
        }

        new_columns = {
            'approved_by': 'VARCHAR(120)',
            'posted_by': 'VARCHAR(120)',
            'locked_by': 'VARCHAR(120)',
            'posting_reference': 'VARCHAR(80)',
            'is_locked': 'BOOLEAN DEFAULT FALSE'
        }

        for column_name, column_sql in new_columns.items():
            if column_name not in existing_columns:
                db.session.execute(
                    db.text(
                        f'ALTER TABLE opening_balance_batch '
                        f'ADD COLUMN {column_name} {column_sql}'
                    )
                )
        # ---------------------------------------------------------
    # OpeningBalance table
    # ---------------------------------------------------------

    if 'opening_balance' in table_names:

        existing_columns = {
            column['name']
            for column in inspector.get_columns('opening_balance')
        }

        new_columns = {
            'is_posted': 'BOOLEAN DEFAULT FALSE',
            'posted_on': 'TIMESTAMP',
            'posted_by': 'VARCHAR(120)',
            'created_loan_id': 'INTEGER',
            'created_fine_id': 'INTEGER'
        }

        for column_name, column_sql in new_columns.items():
            if column_name not in existing_columns:
                db.session.execute(
                    db.text(
                        f'ALTER TABLE opening_balance '
                        f'ADD COLUMN {column_name} {column_sql}'
                    )
                )
        # ---------------------------------------------------------
    # Loan table
    # ---------------------------------------------------------

    if 'loan' in table_names:

        existing_columns = {
            column['name']
            for column in inspector.get_columns('loan')
        }

        if 'opening_interest' not in existing_columns:
            db.session.execute(
                db.text(
                    'ALTER TABLE loan '
                    'ADD COLUMN opening_interest NUMERIC(12,2) DEFAULT 0'
                )
            )

        
    db.session.commit()
   
def ensure_settings_columns():
    columns = {
        # Existing organisation and notification fields
        'logo_url': 'VARCHAR(500)',
        'organization_address': 'VARCHAR(250)',
        'organization_phone': 'VARCHAR(50)',
        'organization_email': 'VARCHAR(120)',
        'registration_number': 'VARCHAR(100)',
        'sms_provider': "VARCHAR(50) DEFAULT 'Manual'",
        'sms_api_key': 'VARCHAR(500)',
        'sms_sender_id': 'VARCHAR(100)',
        'whatsapp_enabled': 'BOOLEAN DEFAULT FALSE',
        'sms_username': 'VARCHAR(100)',

        # Branding and product identity
        'short_name': "VARCHAR(80) DEFAULT 'Village Banking'",
        'motto': (
            "VARCHAR(200) "
            "DEFAULT 'Empowering Community Finance'"
        ),
        'developer_name': (
            "VARCHAR(150) "
            "DEFAULT 'SL Consulting Limited'"
        ),
        'product_name': (
            "VARCHAR(150) "
            "DEFAULT 'SL Village Banking Pro'"
        ),
        'product_version': "VARCHAR(50) DEFAULT '1.0.0'",
        'logo': 'VARCHAR(255)',
        'report_logo': 'VARCHAR(255)',
        'watermark_logo': 'VARCHAR(255)',
        'favicon': 'VARCHAR(255)',

        # Theme and visual customization
        'theme_name': (
            "VARCHAR(80) DEFAULT 'Professional Blue'"
        ),
        'primary_color': (
            "VARCHAR(20) DEFAULT '#0D6EFD'"
        ),
        'secondary_color': (
            "VARCHAR(20) DEFAULT '#198754'"
        ),
        'accent_color': (
            "VARCHAR(20) DEFAULT '#6F42C1'"
        ),
        'sidebar_color': (
            "VARCHAR(20) DEFAULT '#17324D'"
        ),
        'sidebar_text_color': (
            "VARCHAR(20) DEFAULT '#FFFFFF'"
        ),
        'navbar_color': (
            "VARCHAR(20) DEFAULT '#FFFFFF'"
        ),
        'button_color': (
            "VARCHAR(20) DEFAULT '#0D6EFD'"
        ),
        'background_color': (
            "VARCHAR(20) DEFAULT '#F4F7FB'"
        ),
        'card_color': (
            "VARCHAR(20) DEFAULT '#FFFFFF'"
        ),
        'table_header_color': (
            "VARCHAR(20) DEFAULT '#EAF2F8'"
        ),
        'success_color': (
            "VARCHAR(20) DEFAULT '#198754'"
        ),
        'warning_color': (
            "VARCHAR(20) DEFAULT '#FFC107'"
        ),
        'danger_color': (
            "VARCHAR(20) DEFAULT '#DC3545'"
        ),

        # Contact information
        'phone': 'VARCHAR(50)',
        'email': 'VARCHAR(120)',
        'website': 'VARCHAR(150)',
        'postal_address': 'TEXT',
        'physical_address': 'TEXT',

        # Currency settings
        'currency': "VARCHAR(20) DEFAULT 'ZMW'",
        'currency_symbol': "VARCHAR(10) DEFAULT 'K'",
        'decimal_places': 'INTEGER DEFAULT 2',

        # Financial defaults
        'default_interest_rate': (
            'NUMERIC(10,2) DEFAULT 15.00'
        ),
        'default_loan_term': 'INTEGER DEFAULT 6',
        'penalty_rate': (
            'NUMERIC(10,2) DEFAULT 0.00'
        ),
        'share_out_month': (
            "VARCHAR(20) DEFAULT 'December'"
        ),

        # Meeting and dashboard settings
        'committee_meeting_frequency': (
            "VARCHAR(50) DEFAULT 'Monthly'"
        ),
        'member_meeting_frequency': (
            "VARCHAR(50) DEFAULT 'Quarterly'"
        ),
        'enable_ai_advisor': (
            'BOOLEAN DEFAULT TRUE'
        ),
        'enable_notifications': (
            'BOOLEAN DEFAULT TRUE'
        ),
        'enable_dashboard_charts': (
            'BOOLEAN DEFAULT TRUE'
        ),

        # Audit fields
        'created_at': 'TIMESTAMP',
        'updated_on': 'TIMESTAMP'
    }

    inspector = inspect(db.engine)

    existing_columns = {
        column['name']
        for column in inspector.get_columns(
            'system_setting'
        )
    }

    for column, definition in columns.items():
        if column in existing_columns:
            continue

        try:
            db.session.execute(
                db.text(
                    f'ALTER TABLE system_setting '
                    f'ADD COLUMN {column} {definition}'
                )
            )

            db.session.commit()

            print(
                f'Added system_setting.{column}'
            )

        except Exception as exc:
            db.session.rollback()

            print(
                f'Could not add '
                f'system_setting.{column}: {exc}'
            )


def money(value):
    return Decimal(value or 0).quantize(Decimal('0.01'))

def normalize_import_header(value):
    value = str(value or '').strip().lower()

    value = re.sub(
        r'[^a-z0-9]+',
        '_',
        value
    )

    return value.strip('_')

@app.route('/uploaded-logo/<path:filename>')
def uploaded_logo(filename):
    return send_from_directory(
        LOGO_UPLOAD_FOLDER,
        filename
    )



@app.template_filter('kwacha')
def kwacha(value):
    if value is None:
        value = Decimal('0.00')

    settings = get_system_settings()
    amount = money(value)

    return f"{settings.currency_symbol} {amount:,.2f}"
def user_can(permission):
    user = session.get('user') or {}
    role = user.get('role')
    return permission in ROLE_PERMISSIONS.get(role, [])

@app.context_processor
def inject_globals():

    setting = SystemSetting.query.first()

    client_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    return dict(
        payment_methods=PAYMENT_METHODS,
        interest_percent=int(INTEREST_RATE * 100),
        client_name=client_name,
        setting=setting,
        producer_name=PRODUCER_NAME,
        current_year=date.today().year,
        current_user=session.get('user'),
        user_can=user_can,
        roles=ROLES,
        fine_categories=FINE_CATEGORIES,
        fine_statuses=FINE_STATUSES,
        welfare_categories=WELFARE_CATEGORIES,
        welfare_statuses=WELFARE_STATUSES,
        attendance_statuses=ATTENDANCE_STATUSES,
        notification_channels=NOTIFICATION_CHANNELS,
        notification_types=NOTIFICATION_TYPES,
        account_types=ACCOUNT_TYPES
    )
def get_settings():
    setting = SystemSetting.query.first()

    if not setting:
        setting = SystemSetting()
        db.session.add(setting)
        db.session.commit()

    return setting

def get_system_settings():
    setting = SystemSetting.query.first()

    if not setting:
        setting = SystemSetting()
        db.session.add(setting)
        db.session.commit()

    return setting

@app.context_processor
def inject_system_settings():
    return {
        'settings': get_system_settings()
    }

def log_audit(action, entity=None, entity_id=None, details=None):
    user = session.get('user') or {}
    try:
        entry = AuditLog(
            user_id=user.get('id'),
            username=user.get('username', 'system'),
            full_name=user.get('full_name', 'System'),
            role=user.get('role', 'System'),
            action=action,
            entity=entity,
            entity_id=str(entity_id) if entity_id is not None else None,
            details=details,
            ip_address=request.headers.get('X-Forwarded-For', request.remote_addr) if request else None,
        )
        db.session.add(entry)
        db.session.commit()
    except Exception:
        db.session.rollback()


def seed_chart_of_accounts():
    default_accounts = [
        ('1000', 'Cash on Hand', 'Asset', 'Debit'),
        ('1010', 'Bank Account', 'Asset', 'Debit'),
        ('1020', 'Mobile Money', 'Asset', 'Debit'),
        ('1100', 'Loans Receivable', 'Asset', 'Debit'),
        ('2000', 'Member Savings / Contributions', 'Liability', 'Credit'),
        ('2010', 'Welfare Fund Payable', 'Liability', 'Credit'),
        ('3000', 'Accumulated Surplus', 'Equity', 'Credit'),
        ('4000', 'Loan Interest Income', 'Income', 'Credit'),
        ('4010', 'Fines and Penalties Income', 'Income', 'Credit'),
        ('4020', 'Other Income', 'Income', 'Credit'),
        ('5000', 'Administrative Expenses', 'Expense', 'Debit'),
        ('5010', 'Bank Charges', 'Expense', 'Debit'),
        ('5020', 'Transport and Meetings', 'Expense', 'Debit'),
        ('5030', 'Welfare Support Expense', 'Expense', 'Debit'),
        ('5040', 'Share-Out / Dividend Distribution', 'Expense', 'Debit'),
        ('2100', 'Member Loan Payable', 'Liability', 'Credit'),
    ]
    for code, name, account_type, normal_balance in default_accounts:
        if not Account.query.filter_by(code=code).first():
            db.session.add(Account(code=code, name=name, account_type=account_type, normal_balance=normal_balance))
    db.session.commit()


def account_by_code(code):
    return Account.query.filter_by(code=code).first()


def cash_account_for_method(method):
    if method == 'Bank Transfer':
        return account_by_code('1010')
    if method == 'Mobile Money':
        return account_by_code('1020')
    return account_by_code('1000')


def post_journal(
    entry_date=None,
    description='',
    reference='',
    source_type='',
    source_id=None,
    lines=None,
    debit_account_code=None,
    credit_account_code=None,
    amount=None,
    commit=True
):
    if source_type and source_id and JournalEntry.query.filter_by(
        source_type=source_type,
        source_id=str(source_id)
    ).first():
        return None

    if lines is None:
        debit_account = Account.query.filter_by(code=debit_account_code).first()
        credit_account = Account.query.filter_by(code=credit_account_code).first()

        if not debit_account or not credit_account:
            raise Exception(
                f'Chart of Accounts missing: {debit_account_code} or {credit_account_code}'
            )

        lines = [
            {'account': debit_account, 'debit': money(amount)},
            {'account': credit_account, 'credit': money(amount)}
        ]

    debit_total = money(
        sum((money(line.get('debit', 0)) for line in lines), Decimal('0.00'))
    )

    credit_total = money(
        sum((money(line.get('credit', 0)) for line in lines), Decimal('0.00'))
    )

    if debit_total != credit_total:
        raise ValueError('Journal entry is not balanced.')

    entry = JournalEntry(
        entry_date=entry_date or date.today(),
        description=description,
        reference=reference,
        source_type=source_type,
        source_id=str(source_id) if source_id is not None else None,
        posted_by=(session.get('user') or {}).get('username')
    )

    db.session.add(entry)
    db.session.flush()

    for line in lines:
        db.session.add(JournalLine(
            journal_entry_id=entry.id,
            account_id=line['account'].id,
            debit=money(line.get('debit', 0)),
            credit=money(line.get('credit', 0)),
            memo=line.get('memo')
        ))

    if commit:
        db.session.commit()
    else:
        db.session.flush()

    return entry

def sync_operational_transactions_to_gl():
    seed_chart_of_accounts()
    created = 0
    member_savings = account_by_code('2000')
    loans_receivable = account_by_code('1100')
    interest_income = account_by_code('4000')
    fines_income = account_by_code('4010')
    welfare_payable = account_by_code('2010')
    welfare_expense = account_by_code('5030')
    # Contributions: Dr payment account / Cr member savings
    for c in Contribution.query.all():
        if not JournalEntry.query.filter_by(source_type='Contribution', source_id=str(c.id)).first():
            post_journal(c.paid_on, f'Contribution from {c.member.full_name} for {c.month}', c.reference, 'Contribution', c.id, [
                {'account': cash_account_for_method(c.method), 'debit': c.amount},
                {'account': member_savings, 'credit': c.amount},
            ])
            created += 1
    # Disbursed loans: Dr loans receivable / Cr cash on hand
    for l in Loan.query.filter(
    Loan.status.in_(['Disbursed', 'Partially Paid', 'Paid'])
        ).all():
        if not JournalEntry.query.filter_by(source_type='Loan', source_id=str(l.id)).first():
            post_journal(l.issued_on, f'Loan issued to {l.member.full_name}', f'Loan-{l.id}', 'Loan', l.id, [
                {'account': loans_receivable, 'debit': l.principal},
                {'account': account_by_code('1000'), 'credit': l.principal},
            ])
            created += 1
    # Repayments: Dr payment account / Cr loans receivable and interest income estimate
    for r in Repayment.query.all():
        if not JournalEntry.query.filter_by(source_type='Repayment', source_id=str(r.id)).first():
            interest_part = money(min(money(r.amount), money(r.loan.interest_amount)))
            principal_part = money(money(r.amount) - interest_part)
            lines = [{'account': cash_account_for_method(r.method), 'debit': r.amount}]
            if principal_part > 0:
                lines.append({'account': loans_receivable, 'credit': principal_part})
            if interest_part > 0:
                lines.append({'account': interest_income, 'credit': interest_part})
            post_journal(r.paid_on, f'Loan repayment from {r.loan.member.full_name}', r.reference, 'Repayment', r.id, lines)
            created += 1
    # Fine payments: Dr payment account / Cr fines income
    for fp in FinePayment.query.all():
        if not JournalEntry.query.filter_by(source_type='FinePayment', source_id=str(fp.id)).first():
            post_journal(fp.paid_on, f'Fine payment from {fp.fine.member.full_name}', fp.reference, 'FinePayment', fp.id, [
                {'account': cash_account_for_method(fp.method), 'debit': fp.amount},
                {'account': fines_income, 'credit': fp.amount},
            ])
            created += 1
    # Welfare contributions: Dr payment account / Cr welfare fund payable
    if 'WelfareContribution' in globals():
        for wc in WelfareContribution.query.all():
            if not JournalEntry.query.filter_by(source_type='WelfareContribution', source_id=str(wc.id)).first():
                post_journal(wc.paid_on, f'Welfare contribution from {wc.member.full_name}', wc.reference, 'WelfareContribution', wc.id, [
                    {'account': cash_account_for_method(wc.method), 'debit': wc.amount},
                    {'account': welfare_payable, 'credit': wc.amount},
                ])
                created += 1
    # Welfare paid claims: Dr welfare support expense / Cr cash on hand
    if 'WelfareClaim' in globals():
        for claim in WelfareClaim.query.filter_by(status='Paid').all():
            if not JournalEntry.query.filter_by(source_type='WelfareClaim', source_id=str(claim.id)).first():
                post_journal(claim.paid_on or date.today(), f'Welfare payment to {claim.member.full_name}', f'Welfare-{claim.id}', 'WelfareClaim', claim.id, [
                    {'account': welfare_expense, 'debit': claim.amount},
                    {'account': account_by_code('1000'), 'credit': claim.amount},
                ])
                created += 1
    return created


def ledger_balances(start_date=None, end_date=None):
    query = JournalLine.query.join(JournalEntry).join(Account)
    if start_date:
        query = query.filter(JournalEntry.entry_date >= start_date)
    if end_date:
        query = query.filter(JournalEntry.entry_date <= end_date)
    balances = []
    for account in Account.query.order_by(Account.code).all():
        lines = JournalLine.query.join(JournalEntry).filter(JournalLine.account_id == account.id)
        if start_date:
            lines = lines.filter(JournalEntry.entry_date >= start_date)
        if end_date:
            lines = lines.filter(JournalEntry.entry_date <= end_date)
        debit = money(sum((line.debit for line in lines.all()), Decimal('0.00')))
        credit = money(sum((line.credit for line in lines.all()), Decimal('0.00')))
        balance = money(debit - credit) if account.normal_balance == 'Debit' else money(credit - debit)
        balances.append({'account': account, 'debit': debit, 'credit': credit, 'balance': balance})
    return balances

def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get('user'):
            return redirect(url_for('login', next=request.path))
        return view(*args, **kwargs)
    return wrapped

def role_required(permission):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not session.get('user'):
                return redirect(url_for('login', next=request.path))
            if not user_can(permission):
                flash('You do not have permission to access that section.', 'error')
                return redirect(url_for('dashboard'))
            return view(*args, **kwargs)
        return wrapped
    return decorator


def database_file_path():
    # Flask stores sqlite:///village_banking.db inside the instance folder.
    return os.path.join(app.instance_path, 'village_banking.db')

def backups_folder():
    folder = os.path.join(app.instance_path, 'backups')
    os.makedirs(folder, exist_ok=True)
    return folder

def make_backup_filename(prefix='backup'):
    return f"{prefix}_{datetime.now().strftime('%Y_%m_%d_%H%M%S')}.db"

def prune_old_backups(keep=BACKUP_RETENTION):
    folder = backups_folder()
    files = sorted(Path(folder).glob('*.db'), key=lambda x: x.stat().st_mtime, reverse=True)
    for old in files[keep:]:
        try:
            old.unlink()
        except OSError:
            pass

def create_database_backup(prefix='backup', notes=None):
    os.makedirs(app.instance_path, exist_ok=True)
    src = database_file_path()
    if not os.path.exists(src):
        raise FileNotFoundError('Database file was not found. Start the application once before creating backups.')
    filename = make_backup_filename(prefix)
    dst = os.path.join(backups_folder(), filename)
    db.session.remove()
    db.engine.dispose()
    shutil.copy2(src, dst)
    size = os.path.getsize(dst)
    user = session.get('user') or {}
    record = BackupRecord(filename=filename, file_size=size, backup_type='Manual', created_by=user.get('full_name') or user.get('username') or 'System', notes=notes)
    db.session.add(record)
    db.session.commit()
    prune_old_backups()
    return record

def post_to_cash_book(
    entry_date,
    entry_type,
    category,
    amount,
    description='',
    method='',
    reference='',
    source_type='',
    source_id=None
    ):
    user = session.get('user') or {}

    entry = CashBookEntry(
        entry_date=entry_date,
        entry_type=entry_type,
        category=category,
        description=description,
        amount=money(amount),
        method=method,
        reference=reference,
        source_type=source_type,
        source_id=source_id,
        created_by=user.get('username')
    )

    db.session.add(entry)

def cash_account(method):
    if method == 'Bank Transfer':
        return '1010'

    if method == 'Mobile Money':
        return '1020'

    return '1000'

def get_next_member_number():
    members = Member.query.with_entities(
        Member.member_no
    ).all()

    highest_number = 0

    for (member_no,) in members:
        if not member_no:
            continue

        match = re.fullmatch(
            r'M(\d+)',
            member_no.strip(),
            re.IGNORECASE
        )

        if match:
            number = int(match.group(1))
            highest_number = max(
                highest_number,
                number
            )

    return f'M{highest_number + 1:03d}'

@app.route('/system-settings', methods=['GET', 'POST'])
@login_required
@role_required('settings')
def system_settings():
    settings = get_system_settings()

    if request.method == 'POST':
        settings.organisation_name = request.form.get('organisation_name')
        settings.short_name = request.form.get('short_name')
        settings.motto = request.form.get('motto')
        settings.developer_name = request.form.get('developer_name')
        settings.product_name = request.form.get('product_name')
        settings.product_version = request.form.get('product_version')

        settings.phone = request.form.get('phone')
        settings.email = request.form.get('email')
        settings.website = request.form.get('website')
        settings.postal_address = request.form.get('postal_address')
        settings.physical_address = request.form.get('physical_address')
        settings.registration_number = request.form.get('registration_number')

        settings.currency = request.form.get('currency')
        settings.currency_symbol = request.form.get('currency_symbol')
        settings.decimal_places = int(request.form.get('decimal_places') or 2)

        settings.default_interest_rate = money(request.form.get('default_interest_rate') or 0)
        settings.default_loan_term = int(request.form.get('default_loan_term') or 0)
        settings.penalty_rate = money(request.form.get('penalty_rate') or 0)
        settings.share_out_month = request.form.get('share_out_month')

        settings.committee_meeting_frequency = request.form.get('committee_meeting_frequency')
        settings.member_meeting_frequency = request.form.get('member_meeting_frequency')

        settings.enable_ai_advisor = bool(request.form.get('enable_ai_advisor'))
        settings.enable_notifications = bool(request.form.get('enable_notifications'))
        settings.enable_dashboard_charts = bool(request.form.get('enable_dashboard_charts'))

        db.session.commit()

        log_audit(
            'UPDATE_SYSTEM_SETTINGS',
            'SystemSetting',
            settings.id,
            'System branding and settings were updated'
        )

        flash('System settings updated successfully.')
        return redirect(url_for('system_settings'))

    return render_template(
        'system_settings.html',
        settings=settings
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(
            username=request.form.get('username', '').strip(),
            active=True
        ).first()

        if user and check_password_hash(user.password_hash, request.form.get('password', '')):
            session['user'] = {
                'id': user.id,
                'username': user.username,
                'full_name': user.full_name,
                'role': user.role
            }

            log_audit('LOGIN_SUCCESS', 'User', user.id, f'{user.full_name} logged in')
            flash('Welcome back. You are logged in securely.')

            return redirect(request.args.get('next') or url_for('home'))

        flash('Invalid username or password.', 'error')

    return render_template('login.html')

@app.route('/logout')
def logout():
    log_audit('LOGOUT', 'User', session.get('user', {}).get('id'), 'User logged out')
    session.clear()
    flash('You have been logged out.')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def home():
    return render_template('home.html')

@app.route('/settings', methods=['GET', 'POST'])
@login_required
@role_required('settings')
def settings():
    setting = SystemSetting.query.first()

    if not setting:
        setting = SystemSetting()
        db.session.add(setting)
        db.session.commit()

    if request.method == 'POST':
        # Organisation settings
        setting.organisation_name = (
            request.form.get('organization_name')
            or 'Your Organisation Name'
        )

        setting.short_name = (
            request.form.get('short_name')
            or 'Village Banking'
        )

        setting.motto = (
            request.form.get('motto')
            or 'Empowering Community Finance'
        )

        setting.registration_number = request.form.get(
            'registration_number'
        )

        setting.organization_address = request.form.get(
            'organization_address'
        )

        setting.organization_phone = request.form.get(
            'organization_phone'
        )

        setting.organization_email = request.form.get(
            'organization_email'
        )

        setting.website = request.form.get('website')
        setting.postal_address = request.form.get('postal_address')
        setting.physical_address = request.form.get('physical_address')

        # Product identity
        setting.product_name = (
            request.form.get('product_name')
            or 'SL Village Banking Pro'
        )

        setting.product_version = (
            request.form.get('product_version')
            or '1.0.0'
        )

        setting.developer_name = (
            request.form.get('developer_name')
            or 'SL Consulting Limited'
        )

        # Currency configuration
        setting.currency = (
            request.form.get('currency')
            or 'ZMW'
        )

        setting.currency_symbol = (
            request.form.get('currency_symbol')
            or 'K'
        )

        setting.decimal_places = int(
            request.form.get('decimal_places') or 2
        )

        # Financial settings
        setting.contribution_amount = money(
            request.form.get('contribution_amount') or 0
        )

        setting.savings_interest_rate = money(
            request.form.get('savings_interest_rate') or 0
        )

        setting.loan_interest_rate = money(
            request.form.get('loan_interest_rate') or 0
        )

        setting.welfare_contribution_amount = money(
            request.form.get('welfare_contribution_amount') or 0
        )

        setting.default_interest_rate = money(
            request.form.get('default_interest_rate') or 0
        )

        setting.default_loan_term = int(
            request.form.get('default_loan_term') or 6
        )

        setting.penalty_rate = money(
            request.form.get('penalty_rate') or 0
        )

        setting.share_out_month = (
            request.form.get('share_out_month')
            or 'December'
        )

        # Meeting defaults
        setting.committee_meeting_frequency = (
            request.form.get('committee_meeting_frequency')
            or 'Monthly'
        )

        setting.member_meeting_frequency = (
            request.form.get('member_meeting_frequency')
            or 'Quarterly'
        )

        # Dashboard options
        setting.enable_ai_advisor = (
            'enable_ai_advisor' in request.form
        )

        setting.enable_notifications = (
            'enable_notifications' in request.form
        )

        setting.enable_dashboard_charts = (
            'enable_dashboard_charts' in request.form
        )

        # Communication settings
        setting.sms_provider = (
            request.form.get('sms_provider')
            or 'Manual'
        )

        setting.sms_api_key = request.form.get('sms_api_key')
        setting.sms_sender_id = request.form.get('sms_sender_id')
        setting.sms_username = request.form.get('sms_username')

        setting.whatsapp_enabled = (
            'whatsapp_enabled' in request.form
        )

        # Theme configuration
        setting.theme_name = (
            request.form.get('theme_name')
            or 'Professional Blue'
        )

        setting.primary_color = (
            request.form.get('primary_color')
            or '#0D6EFD'
        )

        setting.secondary_color = (
            request.form.get('secondary_color')
            or '#198754'
        )

        setting.accent_color = (
            request.form.get('accent_color')
            or '#6F42C1'
        )

        setting.sidebar_color = (
            request.form.get('sidebar_color')
            or '#17324D'
        )

        setting.sidebar_text_color = (
            request.form.get('sidebar_text_color')
            or '#FFFFFF'
        )

        setting.navbar_color = (
            request.form.get('navbar_color')
            or '#FFFFFF'
        )

        setting.button_color = (
            request.form.get('button_color')
            or '#0D6EFD'
        )

        setting.background_color = (
            request.form.get('background_color')
            or '#F4F7FB'
        )

        setting.card_color = (
            request.form.get('card_color')
            or '#FFFFFF'
        )

        setting.table_header_color = (
            request.form.get('table_header_color')
            or '#EAF2F8'
        )

        setting.success_color = (
            request.form.get('success_color')
            or '#198754'
        )

        setting.warning_color = (
            request.form.get('warning_color')
            or '#FFC107'
        )

        setting.danger_color = (
            request.form.get('danger_color')
            or '#DC3545'
        )

        setting.updated_on = datetime.utcnow()

        db.session.commit()

        log_audit(
            'UPDATE_SETTINGS',
            'SystemSetting',
            setting.id,
            (
                'Organisation branding, theme and system '
                'settings updated'
            )
        )

        flash(
            'Branding and system settings updated successfully.',
            'success'
        )

        return redirect(url_for('settings'))

    return render_template(
        'settings.html',
        setting=setting
    )

@app.route(
    '/logo-management',
    methods=['GET', 'POST']
)
@login_required
@role_required('settings')
def logo_management():
    setting = SystemSetting.query.first()

    if not setting:
        setting = SystemSetting()
        db.session.add(setting)
        db.session.commit()

    if request.method == 'POST':
        action = request.form.get('action')

        try:
            if action == 'upload_system_logo':
                filename = save_logo_file(
                    request.files.get('system_logo'),
                    'system_logo'
                )

                if filename:
                    setting.logo = filename

            elif action == 'upload_report_logo':
                filename = save_logo_file(
                    request.files.get('report_logo'),
                    'report_logo'
                )

                if filename:
                    setting.report_logo = filename

            elif action == 'upload_watermark_logo':
                filename = save_logo_file(
                    request.files.get(
                        'watermark_logo'
                    ),
                    'watermark_logo'
                )

                if filename:
                    setting.watermark_logo = filename

            elif action == 'upload_favicon':
                filename = save_logo_file(
                    request.files.get('favicon'),
                    'favicon'
                )

                if filename:
                    setting.favicon = filename

            elif action == 'remove_system_logo':
                delete_logo_file(setting.logo)
                setting.logo = None

            elif action == 'remove_report_logo':
                delete_logo_file(
                    setting.report_logo
                )
                setting.report_logo = None

            elif action == 'remove_watermark_logo':
                delete_logo_file(
                    setting.watermark_logo
                )
                setting.watermark_logo = None

            elif action == 'remove_favicon':
                delete_logo_file(setting.favicon)
                setting.favicon = None

            else:
                flash(
                    'Unknown logo action.',
                    'error'
                )

                return redirect(
                    url_for('logo_management')
                )

            db.session.commit()

            log_audit(
                'UPDATE_LOGOS',
                'SystemSetting',
                setting.id,
                f'Logo action completed: {action}'
            )

            flash(
                'Logo settings updated successfully.',
                'success'
            )

        except ValueError as exc:
            flash(str(exc), 'error')

        except Exception as exc:
            db.session.rollback()

            flash(
                f'Logo update failed: {exc}',
                'error'
            )

        return redirect(
            url_for('logo_management')
        )

    return render_template(
        'logo_management.html',
        setting=setting
    )


@app.route('/dashboard')
@login_required
@role_required('dashboard')
def dashboard():
     return redirect(url_for('executive_dashboard'))
   
@app.route('/executive-dashboard')
@login_required
@role_required('dashboard')
def executive_dashboard():
    balances = ledger_balances()

    balance_map = {
        b['account'].code: b['balance']
        for b in balances
    }
    # Get the next scheduled meeting
    next_meeting = Meeting.query.filter(
        Meeting.meeting_date >= date.today()
    ).order_by(
        Meeting.meeting_date.asc()
    ).first()

    cash_on_hand = money(balance_map.get('1000', Decimal('0.00')))
    bank_account = money(balance_map.get('1010', Decimal('0.00')))
    mobile_money = money(balance_map.get('1020', Decimal('0.00')))
    total_cash = money(cash_on_hand + bank_account + mobile_money)

    loans_receivable = money(balance_map.get('1100', Decimal('0.00')))
    member_savings = money(balance_map.get('2000', Decimal('0.00')))
    welfare_fund = money(balance_map.get('2010', Decimal('0.00')))

    income = Decimal('0.00')
    expenses = Decimal('0.00')

    for b in balances:
        account = b['account']

        if account.account_type == 'Income':
            income += b['balance']

        elif account.account_type == 'Expense':
            expenses += b['balance']

    total_income = money(income)
    total_expenses = money(expenses)
    current_surplus = money(total_income - total_expenses)

    active_members = Member.query.count()

    active_loans = Loan.query.filter(
        Loan.status.in_(['Disbursed', 'Partially Paid'])
    ).count()

    pending_welfare_claims = WelfareClaim.query.filter_by(
        status='Pending'
    ).count()

    recent_journals = JournalEntry.query.order_by(
        JournalEntry.entry_date.desc(),
        JournalEntry.id.desc()
    ).limit(8).all()
        # -----------------------------
    # Financial Performance Chart
    # -----------------------------

    if db.engine.dialect.name == 'sqlite':
        contribution_month = func.strftime('%Y-%m', Contribution.paid_on)
    else:
        contribution_month = func.to_char(Contribution.paid_on, 'YYYY-MM')

    monthly_contributions = (
        db.session.query(
            contribution_month,
            func.coalesce(func.sum(Contribution.amount), 0)
        )
        .group_by(contribution_month)
        .order_by(contribution_month)
        .all()
    )

    contribution_labels = [m for m, _ in monthly_contributions]
    contribution_values = [float(v) for _, v in monthly_contributions]


    if db.engine.dialect.name == 'sqlite':
        member_month = func.strftime('%Y-%m', Member.created_at)
    else:
        member_month = func.to_char(Member.created_at, 'YYYY-MM')

    member_growth = (
        db.session.query(
            member_month,
            func.count(Member.id)
        )
        .group_by(member_month)
        .order_by(member_month)
        .all()
    )

    member_labels = [m for m, _ in member_growth]
    member_values = [v for _, v in member_growth]

    # -----------------------------
    # Loan Portfolio
    # -----------------------------

    current_loans = Loan.query.filter_by(status='Disbursed').count()

    paid_loans = Loan.query.filter_by(status='Paid').count()

    overdue_loans = 0

    for loan in Loan.query.filter_by(status='Disbursed').all():
        if loan.balance > 0 and loan.due_on and loan.due_on < date.today():
            overdue_loans += 1
    today = date.today()
    today_activity = []

    today_contributions = Contribution.query.filter(
        Contribution.paid_on == today
    ).order_by(
        Contribution.id.desc()
    ).limit(5).all()

    for c in today_contributions:
        today_activity.append({
            "icon": "fa-coins",
            "level": "good",
            "title": f"{c.member.full_name} saved {money(c.amount)}",
            "time": "Today",
            "message": "Savings recorded."
        })

    today_members = Member.query.filter(
    Member.created_at == today
    ).order_by(
        Member.id.desc()
    ).limit(5).all()

    for m in today_members:
        today_activity.append({
            "icon": "fa-user-plus",
            "level": "blue",
            "title": f"{m.full_name} joined the group",
            "time": "Today",
            "message": "New member registered."
        })

    today_loans = Loan.query.filter(
        Loan.approved_on == today
    ).order_by(
        Loan.id.desc()
    ).limit(5).all()

    for loan in today_loans:
        today_activity.append({
            "icon": "fa-hand-holding-dollar",
            "level": "purple",
            "title": f"Loan approved for {loan.member.full_name}",
            "time": "Today",
            "message": f"Amount approved: {money(loan.principal)}"
        })

    today_activity = today_activity[:8]

    dashboard_service = DashboardService(
        total_cash=total_cash,
        overdue_loans=overdue_loans,
        pending_welfare_claims=pending_welfare_claims,
        next_meeting=next_meeting,
        today_activity=today_activity,
        active_loans=active_loans,
        active_members=active_members,
        welfare_fund=welfare_fund
    )

    dashboard_data = dashboard_service.build()

   
    today_checklist = []

    today_checklist.append({
        "done": False,
        "icon": "fa-coins",
        "title": "Record today's savings",
        "message": "Enter member savings received today.",
        "link": url_for("contributions")
    })

    if overdue_loans > 0:
        today_checklist.append({
            "done": False,
            "icon": "fa-hand-holding-dollar",
            "title": "Follow up overdue loans",
            "message": f"{overdue_loans} loan(s) need attention.",
            "link": url_for("loans")
        })
    else:
        today_checklist.append({
            "done": True,
            "icon": "fa-circle-check",
            "title": "Loan follow-up checked",
            "message": "No overdue loan follow-up required.",
            "link": url_for("loans")
        })

    if pending_welfare_claims > 0:
        today_checklist.append({
            "done": False,
            "icon": "fa-heart",
            "title": "Review welfare claims",
            "message": f"{pending_welfare_claims} claim(s) waiting.",
            "link": url_for("welfare")
        })
    else:
        today_checklist.append({
            "done": True,
            "icon": "fa-heart-circle-check",
            "title": "Emergency fund checked",
            "message": "No pending welfare claims.",
            "link": url_for("welfare")
        })

    today_checklist.append({
        "done": False,
        "icon": "fa-building-columns",
        "title": "Check bank position",
        "message": "Review bank reconciliation when needed.",
        "link": url_for("bank_reconciliation")
    })

    # -----------------------------
    # Top Savers
    # -----------------------------

    top_savers = (
        db.session.query(
            Member.id,
            Member.member_no,
            Member.full_name,
            func.coalesce(func.sum(Contribution.amount), 0).label('total_saved')
        )
        .join(Contribution, Contribution.member_id == Member.id)
        .group_by(Member.id, Member.member_no, Member.full_name)
        .order_by(func.coalesce(func.sum(Contribution.amount), 0).desc())
        .limit(10)
        .all()
    )
    # -----------------------------
    # Members Requiring Follow-up
    # -----------------------------

    current_month = date.today().strftime("%Y-%m")

    saved_this_month_ids = [
        r[0]
        for r in db.session.query(Contribution.member_id)
            .filter(Contribution.month == current_month)
            .distinct()
            .all()
    ]

    members_requiring_followup = []

    for member in Member.query.order_by(Member.full_name).all():

        if member.id in saved_this_month_ids:
            continue

        last_contribution = (
            Contribution.query
            .filter_by(member_id=member.id)
            .order_by(
                Contribution.paid_on.desc(),
                Contribution.id.desc()
            )
            .first()
        )

        members_requiring_followup.append({
            "id": member.id,
            "member_no": member.member_no,
            "full_name": member.full_name,
            "last_contribution": (
                last_contribution.month
                if last_contribution
                else "Never"
            )
        })
    # -----------------------------
    # Executive Financial Ratios
    # -----------------------------

    average_savings_per_member = money(0)
    if active_members > 0:
        average_savings_per_member = money(member_savings / active_members)

    average_loan_size = money(0)
    if active_loans > 0:
        average_loan_size = money(loans_receivable / active_loans)

    liquidity_ratio = Decimal("0.00")
    if member_savings > 0:
        liquidity_ratio = money(total_cash / member_savings)

    total_loans_count = Loan.query.count()
    loan_recovery_rate = Decimal("100.00")

    if total_loans_count > 0:
        loan_recovery_rate = money((paid_loans / total_loans_count) * 100)

    portfolio_at_risk = Decimal("0.00")

    if current_loans > 0:
        portfolio_at_risk = money((overdue_loans / current_loans) * 100)

    financial_ratios = [
        {
            "title": "Average Savings",
            "value": kwacha(average_savings_per_member),
            "status": "Per Member",
            "level": "good",
            "icon": "fa-piggy-bank"
        },
        {
            "title": "Average Loan Size",
            "value": kwacha(average_loan_size),
            "status": "Active Loans",
            "level": "blue",
            "icon": "fa-hand-holding-dollar"
        },
        {
            "title": "Liquidity Ratio",
            "value": liquidity_ratio,
            "status": "Cash vs Savings",
            "level": "good" if liquidity_ratio >= 0.5 else "watch",
            "icon": "fa-droplet"
        },
        {
            "title": "Loan Recovery",
            "value": f"{loan_recovery_rate}%",
            "status": "Excellent" if loan_recovery_rate >= 95 else "Monitor",
            "level": "good" if loan_recovery_rate >= 95 else "watch",
            "icon": "fa-circle-check"
        },
        {
            "title": "Portfolio at Risk",
            "value": f"{portfolio_at_risk}%",
            "status": "Healthy" if portfolio_at_risk < 5 else "Needs Attention",
            "level": "good" if portfolio_at_risk < 5 else "danger",
            "icon": "fa-triangle-exclamation"
        }
    ]

    members_requiring_followup = members_requiring_followup[:10]

    # -----------------------------
    # Executive AI Advisor
    # -----------------------------

    executive_ai_advice = []

    if total_cash > 0:
        executive_ai_advice.append({
            "level": "good",
            "icon": "fa-wallet",
            "title": "Cash Position",
            "message": f"The group currently has {kwacha(total_cash)} available in cash, bank and mobile money."
        })
    else:
        executive_ai_advice.append({
            "level": "danger",
            "icon": "fa-wallet",
            "title": "Cash Position Warning",
            "message": "The group has no available cash recorded. The committee should review the cash book."
        })

    if len(members_requiring_followup) > 0:
        executive_ai_advice.append({
            "level": "watch",
            "icon": "fa-user-clock",
            "title": "Contribution Follow-up Required",
            "message": f"{len(members_requiring_followup)} member(s) shown on this dashboard have not contributed this month. Follow-up is recommended before month-end."
        })
    else:
        executive_ai_advice.append({
            "level": "good",
            "icon": "fa-circle-check",
            "title": "Contribution Performance",
            "message": "All listed members have contributed this month. This is a strong participation signal."
        })

    if portfolio_at_risk < 5:
        executive_ai_advice.append({
            "level": "good",
            "icon": "fa-shield-halved",
            "title": "Loan Portfolio Health",
            "message": f"Portfolio at Risk is {portfolio_at_risk}%, which is within a healthy range."
        })
    else:
        executive_ai_advice.append({
            "level": "danger",
            "icon": "fa-triangle-exclamation",
            "title": "Loan Portfolio Risk",
            "message": f"Portfolio at Risk is {portfolio_at_risk}%. The committee should urgently follow up overdue loans."
        })

    if liquidity_ratio >= 0.5:
        executive_ai_advice.append({
            "level": "good",
            "icon": "fa-droplet",
            "title": "Liquidity Position",
            "message": f"The liquidity ratio is {liquidity_ratio}, meaning the group has reasonable liquid funds compared with member savings."
        })
    else:
        executive_ai_advice.append({
            "level": "watch",
            "icon": "fa-droplet",
            "title": "Liquidity Watch",
            "message": f"The liquidity ratio is {liquidity_ratio}. The committee should monitor cash availability carefully."
        })

    if pending_welfare_claims > 0:
        executive_ai_advice.append({
            "level": "watch",
            "icon": "fa-heart",
            "title": "Welfare Claims",
            "message": f"There are {pending_welfare_claims} pending welfare claim(s) requiring committee review."
        })  

    return render_template(
        'executive_dashboard.html',
        cash_on_hand=cash_on_hand,
        bank_account=bank_account,
        mobile_money=mobile_money,
        total_cash=total_cash,
        loans_receivable=loans_receivable,
        member_savings=member_savings,
        welfare_fund=welfare_fund,
        total_income=total_income,
        total_expenses=total_expenses,
        current_surplus=current_surplus,
        active_members=active_members,
        active_loans=active_loans,
        pending_welfare_claims=pending_welfare_claims,
        recent_journals=recent_journals,
        contribution_labels=contribution_labels,
        contribution_values=contribution_values,
        member_labels=member_labels,
        member_values=member_values,
        current_loans=current_loans,
        paid_loans=paid_loans,
        overdue_loans=overdue_loans,
        today=date.today(),
        **dashboard_data,
        today_checklist=today_checklist,
        top_savers=top_savers,
        members_requiring_followup=members_requiring_followup,
        financial_ratios=financial_ratios,
        executive_ai_advice=executive_ai_advice,
    )

@app.route('/members')
@login_required
@role_required('members')
def members():
    q = request.args.get('q', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 25

    query = Member.query

    if q:
        search_term = f'%{q}%'

        query = query.filter(
            db.or_(
                Member.full_name.ilike(search_term),
                Member.member_no.ilike(search_term),
                Member.phone.ilike(search_term),
                Member.national_id.ilike(search_term),
                Member.group_name.ilike(search_term),
                Member.member_type.ilike(search_term),
                Member.committee_position.ilike(search_term),
            )
        )

    pagination = query.order_by(
        Member.member_no.asc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    total_members = Member.query.count()

    active_members = Member.query.filter_by(
        status='Active'
    ).count()

    inactive_members = max(
        total_members - active_members,
        0
    )

    committee_members = Member.query.filter(
        Member.committee_position.isnot(None),
        Member.committee_position != ''
    ).count()

    current_month = date.today().strftime('%Y-%m')

    if db.engine.dialect.name == 'sqlite':
        member_month_expression = db.func.strftime(
            '%Y-%m',
            Member.created_at
        )
    else:
        member_month_expression = db.func.to_char(
            Member.created_at,
            'YYYY-MM'
        )

    new_members_this_month = Member.query.filter(
        member_month_expression == current_month
    ).count()

    search_result_count = pagination.total

    return render_template(
        'members.html',
        members=pagination.items,
        pagination=pagination,
        q=q,
        search_result_count=search_result_count,
        total_members=total_members,
        active_members=active_members,
        inactive_members=inactive_members,
        committee_members=committee_members,
        new_members_this_month=new_members_this_month
    )
@app.route('/member/<int:member_id>/savings')
@login_required
@role_required('members')
def member_savings_statement(member_id):
    member = Member.query.get_or_404(member_id)

    contributions = Contribution.query.filter_by(
        member_id=member.id
    ).order_by(
        Contribution.paid_on.desc(),
        Contribution.id.desc()
    ).all()

    savings_interest = SavingsInterest.query.filter_by(
        member_id=member.id
    ).order_by(
        SavingsInterest.month.desc(),
        SavingsInterest.id.desc()
    ).all()

    distributions = Distribution.query.filter_by(
        member_id=member.id
    ).order_by(
        Distribution.paid_on.desc(),
        Distribution.id.desc()
    ).all()

    total_contributions = money(sum((c.amount for c in contributions), Decimal('0.00')))
    total_interest = money(sum((s.interest_amount for s in savings_interest), Decimal('0.00')))
    total_distributions = money(sum((d.amount for d in distributions), Decimal('0.00')))

    savings_balance = money(total_contributions + total_interest - total_distributions)

    loan_balance = money(
    sum(
        (l.balance for l in Loan.query.filter_by(member_id=member.id).all()),
        Decimal('0.00')
        )
    )

    fine_balance = money(
        sum(
            (
                f.balance
                for f in FinePenalty.query.filter_by(member_id=member.id).all()
                if f.status != 'Waived'
            ),
            Decimal('0.00')
        )
    )

    last_contribution = contributions[0] if contributions else None

    if savings_balance > 0 and loan_balance <= 0 and fine_balance <= 0:
        member_statement_status = "Good Standing"
        member_statement_message = "This member has savings and no outstanding loans or fines."
    elif loan_balance > 0 or fine_balance > 0:
        member_statement_status = "Needs Review"
        member_statement_message = "This member has outstanding balances requiring attention."
    else:
        member_statement_status = "No Savings Activity"
        member_statement_message = "No active savings balance is currently shown."

    return render_template(
        'member_savings_statement.html',
        member=member,
        contributions=contributions,
        savings_interest=savings_interest,
        distributions=distributions,
        total_contributions=total_contributions,
        total_interest=total_interest,
        total_distributions=total_distributions,
        savings_balance=savings_balance,
        loan_balance=loan_balance,
        fine_balance=fine_balance,
        last_contribution=last_contribution,
        member_statement_status=member_statement_status,
        member_statement_message=member_statement_message
    )

@app.route('/member/<int:member_id>')
@login_required
@role_required('members')
def member_profile(member_id):
    member = Member.query.get_or_404(member_id)

    total_contributions = money(
        db.session.query(db.func.coalesce(db.func.sum(Contribution.amount), 0))
        .filter(Contribution.member_id == member.id)
        .scalar()
    )

    savings_interest = money(
        db.session.query(db.func.coalesce(db.func.sum(SavingsInterest.interest_amount), 0))
        .filter(SavingsInterest.member_id == member.id)
        .scalar()
    )

    loans = Loan.query.filter_by(member_id=member.id).all()
    loan_balance = money(sum((l.balance for l in loans), Decimal('0.00')))

    fines = FinePenalty.query.filter_by(member_id=member.id).all()
    fine_balance = money(sum((f.balance for f in fines if f.status != 'Waived'), Decimal('0.00')))

    welfare_contributions = money(
        db.session.query(db.func.coalesce(db.func.sum(WelfareContribution.amount), 0))
        .filter(WelfareContribution.member_id == member.id)
        .scalar()
    )

    welfare_paid = money(
        db.session.query(db.func.coalesce(db.func.sum(WelfareClaim.amount_approved), 0))
        .filter(WelfareClaim.member_id == member.id)
        .filter(WelfareClaim.status == 'Paid')
        .scalar()
    )

    recent_contributions = Contribution.query.filter_by(member_id=member.id).order_by(Contribution.paid_on.desc()).limit(5).all()
    recent_loans = Loan.query.filter_by(member_id=member.id).order_by(Loan.issued_on.desc()).limit(5).all()
    recent_fines = FinePenalty.query.filter_by(member_id=member.id).order_by(FinePenalty.id.desc()).limit(5).all()
    member_health = "Excellent"
    member_health_level = "good"
    member_health_score = 100
    member_health_message = "This member is in good standing."

    if fine_balance > 0:
        member_health_score -= 20
        member_health = "Needs Follow-up"
        member_health_level = "watch"
        member_health_message = "This member has unpaid fines or penalties."

    if loan_balance > 0:
        member_health_score -= 15
        member_health = "Active Borrower"
        member_health_level = "watch"
        member_health_message = "This member currently has an outstanding loan balance."

    if member.status != "Active":
        member_health_score -= 30
        member_health = "Inactive"
        member_health_level = "danger"
        member_health_message = "This member is currently not active."

    if member_health_score >= 85:
        member_health = "Excellent Member"
        member_health_level = "good"
    elif member_health_score >= 60:
        member_health = "Needs Follow-up"
        member_health_level = "watch"
    else:
        member_health = "High Attention"
        member_health_level = "danger"

    intelligence = MemberIntelligenceService(
        member=member,
        total_contributions=total_contributions,
        loan_balance=loan_balance,
        fine_balance=fine_balance,
        recent_contributions=recent_contributions,
        recent_loans=recent_loans,
        recent_fines=recent_fines
    )

    intelligence_data = intelligence.build()

    return render_template(
        "member_profile.html",
        member=member,
        total_contributions=total_contributions,
        savings_interest=savings_interest,
        loan_balance=loan_balance,
        fine_balance=fine_balance,
        welfare_contributions=welfare_contributions,
        welfare_paid=welfare_paid,
        recent_contributions=recent_contributions,
        recent_loans=recent_loans,
        recent_fines=recent_fines,
        **intelligence_data
    )

@app.route('/members/new', methods=['GET', 'POST'])
@login_required
@role_required('members')
def member_new():
    next_member_no = get_next_member_number()

    if request.method == 'POST':
        member_no = (
            request.form.get('member_no')
            or next_member_no
        ).strip().upper()

        existing = Member.query.filter(
            db.func.upper(Member.member_no) == member_no
        ).first()

        if existing:
            flash(
                (
                    f'Member number {member_no} '
                    'already exists in the database.'
                ),
                'error'
            )

            return render_template(
                'member_form.html',
                member=None,
                next_member_no=next_member_no,
                form_data=request.form
            )

        full_name = (
            request.form.get('full_name')
            or ''
        ).strip()

        if not full_name:
            flash(
                'Member name is required.',
                'error'
            )

            return render_template(
                'member_form.html',
                member=None,
                next_member_no=next_member_no,
                form_data=request.form
            )

        member = Member(
            member_no=member_no,
            full_name=full_name,
            phone=(
                request.form.get('phone')
                or ''
            ).strip() or None,
            national_id=(
                request.form.get('national_id')
                or ''
            ).strip() or None,
            group_name=(
                request.form.get('group_name')
                or ''
            ).strip() or None,
            member_type=(
                request.form.get('member_type')
                or 'Ordinary Member'
            ),
            committee_position=(
                request.form.get(
                    'committee_position'
                )
                or None
            )
        )

        db.session.add(member)
        db.session.commit()

        log_audit(
            'CREATE_MEMBER',
            'Member',
            member.id,
            (
                f'{member.member_no} - '
                f'{member.full_name}'
            )
        )

        flash(
            'Member added successfully.',
            'success'
        )

        return redirect(
            url_for('members')
        )

    return render_template(
        'member_form.html',
        member=None,
        next_member_no=next_member_no,
        form_data={}
    )

@app.route('/members/<int:member_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('members')
def member_edit(member_id):
    member = Member.query.get_or_404(member_id)

    if request.method == 'POST':
        member.member_no = request.form['member_no'].strip()
        member.full_name = request.form['full_name'].strip()
        member.phone = request.form.get('phone')
        member.national_id = request.form.get('national_id')
        member.group_name = request.form.get('group_name')
        member.status = request.form.get('status') or 'Active'

        member.member_type = request.form.get('member_type') or 'Ordinary Member'
        member.committee_position = request.form.get('committee_position') or None
        
        db.session.commit()
        log_audit('UPDATE_MEMBER', 'Member', member.id, f'{member.member_no} - {member.full_name}')
        flash('Member updated successfully.')
        return redirect(url_for('members'))

    return render_template(
    'member_form.html',
    member=member,
    form_data={}
        )


@app.route('/members/<int:member_id>/toggle', methods=['POST'])
@login_required
@role_required('members')
def member_toggle(member_id):
    member = Member.query.get_or_404(member_id)

    member.status = 'Inactive' if member.status == 'Active' else 'Active'

    db.session.commit()
    log_audit('TOGGLE_MEMBER_STATUS', 'Member', member.id, f'{member.member_no} status changed to {member.status}')
    flash(f'Member status changed to {member.status}.')
    return redirect(url_for('members'))

@app.route('/members/import', methods=['GET', 'POST'])
@login_required
@role_required('members')
def members_import():
    results = {
        'processed': 0,
        'created': 0,
        'updated': 0,
        'skipped': 0,
        'errors': []
    }

    if request.method == 'POST':
        uploaded_file = request.files.get('file')

        if not uploaded_file or not uploaded_file.filename:
            flash(
                'Please select a CSV or Excel file to import.',
                'error'
            )

            return render_template(
                'members_import.html',
                results=results
            )

        filename = uploaded_file.filename.lower()

        try:
            rows = []

            if filename.endswith('.xlsx'):
                workbook = load_workbook(
                    uploaded_file,
                    data_only=True
                )

                sheet = workbook.active

                headers = [
                    normalize_import_header(cell.value)
                    for cell in sheet[1]
                ]

                for line_no, values in enumerate(
                    sheet.iter_rows(
                        min_row=2,
                        values_only=True
                    ),
                    start=2
                ):
                    row = dict(zip(headers, values))
                    rows.append((line_no, row))

            elif filename.endswith('.csv'):
                raw = uploaded_file.stream.read()

                try:
                    decoded = raw.decode('utf-8-sig')
                except UnicodeDecodeError:
                    decoded = raw.decode('latin-1')

                stream = io.StringIO(
                    decoded,
                    newline=None
                )

                reader = csv.DictReader(stream)

                original_headers = reader.fieldnames or []

                normalized_headers = [
                    normalize_import_header(header)
                    for header in original_headers
                ]

                for line_no, values in enumerate(
                    reader,
                    start=2
                ):
                    normalized_row = {}

                    for original, normalized in zip(
                        original_headers,
                        normalized_headers
                    ):
                        normalized_row[normalized] = (
                            values.get(original)
                        )

                    rows.append(
                        (line_no, normalized_row)
                    )

            else:
                flash(
                    (
                        'Unsupported file type. '
                        'Please upload an .xlsx or .csv file.'
                    ),
                    'error'
                )

                return render_template(
                    'members_import.html',
                    results=results
                )

            if not rows:
                flash(
                    'The selected file contains no member records.',
                    'warning'
                )

                return render_template(
                    'members_import.html',
                    results=results
                )

            for line_no, row in rows:
                results['processed'] += 1

                member_no = str(
                    row.get('member_no') or ''
                ).strip().upper()

                full_name = str(
                    row.get('full_name') or ''
                ).strip()

                if not full_name:
                    results['skipped'] += 1

                    results['errors'].append(
                        (
                            f'Line {line_no}: '
                            'full_name is required.'
                        )
                    )

                    continue

                if not member_no:
                    member_no = get_next_member_number()

                phone = str(
                    row.get('phone') or ''
                ).strip() or None

                national_id = str(
                    row.get('national_id') or ''
                ).strip() or None

                group_name = str(
                    row.get('group_name')
                    or row.get('group')
                    or ''
                ).strip() or None

                status = str(
                    row.get('status') or 'Active'
                ).strip() or 'Active'

                member_type = str(
                    row.get('member_type')
                    or 'Ordinary Member'
                ).strip() or 'Ordinary Member'

                committee_position = str(
                    row.get('committee_position')
                    or ''
                ).strip() or None

                member = Member.query.filter(
                    db.func.upper(
                        Member.member_no
                    ) == member_no
                ).first()

                if member:
                    member.full_name = full_name
                    member.phone = phone
                    member.national_id = national_id
                    member.group_name = group_name
                    member.status = status
                    member.member_type = member_type
                    member.committee_position = (
                        committee_position
                    )

                    results['updated'] += 1

                else:
                    if national_id:
                        duplicate_national_id = (
                            Member.query.filter_by(
                                national_id=national_id
                            ).first()
                        )

                        if duplicate_national_id:
                            results['skipped'] += 1

                            results['errors'].append(
                                (
                                    f'Line {line_no}: '
                                    f'National ID {national_id} '
                                    'already belongs to '
                                    f'{duplicate_national_id.full_name}.'
                                )
                            )

                            continue

                    member = Member(
                        member_no=member_no,
                        full_name=full_name,
                        phone=phone,
                        national_id=national_id,
                        group_name=group_name,
                        status=status,
                        member_type=member_type,
                        committee_position=(
                            committee_position
                        )
                    )

                    db.session.add(member)
                    db.session.flush()

                    results['created'] += 1

            db.session.commit()

            log_audit(
                'IMPORT_MEMBERS',
                'Member',
                None,
                (
                    'Bulk member import completed. '
                    f'Processed: {results["processed"]}; '
                    f'Created: {results["created"]}; '
                    f'Updated: {results["updated"]}; '
                    f'Skipped: {results["skipped"]}'
                )
            )

            flash(
                (
                    'Import completed. '
                    f'Created {results["created"]}, '
                    f'updated {results["updated"]}, '
                    f'skipped {results["skipped"]}.'
                ),
                'success'
            )

        except Exception as exc:
            db.session.rollback()

            flash(
                f'Import failed: {exc}',
                'error'
            )

        return render_template(
            'members_import.html',
            results=results
        )

    return render_template(
        'members_import.html',
        results=results
    )

@app.route('/contributions', methods=['GET', 'POST'])
@login_required
@role_required('contributions')
def contributions():

    if request.method == 'POST':
        c = Contribution(
            member_id=int(request.form['member_id']),
            month=request.form['month'],
            amount=money(request.form['amount']),
            method=request.form['method'],
            reference=request.form.get('reference'),
            paid_on=parse_date(request.form.get('paid_on'))
        )

        db.session.add(c)
        db.session.flush()

        post_to_cash_book(
            entry_date=c.paid_on,
            entry_type='In',
            category='Member Contribution',
            amount=c.amount,
            description=f'{c.member.member_no} - {c.member.full_name}',
            method=c.method,
            reference=c.reference,
            source_type='Contribution',
            source_id=c.id
        )

        post_journal(
            entry_date=c.paid_on,
            description=f'Member contribution - {c.member.member_no} - {c.member.full_name}',
            debit_account_code=cash_account(c.method),
            credit_account_code='2000',
            amount=c.amount,
            source_type='Contribution',
            source_id=c.id
        )

        db.session.commit()

        log_audit(
            'RECORD_CONTRIBUTION',
            'Contribution',
            c.id,
            f'{c.member.full_name} paid {kwacha(c.amount)} for {c.month} via {c.method}'
        )

        flash('Contribution recorded.')
        return redirect(url_for('contributions'))

    page = request.args.get('page', 1, type=int)
    per_page = 25

    search = request.args.get('search', '').strip()
    month_filter = request.args.get('month', '').strip()
    method_filter = request.args.get('method', '').strip()

    query = Contribution.query.join(Member)

    if search:
            query = query.filter(
                db.or_(
                    Member.full_name.ilike(f'%{search}%'),
                    Member.member_no.ilike(f'%{search}%'),
                    Contribution.reference.ilike(f'%{search}%')
                )
            )

    if month_filter:
            query = query.filter(Contribution.month == month_filter)

    if method_filter:
            query = query.filter(Contribution.method == method_filter)

    pagination = query.order_by(
            Contribution.paid_on.desc(),
            Contribution.id.desc()
        ).paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )
    today = date.today()
    current_month = today.strftime('%Y-%m')

    total_savings = money(
                db.session.query(
                    db.func.coalesce(db.func.sum(Contribution.amount), 0)
                ).scalar()
            )

    this_month_savings = money(
        db.session.query(
            db.func.coalesce(db.func.sum(Contribution.amount), 0)
        )
        .filter(Contribution.month == current_month)
        .scalar()
    )

    saved_this_month = db.session.query(
        Contribution.member_id
    ).filter(
        Contribution.month == current_month
    ).distinct().count()

    total_members = Member.query.count()
    missing_this_month = max(total_members - saved_this_month, 0)

    return render_template(
        'contributions.html',
        contributions=pagination.items,
        pagination=pagination,
        members=Member.query.order_by(Member.full_name).all(),
        total_savings=total_savings,
        this_month_savings=this_month_savings,
        saved_this_month=saved_this_month,
        missing_this_month=missing_this_month,
        search=search,
        month_filter=month_filter,
        method_filter=method_filter,
    )
@app.route('/contributions/passbook/<int:member_id>')
@login_required
@role_required('contributions')
def contribution_passbook(member_id):
    member = Member.query.get_or_404(member_id)

    contributions = Contribution.query.filter_by(
        member_id=member.id
    ).order_by(
        Contribution.paid_on.asc(),
        Contribution.id.asc()
    ).all()

    running_balance = money(0)
    passbook_rows = []

    for c in contributions:
        running_balance += money(c.amount)
        passbook_rows.append({
            'date': c.paid_on,
            'month': c.month,
            'description': f'Savings contribution for {c.month}',
            'method': c.method,
            'reference': c.reference,
            'deposit': c.amount,
            'balance': running_balance
        })

    return render_template(
        'contribution_passbook.html',
        member=member,
        passbook_rows=passbook_rows,
        running_balance=running_balance
    )

@app.route('/contributions/passbook/<int:member_id>/pdf')
@login_required
@role_required('contributions')
def contribution_passbook_pdf(member_id):
    system_setting = get_system_settings()

    member = Member.query.get_or_404(member_id)

    contributions = Contribution.query.filter_by(
        member_id=member.id
    ).order_by(
        Contribution.paid_on.asc(),
        Contribution.id.asc()
    ).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(system_setting.organisation_name.upper(), styles["Title"]))
    elements.append(
    Paragraph(
        f"Running on {system_setting.product_name} | Developed by {system_setting.developer_name}",
        styles["Normal"]
    )
)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("MEMBER SAVINGS STATEMENT", styles["Heading2"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"<b>Member No:</b> {member.member_no}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Member Name:</b> {member.full_name}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    data = [["Date", "Month", "Description", "Method", "Reference", "Deposit", "Balance"]]

    running_balance = money(0)

    for c in contributions:
        running_balance += money(c.amount)
        data.append([
            str(c.paid_on),
            c.month,
            f"Savings contribution for {c.month}",
            c.method,
            c.reference or "-",
            kwacha(c.amount),
            kwacha(running_balance)
        ])

    data.append(["", "", "", "", "Current Balance", "", kwacha(running_balance)])

    table = Table(data, repeatRows=1)

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f2f2f2")),
        ("FONTNAME", (4, -1), (-1, -1), "Helvetica-Bold"),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Generated on: {date.today()}", styles["Normal"]))

    doc.build(elements)

    buffer.seek(0)

    filename = f"savings_statement_{member.member_no}.pdf"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )

# ------------------------------------------------------------
# LOAN PROCESSING WORKFLOW
#
# Applied    -> Review
# Reviewed   -> Approve
# Approved   -> Disburse
# Disbursed  -> Repayment
# Paid       -> Closed
#
# Rejection may occur before disbursement.
#
# Route Responsibilities:
# loan_review()    : Applied   -> Reviewed
# loan_approve()   : Reviewed  -> Approved
# loan_disburse()  : Approved  -> Disbursed
# repayments()     : Disbursed -> Partially Paid/Paid
# ------------------------------------------------------------
@app.route('/loans', methods=['GET','POST'])
@login_required
@role_required('loans')
def loans():
    if request.method == 'POST':
        issued = parse_date(request.form.get('issued_on'))
        due = parse_date(request.form.get('due_on')) or issued + timedelta(days=30)

        l = Loan(
            member_id=int(request.form['member_id']),
            principal=money(request.form['principal']),
            due_on=due,
            issued_on=issued,
            purpose=request.form.get('purpose'),
            status='Applied'
        )

        db.session.add(l)
        db.session.commit()

        log_audit(
            'LOAN_APPLICATION_CREATED',
            'Loan',
            l.id,
            f'{l.member.full_name} applied for {kwacha(l.principal)}; due {l.due_on}'
        )

        flash('Loan application created. It must be reviewed, approved, then disbursed before repayment can be recorded.')
        return redirect(url_for('loans'))

    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = Loan.query.order_by(
        Loan.issued_on.desc(),
        Loan.id.desc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    # -------------------------------------------------
# Loan Intelligence Centre Summary
# -------------------------------------------------

    total_loans = Loan.query.count()

    applied_loans = Loan.query.filter_by(status='Applied').count()

    approved_loans = Loan.query.filter_by(status='Approved').count()

    disbursed_loans = Loan.query.filter_by(status='Disbursed').count()

    paid_loans = Loan.query.filter_by(status='Paid').count()

    overdue_loans = 0

    portfolio_balance = Decimal("0.00")

    for loan in Loan.query.filter_by(status="Disbursed").all():

        portfolio_balance += loan.balance

        if loan.balance > 0 and loan.due_on and loan.due_on < date.today():
            overdue_loans += 1

    portfolio_balance = money(portfolio_balance)

    active_loans = Loan.query.filter(
            Loan.status.in_(["Disbursed", "Partially Paid"])
        ).all()
    
    loan_intelligence = LoanIntelligenceService(
            total_loans=total_loans,
            applied_loans=applied_loans,
            approved_loans=approved_loans,
            disbursed_loans=disbursed_loans,
            paid_loans=paid_loans,
            overdue_loans=overdue_loans,
            portfolio_balance=portfolio_balance,
            active_loans=active_loans
        )
    
    loan_data = loan_intelligence.build()

    return render_template(
        "loans.html",
        loans=pagination.items,
        pagination=pagination,
        members=Member.query.order_by(Member.full_name).all(),
        settings=get_settings(),

        total_loans=total_loans,
        applied_loans=applied_loans,
        approved_loans=approved_loans,
        disbursed_loans=disbursed_loans,
        paid_loans=paid_loans,
        overdue_loans=overdue_loans,
        portfolio_balance=portfolio_balance,

        **loan_data
    )

@app.route('/loans/<int:loan_id>/statement.pdf')
@login_required
@role_required('loans')
def loan_statement_pdf(loan_id):
    loan = Loan.query.get_or_404(loan_id)

    repayments = Repayment.query.filter_by(
        loan_id=loan.id
    ).order_by(
        Repayment.paid_on,
        Repayment.id
    ).all()

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm
    )

    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'LoanStatementTitle',
        parent=styles['Title'],
        fontSize=16,
        leading=20
    )

    normal = styles['Normal']

    story.append(Paragraph('Loan Statement', title_style))
    story.append(Spacer(1, 8))

    story.append(Paragraph(f'<b>Member:</b> {loan.member.full_name}', normal))
    story.append(Paragraph(f'<b>Member No:</b> {loan.member.member_no}', normal))
    story.append(Paragraph(f'<b>Loan No:</b> {loan.loan_no or f"LN{loan.id:04d}"}', normal))
    story.append(Paragraph(f'<b>Status:</b> {loan.status}', normal))
    story.append(Spacer(1, 8))

    summary_data = [
        [
            Paragraph('<b>Principal</b>', normal),
            kwacha(loan.principal),
            Paragraph('<b>Interest</b>', normal),
            kwacha(loan.interest_amount)
        ],
        [
            Paragraph('<b>Total Due</b>', normal),
            kwacha(loan.total_due),
            Paragraph('<b>Total Paid</b>', normal),
            kwacha(loan.total_paid)
        ],
        [
            Paragraph('<b>Balance</b>', normal),
            kwacha(loan.balance),
            Paragraph('<b>Status</b>', normal),
            loan.status or '-'
        ],
        [
            Paragraph('<b>Issued On</b>', normal),
            str(loan.issued_on or '-'),
            Paragraph('<b>Due On</b>', normal),
            str(loan.due_on or '-')
        ],
        [
            Paragraph('<b>Purpose</b>', normal),
            Paragraph(loan.purpose or '-', normal),
            Paragraph('<b>Loan No</b>', normal),
            loan.loan_no or f'LN{loan.id:04d}'
        ],
        [
            Paragraph('<b>Disbursement<br/>Method</b>', normal),
            loan.disbursement_method or '-',
            Paragraph('<b>Reference</b>', normal),
            loan.disbursement_reference or '-'
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[38 * mm, 42 * mm, 34 * mm, 46 * mm]
    )

    summary_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),

        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f4f8')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f0f4f8')),

        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),

        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('WORDWRAP', (0, 0), (-1, -1), True),
    ]))

    story.append(summary_table)
    story.append(Spacer(1, 12))

    workflow_data = [
        ['Reviewed By', loan.reviewed_by or '-', 'Reviewed On', str(loan.reviewed_on or '-')],
        ['Approved By', loan.approved_by or '-', 'Approved On', str(loan.approved_on or '-')],
        ['Disbursed By', loan.disbursed_by or '-', 'Disbursed On', str(loan.disbursed_on or '-')],
        ['Rejected On', str(loan.rejected_on or '-'), 'Reason', loan.rejection_reason or '-'],
    ]

    workflow_table = Table(workflow_data, colWidths=[32 * mm, 48 * mm, 32 * mm, 48 * mm])
    workflow_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f4f8')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f0f4f8')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    story.append(Paragraph('Workflow Details', styles['Heading2']))
    story.append(workflow_table)
    story.append(Spacer(1, 12))

    repayment_data = [['Date', 'Method', 'Reference', 'Amount']]

    for r in repayments:
        repayment_data.append([
            str(r.paid_on or '-'),
            r.method or '-',
            r.reference or '-',
            kwacha(r.amount),
        ])

    if len(repayment_data) == 1:
        repayment_data.append(['-', '-', 'No repayments recorded', '-'])

    repayment_table = Table(repayment_data, colWidths=[32 * mm, 38 * mm, 58 * mm, 32 * mm])
    repayment_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4f68')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    story.append(Paragraph('Repayment History', styles['Heading2']))
    story.append(repayment_table)

    doc.build(story)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'loan_statement_{loan.id}.pdf',
        mimetype='application/pdf'
    )

@app.route('/loans/<int:loan_id>')
@login_required
@role_required('loans')
def loan_details(loan_id):
    loan = Loan.query.get_or_404(loan_id)

    repayments = Repayment.query.filter_by(
        loan_id=loan.id
    ).order_by(
        Repayment.paid_on.desc(),
        Repayment.id.desc()
    ).all()

    return render_template(
    'loan_details.html',
    loan=loan,
    repayments=repayments,
    members=Member.query.order_by(Member.member_no).all()
    )

@app.route('/loans/<int:loan_id>/guarantor', methods=['POST'])
@login_required
@role_required('loans')
def loan_add_guarantor(loan_id):
    loan = Loan.query.get_or_404(loan_id)

    guarantor_member_id = request.form.get('guarantor_member_id', type=int)

    if not guarantor_member_id:
        flash('Please select a guarantor.', 'error')
        return redirect(url_for('loan_details', loan_id=loan.id))

    if guarantor_member_id == loan.member_id:
        flash('A borrower cannot guarantee their own loan.', 'error')
        return redirect(url_for('loan_details', loan_id=loan.id))

    existing = LoanGuarantor.query.filter_by(
        loan_id=loan.id,
        guarantor_member_id=guarantor_member_id
    ).first()

    if existing:
        flash('This member is already a guarantor for this loan.', 'error')
        return redirect(url_for('loan_details', loan_id=loan.id))

    guarantor = LoanGuarantor(
        loan_id=loan.id,
        guarantor_member_id=guarantor_member_id,
        guaranteed_amount=money(request.form.get('guaranteed_amount') or 0),
        notes=request.form.get('notes')
    )

    db.session.add(guarantor)
    db.session.commit()

    log_audit(
        'ADD_LOAN_GUARANTOR',
        'LoanGuarantor',
        guarantor.id,
        f'Guarantor added to loan {loan.loan_no or loan.id}'
    )

    flash('Guarantor added successfully.')
    return redirect(url_for('loan_details', loan_id=loan.id))

@app.route('/loans/<int:loan_id>/review', methods=['POST'])
@login_required
@role_required('loans')
def loan_review(loan_id):
    loan = Loan.query.get_or_404(loan_id)

    if loan.status != 'Applied':
        flash('Only applied loans can be reviewed.', 'error')
        return redirect(url_for('loans'))

    user = session.get('user') or {}

    loan.status = 'Reviewed'
    loan.reviewed_by = user.get('full_name') or user.get('username')
    loan.reviewed_on = date.today()

    db.session.commit()

    log_audit(
        'LOAN_REVIEWED',
        'Loan',
        loan.id,
        f'{loan.member.full_name} loan reviewed by {loan.reviewed_by}'
    )

    flash('Loan application marked as reviewed.')
    return redirect(url_for('loans'))


# ------------------------------------------------------------
# LOAN REPAYMENT WORKFLOW
#
# Loan Status Flow:
# Applied
#   -> Reviewed
#   -> Approved
#   -> Disbursed
#   -> Partially Paid
#   -> Paid
#
# Repayments may ONLY be recorded for:
#   - Disbursed
#   - Partially Paid
#
# Repayments are NOT allowed for:
#   - Applied
#   - Reviewed
#   - Approved
#   - Paid
#   - Rejected
#
# After repayment:
#   - If balance > 0  -> status becomes 'Partially Paid'
#   - If balance = 0  -> status becomes 'Paid'
#
# This workflow preserves a complete audit trail:
# Review -> Approval -> Disbursement -> Repayment -> Closure
# ------------------------------------------------------------

@app.route('/repayments', methods=['POST'])
@login_required
@role_required('repayments')
def repayments():
    loan = Loan.query.get_or_404(int(request.form['loan_id']))

    if loan.status not in ['Disbursed', 'Partially Paid']:
        flash('Repayment can only be recorded for disbursed loans.', 'error')
        return redirect(url_for('loans'))

    amount = money(request.form['amount'])

    if amount <= 0:
        flash('Repayment amount must be greater than zero.', 'error')
        return redirect(url_for('loans'))

    if amount > loan.balance:
        flash('Repayment amount cannot exceed the outstanding loan balance.', 'error')
        return redirect(url_for('loans'))

    r = Repayment(
        loan_id=loan.id,
        amount=amount,
        method=request.form['method'],
        reference=request.form.get('reference'),
        paid_on=parse_date(request.form.get('paid_on'))
    )

    db.session.add(r)
    db.session.flush()

    post_to_cash_book(
        entry_date=r.paid_on,
        entry_type='In',
        category='Loan Repayment',
        amount=r.amount,
        description=f'{loan.member.member_no} - {loan.member.full_name}',
        method=r.method,
        reference=r.reference,
        source_type='Repayment',
        source_id=r.id
    )

    post_journal(
        entry_date=r.paid_on,
        description=f'Loan repayment - {loan.member.member_no} - {loan.member.full_name}',
        debit_account_code=cash_account(r.method),
        credit_account_code='1100',
        amount=r.amount,
        source_type='Repayment',
        source_id=r.id
    )

    db.session.commit()
    db.session.refresh(loan)
    
    log_audit(
        'RECORD_REPAYMENT',
        'Repayment',
        r.id,
        f'{loan.member.full_name} paid {kwacha(r.amount)} on loan #{loan.id} via {r.method}'
    )

    if loan.balance <= 0:
        loan.status = 'Paid'
        db.session.commit()

        log_audit(
            'CLOSE_LOAN',
            'Loan',
            loan.id,
            f'Loan for {loan.member.full_name} fully repaid'
        )
    else:
        loan.status = 'Partially Paid'
        db.session.commit()

    flash('Repayment recorded.')
    return redirect(url_for('loans'))

@app.route('/loans/<int:loan_id>/approve', methods=['POST'])
@login_required
@role_required('loans')
def loan_approve(loan_id):
    loan = Loan.query.get_or_404(loan_id)

    if loan.status != 'Reviewed':
        flash('Only reviewed loans can be approved.', 'error')
        return redirect(url_for('loans'))

    loan.status = 'Approved'
    loan.approved_on = date.today()

    user = session.get('user') or {}
    loan.approved_by = user.get('full_name') or user.get('username')

    db.session.commit()

    log_audit(
        'LOAN_APPROVED',
        'Loan',
        loan.id,
        f'{loan.member.full_name} loan approved by {loan.approved_by}'
    )
    flash('Loan approved successfully.')
    return redirect(url_for('loans'))


@app.route('/loans/<int:loan_id>/disburse', methods=['POST'])
@login_required
@role_required('loans')
def loan_disburse(loan_id):
    loan = Loan.query.get_or_404(loan_id)

    if loan.status != 'Approved':
        flash('Only approved loans can be disbursed.', 'error')
        return redirect(url_for('loans'))

    loan.status = 'Disbursed'
    loan.disbursed_on = date.today()

    loan.disbursement_method = request.form.get('method')
    loan.disbursement_reference = request.form.get('reference')

    user = session.get('user') or {}
    loan.disbursed_by = user.get('full_name') or user.get('username')

    post_to_cash_book(
        entry_date=loan.disbursed_on,
        entry_type='Out',
        category='Loan Disbursement',
        amount=loan.principal,
        description=f'{loan.member.member_no} - {loan.member.full_name}',
        method=loan.disbursement_method,
        reference=loan.disbursement_reference,
        source_type='Loan',
        source_id=loan.id
    )

    post_journal(
        entry_date=loan.disbursed_on,
        description=f'Loan disbursement - {loan.member.member_no} - {loan.member.full_name}',
        debit_account_code='1100',
        credit_account_code=cash_account(loan.disbursement_method),
        amount=loan.principal,
        source_type='Loan',
        source_id=loan.id
    )

    db.session.commit()

    log_audit(
        'DISBURSE_LOAN',
        'Loan',
        loan.id,
        f'Loan for {loan.member.full_name} disbursed'
    )

    flash('Loan disbursed successfully.')
    return redirect(url_for('loans'))

@app.route('/loans/<int:loan_id>/reject', methods=['POST'])
@login_required
@role_required('loans')
def loan_reject(loan_id):
    loan = Loan.query.get_or_404(loan_id)

    if loan.status not in ['Applied', 'Reviewed']:
        flash('Only applied or reviewed loans can be rejected.', 'error')
        return redirect(url_for('loans'))

    reason = (request.form.get('rejection_reason') or '').strip()

    if not reason:
        flash('Rejection reason is required.', 'error')
        return redirect(url_for('loans'))

    user = session.get('user') or {}

    loan.status = 'Rejected'
    loan.rejected_on = date.today()
    loan.rejection_reason = reason
    loan.reviewed_by = user.get('full_name') or user.get('username')

    db.session.commit()

    log_audit(
        'LOAN_REJECTED',
        'Loan',
        loan.id,
        f'{loan.member.full_name} loan rejected. Reason: {reason}'
    )

    flash('Loan application rejected.')
    return redirect(url_for('loans'))

@app.route('/loans/aging')
@login_required
@role_required('loans')
def loan_aging_report():
    today = date.today()

    loans = Loan.query.filter(
    Loan.status.in_(['Disbursed', 'Partially Paid'])
        ).order_by(
            Loan.due_on.asc(),
            Loan.id.desc()
        ).all()

    loans = [
            loan for loan in loans
            if loan.balance > 0
        ]
    buckets = {
        'Current': [],
        '1-30 Days Overdue': [],
        '31-60 Days Overdue': [],
        '61-90 Days Overdue': [],
        '90+ Days Overdue': [],
    }

    for loan in loans:
        if not loan.due_on or loan.due_on >= today:
            buckets['Current'].append(loan)
            continue

        days_overdue = (today - loan.due_on).days

        if days_overdue <= 30:
            buckets['1-30 Days Overdue'].append(loan)
        elif days_overdue <= 60:
            buckets['31-60 Days Overdue'].append(loan)
        elif days_overdue <= 90:
            buckets['61-90 Days Overdue'].append(loan)
        else:
            buckets['90+ Days Overdue'].append(loan)

    bucket_totals = {
        name: money(sum((loan.balance for loan in items), Decimal('0.00')))
        for name, items in buckets.items()
    }

    grand_total = money(sum(bucket_totals.values(), Decimal('0.00')))

    return render_template(
        'loan_aging.html',
        buckets=buckets,
        bucket_totals=bucket_totals,
        grand_total=grand_total,
        today=today
    )


@app.route('/distributions', methods=['GET', 'POST'])
@login_required
@role_required('distributions')
def distributions():
    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.values.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.values.get('end_month')
        or today_month
    )

    expenses = money(
        request.values.get('expenses') or 0
    )

    other_income = money(
        request.values.get('other_income') or 0
    )

    start_date = datetime.strptime(
        start_month + '-01',
        '%Y-%m-%d'
    ).date()

    end_year, end_mon = [
        int(value)
        for value in end_month.split('-')
    ]

    end_date = (
        date(end_year, end_mon, 28)
        + timedelta(days=4)
    ).replace(day=1) - timedelta(days=1)

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    shareout_rows = shareout_data['rows']

    cycle = get_shareout_cycle(
        start_month,
        end_month,
    )

    cycle_locked = bool(
        cycle and cycle.status == 'Locked'
    )

    if request.method == 'POST':
        if cycle_locked:
            flash(
                'This Share-Out cycle is locked. '
                'No further payments can be recorded.',
                'error'
            )

            return redirect(
                url_for(
                    'distributions',
                    start_month=start_month,
                    end_month=end_month,
                    expenses=expenses,
                    other_income=other_income,
                )
            )
        
        member_id = int(request.form['member_id'])
        payment_amount = money(request.form['amount'])

        member_shareout = next(
            (
                row
                for row in shareout_rows
                if row['member_id'] == member_id
            ),
            None
        )

        if not member_shareout:
            flash(
                'The selected member has no Share-Out allocation '
                'for this period.',
                'error'
            )

            return redirect(
                url_for(
                    'distributions',
                    start_month=start_month,
                    end_month=end_month,
                    expenses=expenses,
                    other_income=other_income,
                )
            )

        already_paid = money(
            db.session.query(
                db.func.coalesce(
                    db.func.sum(Distribution.amount),
                    0
                )
            )
            .filter(Distribution.member_id == member_id)
            .filter(Distribution.paid_on >= start_date)
            .filter(Distribution.paid_on <= end_date)
            .scalar()
        )

        outstanding_balance = money(
            member_shareout['net_shareout'] - already_paid
        )

        if payment_amount <= 0:
            flash(
                'Payment amount must be greater than zero.',
                'error'
            )

            return redirect(
                url_for(
                    'distributions',
                    start_month=start_month,
                    end_month=end_month,
                    expenses=expenses,
                    other_income=other_income,
                )
            )

        if outstanding_balance <= 0:
            flash(
                'This member has no outstanding Share-Out balance.',
                'error'
            )

            return redirect(
                url_for(
                    'distributions',
                    start_month=start_month,
                    end_month=end_month,
                    expenses=expenses,
                    other_income=other_income,
                )
            )

        if payment_amount > outstanding_balance:
            flash(
                (
                    'Payment exceeds the member’s outstanding '
                    f'Share-Out balance of '
                    f'{kwacha(outstanding_balance)}.'
                ),
                'error'
            )

            return redirect(
                url_for(
                    'distributions',
                    start_month=start_month,
                    end_month=end_month,
                    expenses=expenses,
                    other_income=other_income,
                )
            )

        payment_date = (
            parse_date(request.form.get('paid_on'))
            or date.today()
        )

        d = Distribution(
            member_id=member_id,
            amount=payment_amount,
            method=request.form['method'],
            reference=request.form.get('reference'),
            authorized_by=request.form.get('authorized_by'),
            paid_on=payment_date,
        )

        db.session.add(d)
        db.session.flush()

        post_to_cash_book(
            entry_date=d.paid_on,
            entry_type='Out',
            category='Share-Out Payment',
            amount=d.amount,
            description=(
                f'{d.member.member_no} - '
                f'{d.member.full_name}'
            ),
            method=d.method,
            reference=d.reference,
            source_type='Distribution',
            source_id=d.id,
        )

        post_journal(
            entry_date=d.paid_on,
            description=(
                f'Share-out payment - '
                f'{d.member.member_no} - '
                f'{d.member.full_name}'
            ),
            debit_account_code='5040',
            credit_account_code=cash_account(d.method),
            amount=d.amount,
            source_type='Distribution',
            source_id=d.id,
        )

        db.session.commit()

        log_audit(
            'RECORD_DISTRIBUTION',
            'Distribution',
            d.id,
            (
                f'{d.member.full_name} received '
                f'{kwacha(d.amount)} via {d.method}'
            )
        )

        flash('Share-Out payment recorded successfully.', 'success')

        return redirect(
            url_for(
                'distributions',
                start_month=start_month,
                end_month=end_month,
                expenses=expenses,
                other_income=other_income,
            )
        )

    page = request.args.get('page', 1, type=int)
    per_page = 25

    period_query = Distribution.query.filter(
        Distribution.paid_on >= start_date,
        Distribution.paid_on <= end_date,
    )

    pagination = period_query.order_by(
        Distribution.paid_on.desc(),
        Distribution.id.desc(),
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False,
    )

    total_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .scalar()
    )

    payments_by_member = dict(
        db.session.query(
            Distribution.member_id,
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .group_by(Distribution.member_id)
        .all()
    )

    payment_schedule = []

    members_paid = 0
    members_partially_paid = 0
    members_not_paid = 0
    total_outstanding = Decimal('0.00')

    for row in shareout_rows:
        net_shareout = money(row['net_shareout'])

        member_paid = money(
            payments_by_member.get(
                row['member_id'],
                Decimal('0.00')
            )
        )

        outstanding = money(
            net_shareout - member_paid
        )

        if outstanding < 0:
            status = 'Overpaid'
        elif member_paid <= 0:
            status = 'Pending'
            members_not_paid += 1
        elif outstanding > 0:
            status = 'Partially Paid'
            members_partially_paid += 1
        else:
            status = 'Paid'
            members_paid += 1

        if outstanding > 0:
            total_outstanding += outstanding

        payment_schedule.append({
            **row,
            'amount_paid': member_paid,
            'outstanding_balance': outstanding,
            'payment_status': status,
        })

    total_outstanding = money(total_outstanding)

    eligible_members = len([
        row
        for row in shareout_rows
        if money(row['net_shareout']) > 0
    ])

    payment_progress = (
        0
        if eligible_members == 0
        else round(
            (members_paid / eligible_members) * 100
        )
    )

    bank_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .filter(
            Distribution.method == 'Bank Transfer'
        )
        .scalar()
    )

    mobile_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .filter(
            Distribution.method.in_([
                'Mobile Money',
                'Airtel Money',
                'MTN Money',
            ])
        )
        .scalar()
    )

    cash_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .filter(
            Distribution.method == 'Cash'
        )
        .scalar()
    )

    missing_references = period_query.filter(
        db.or_(
            Distribution.reference.is_(None),
            Distribution.reference == '',
        )
    ).count()

    today = date.today()

    today_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on == today)
        .scalar()
    )

    today_transactions = Distribution.query.filter(
        Distribution.paid_on == today
    ).count()

    largest_distribution = period_query.order_by(
        Distribution.amount.desc()
    ).first()

    if eligible_members == 0:
        distribution_message = (
            'No eligible Share-Out payments are available '
            'for the selected period.'
        )
    elif members_not_paid == 0 and members_partially_paid == 0:
        distribution_message = (
            'Distribution completed successfully.'
        )
    elif payment_progress >= 75:
        distribution_message = (
            'Distribution is progressing well.'
        )
    elif payment_progress >= 25:
        distribution_message = (
            'Continue processing outstanding payments.'
        )
    else:
        distribution_message = (
            'Distribution has only just begun.'
        )

    members = [
        row
        for row in payment_schedule
        if (
            row['payment_status']
            in ['Pending', 'Partially Paid']
            and money(row['net_shareout']) > 0
        )
    ]

    return render_template(
        'distributions.html',
        distributions=pagination.items,
        pagination=pagination,
        members=members,
        payment_schedule=payment_schedule,
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
        total_paid=total_paid,
        total_outstanding=total_outstanding,
        members_paid=members_paid,
        members_partially_paid=members_partially_paid,
        members_not_paid=members_not_paid,
        eligible_members=eligible_members,
        bank_paid=bank_paid,
        mobile_paid=mobile_paid,
        cash_paid=cash_paid,
        missing_references=missing_references,
        payment_progress=payment_progress,
        today_paid=today_paid,
        today_transactions=today_transactions,
        largest_distribution=largest_distribution,
        distribution_message=distribution_message,
        shareout_fund=shareout_data['shareout_fund'],
        total_net_payable=shareout_data['total_net_payable'],
        cycle=cycle,
        cycle_locked=cycle_locked,
    )
@app.route('/distributions.pdf')
@login_required
@role_required('distributions')
def distributions_pdf():
    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.args.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.args.get('end_month')
        or today_month
    )

    expenses = money(
        request.args.get('expenses') or 0
    )

    other_income = money(
        request.args.get('other_income') or 0
    )

    start_date = datetime.strptime(
        start_month + '-01',
        '%Y-%m-%d'
    ).date()

    end_year, end_mon = [
        int(value)
        for value in end_month.split('-')
    ]

    end_date = (
        date(end_year, end_mon, 28)
        + timedelta(days=4)
    ).replace(day=1) - timedelta(days=1)

    setting = SystemSetting.query.first()

    organization_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    cycle = get_shareout_cycle(
        start_month,
        end_month,
    )

    cycle_status = (
        cycle.status
        if cycle
        else 'Draft'
    )

    payments = Distribution.query.filter(
        Distribution.paid_on >= start_date,
        Distribution.paid_on <= end_date,
    ).order_by(
        Distribution.paid_on.asc(),
        Distribution.id.asc(),
    ).all()

    total_paid = money(
        sum(
            (
                money(payment.amount)
                for payment in payments
            ),
            Decimal('0.00')
        )
    )

    payments_by_member = dict(
        db.session.query(
            Distribution.member_id,
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .group_by(Distribution.member_id)
        .all()
    )

    payment_schedule = []

    members_paid = 0
    members_partially_paid = 0
    members_pending = 0
    members_overpaid = 0
    members_requiring_review = 0

    total_outstanding = Decimal('0.00')

    for row in shareout_data['rows']:
        member_id = row['member_id']

        net_shareout = money(
            row.get(
                'net_shareout',
                row.get('net_payable', 0)
            )
        )

        amount_paid = money(
            payments_by_member.get(
                member_id,
                Decimal('0.00')
            )
        )

        outstanding_balance = money(
            net_shareout - amount_paid
        )

        if net_shareout <= 0:
            payment_status = 'Review'
            members_requiring_review += 1

        elif outstanding_balance < 0:
            payment_status = 'Overpaid'
            members_overpaid += 1

        elif amount_paid <= 0:
            payment_status = 'Pending'
            members_pending += 1

        elif outstanding_balance > 0:
            payment_status = 'Partially Paid'
            members_partially_paid += 1

        else:
            payment_status = 'Paid'
            members_paid += 1

        if outstanding_balance > 0:
            total_outstanding += outstanding_balance

        payment_schedule.append({
            **row,
            'net_shareout': net_shareout,
            'amount_paid': amount_paid,
            'outstanding_balance': outstanding_balance,
            'payment_status': payment_status,
        })

    total_outstanding = money(
        total_outstanding
    )

    eligible_members = len([
        row
        for row in payment_schedule
        if money(row['net_shareout']) > 0
    ])

    payment_progress = (
        0
        if eligible_members == 0
        else round(
            (members_paid / eligible_members) * 100
        )
    )

    bank_paid = money(
        sum(
            (
                money(payment.amount)
                for payment in payments
                if payment.method == 'Bank Transfer'
            ),
            Decimal('0.00')
        )
    )

    mobile_paid = money(
        sum(
            (
                money(payment.amount)
                for payment in payments
                if payment.method in [
                    'Mobile Money',
                    'Airtel Money',
                    'MTN Money',
                ]
            ),
            Decimal('0.00')
        )
    )

    cash_paid = money(
        sum(
            (
                money(payment.amount)
                for payment in payments
                if payment.method == 'Cash'
            ),
            Decimal('0.00')
        )
    )

    missing_references = sum(
        1
        for payment in payments
        if not payment.reference
    )

    primary_colour = pdf_colour(
        setting.primary_color if setting else None,
        '#0D6EFD'
    )

    secondary_colour = pdf_colour(
        setting.secondary_color if setting else None,
        '#198754'
    )

    table_header_colour = pdf_colour(
        setting.table_header_color if setting else None,
        '#EAF2F8'
    )

    success_colour = pdf_colour(
        setting.success_color if setting else None,
        '#198754'
    )

    warning_colour = pdf_colour(
        setting.warning_color if setting else None,
        '#FFC107'
    )

    danger_colour = pdf_colour(
        setting.danger_color if setting else None,
        '#DC3545'
    )

    buffer = io.BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=17 * mm,
        title=(
            f'Distribution Payment Register '
            f'{start_month} to {end_month}'
        ),
        author=organization_name,
    )

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'DistributionRegisterTitle',
        parent=styles['Heading1'],
        fontSize=15,
        leading=18,
        textColor=primary_colour,
        spaceAfter=7,
    )

    section_style = ParagraphStyle(
        'DistributionRegisterSection',
        parent=styles['Heading2'],
        fontSize=10,
        leading=12,
        textColor=primary_colour,
        spaceBefore=7,
        spaceAfter=6,
    )

    small_style = ParagraphStyle(
        'DistributionRegisterSmall',
        parent=styles['Normal'],
        fontSize=6.5,
        leading=8,
    )

    elements.append(
        build_pdf_branding(
            setting,
            styles
        )
    )

    elements.append(Spacer(1, 6))

    elements.append(
        Paragraph(
            'DISTRIBUTION PAYMENT REGISTER',
            title_style
        )
    )

    period_table = Table(
        [[
            'Share-Out Period',
            f'{start_month} to {end_month}',
            'Cycle Status',
            cycle_status,
        ]],
        colWidths=[
            32 * mm,
            55 * mm,
            29 * mm,
            42 * mm,
        ],
        hAlign='LEFT',
    )

    period_table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), table_header_colour),
            ('BACKGROUND', (2, 0), (2, -1), table_header_colour),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('GRID', (0, 0), (-1, -1), 0.4, primary_colour),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ])
    )

    elements.append(period_table)
    elements.append(Spacer(1, 8))

    summary_data = [
        [
            'Total Paid',
            kwacha(total_paid),
            'Outstanding',
            kwacha(total_outstanding),
            'Progress',
            f'{payment_progress}%',
        ],
        [
            'Eligible Members',
            str(eligible_members),
            'Paid',
            str(members_paid),
            'Partially Paid',
            str(members_partially_paid),
        ],
        [
            'Pending',
            str(members_pending),
            'Overpaid',
            str(members_overpaid),
            'Require Review',
            str(members_requiring_review),
        ],
        [
            'Bank Transfers',
            kwacha(bank_paid),
            'Mobile Money',
            kwacha(mobile_paid),
            'Cash Payments',
            kwacha(cash_paid),
        ],
        [
            'Missing References',
            str(missing_references),
            'Net Payable',
            kwacha(shareout_data['total_net_payable']),
            'Share-Out Fund',
            kwacha(shareout_data['shareout_fund']),
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            32 * mm,
            30 * mm,
            30 * mm,
            30 * mm,
            32 * mm,
            32 * mm,
        ],
        hAlign='LEFT',
    )

    summary_table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), table_header_colour),
            ('BACKGROUND', (2, 0), (2, -1), table_header_colour),
            ('BACKGROUND', (4, 0), (4, -1), table_header_colour),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTNAME', (4, 0), (4, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
            (
                'GRID',
                (0, 0),
                (-1, -1),
                0.35,
                colors.HexColor('#AAB7C4')
            ),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ])
    )

    elements.append(summary_table)
    elements.append(Spacer(1, 8))

    elements.append(
        Paragraph(
            'Member Payment Status',
            section_style
        )
    )

    member_data = [[
        'Member',
        'Net Share-Out',
        'Amount Paid',
        'Outstanding',
        'Status',
    ]]

    for row in payment_schedule:
        member_data.append([
            Paragraph(
                (
                    f"<b>{row['member_no']}</b><br/>"
                    f"{row['full_name']}"
                ),
                small_style
            ),
            kwacha(row['net_shareout']),
            kwacha(row['amount_paid']),
            kwacha(row['outstanding_balance']),
            row['payment_status'],
        ])

    if not payment_schedule:
        member_data.append([
            'No Share-Out allocations found',
            '',
            '',
            '',
            '',
        ])

    member_table = Table(
        member_data,
        repeatRows=1,
        colWidths=[
            62 * mm,
            34 * mm,
            34 * mm,
            34 * mm,
            32 * mm,
        ],
        hAlign='LEFT',
    )

    member_style = [
        ('BACKGROUND', (0, 0), (-1, 0), primary_colour),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6.5),
        ('FONTSIZE', (0, 1), (-1, -1), 6.2),
        ('ALIGN', (1, 1), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        (
            'GRID',
            (0, 0),
            (-1, -1),
            0.3,
            colors.HexColor('#AAB7C4')
        ),
        (
            'ROWBACKGROUNDS',
            (0, 1),
            (-1, -1),
            [
                colors.white,
                table_header_colour,
            ]
        ),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    for row_number, row in enumerate(
        payment_schedule,
        start=1
    ):
        status = row['payment_status']

        if status == 'Paid':
            status_colour = success_colour
        elif status == 'Partially Paid':
            status_colour = warning_colour
        elif status in ['Overpaid', 'Review']:
            status_colour = danger_colour
        else:
            status_colour = primary_colour

        member_style.extend([
            (
                'TEXTCOLOR',
                (4, row_number),
                (4, row_number),
                status_colour
            ),
            (
                'FONTNAME',
                (4, row_number),
                (4, row_number),
                'Helvetica-Bold'
            ),
        ])

    member_table.setStyle(
        TableStyle(member_style)
    )

    elements.append(member_table)
    elements.append(Spacer(1, 8))

    elements.append(
        Paragraph(
            'Payment Transaction Register',
            section_style
        )
    )

    transaction_data = [[
        'Date',
        'Member',
        'Method',
        'Reference',
        'Authorized By',
        'Amount',
    ]]

    for payment in payments:
        transaction_data.append([
            payment.paid_on.strftime('%d %b %Y'),
            Paragraph(
                (
                    f'<b>{payment.member.member_no}</b><br/>'
                    f'{payment.member.full_name}'
                ),
                small_style
            ),
            payment.method,
            payment.reference or 'Missing',
            payment.authorized_by or '-',
            kwacha(payment.amount),
        ])

    if not payments:
        transaction_data.append([
            'No payments recorded',
            '',
            '',
            '',
            '',
            '',
        ])

    transaction_table = Table(
        transaction_data,
        repeatRows=1,
        colWidths=[
            25 * mm,
            50 * mm,
            35 * mm,
            42 * mm,
            38 * mm,
            30 * mm,
        ],
        hAlign='LEFT',
    )

    transaction_style = [
        ('BACKGROUND', (0, 0), (-1, 0), primary_colour),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 6.5),
        ('FONTSIZE', (0, 1), (-1, -1), 6.2),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        (
            'GRID',
            (0, 0),
            (-1, -1),
            0.3,
            colors.HexColor('#AAB7C4')
        ),
        (
            'ROWBACKGROUNDS',
            (0, 1),
            (-1, -1),
            [
                colors.white,
                table_header_colour,
            ]
        ),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    for row_number, payment in enumerate(
        payments,
        start=1
    ):
        if not payment.reference:
            transaction_style.extend([
                (
                    'TEXTCOLOR',
                    (3, row_number),
                    (3, row_number),
                    danger_colour
                ),
                (
                    'FONTNAME',
                    (3, row_number),
                    (3, row_number),
                    'Helvetica-Bold'
                ),
            ])

    transaction_table.setStyle(
        TableStyle(transaction_style)
    )

    elements.append(transaction_table)
    elements.append(Spacer(1, 12))

    certification_data = [
        [
            '________________________',
            '________________________',
            '________________________',
            '________________________',
        ],
        [
            'Prepared By',
            'Verified By',
            'Treasurer',
            'Chairperson',
        ],
        [
            'Date: __________________',
            'Date: __________________',
            'Date: __________________',
            'Date: __________________',
        ],
    ]

    certification_table = Table(
        certification_data,
        colWidths=[
            47 * mm,
            47 * mm,
            47 * mm,
            47 * mm,
        ],
    )

    certification_table.setStyle(
        TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ])
    )

    elements.append(certification_table)

    generated_on = datetime.now().strftime(
        '%d %B %Y at %H:%M'
    )

    elements.append(Spacer(1, 7))

    elements.append(
        Paragraph(
            f'Generated on {generated_on}',
            small_style
        )
    )

    document.build(
        elements,
        onFirstPage=lambda canvas, doc: draw_pdf_footer(
            canvas,
            doc,
            setting
        ),
        onLaterPages=lambda canvas, doc: draw_pdf_footer(
            canvas,
            doc,
            setting
        ),
    )

    buffer.seek(0)

    log_audit(
        'EXPORT_DISTRIBUTION_REGISTER_PDF',
        'Distribution',
        None,
        (
            f'Distribution Payment Register PDF exported '
            f'for {start_month} to {end_month}'
        )
    )

    filename = (
        f'distribution_payment_register_'
        f'{start_month}_to_{end_month}.pdf'
    )

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition':
                f'inline; filename="{filename}"'
        }
    )

@app.route(
    '/share-out-approval',
    methods=['GET', 'POST']
)
@login_required
@role_required('shareout')
def shareout_approval():
    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.values.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.values.get('end_month')
        or today_month
    )

    expenses = money(
        request.values.get('expenses') or 0
    )

    other_income = money(
        request.values.get('other_income') or 0
    )

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    cycle = ShareOutCycle.query.filter_by(
        start_month=start_month,
        end_month=end_month,
    ).first()

    if not cycle:
        cycle = ShareOutCycle(
            start_month=start_month,
            end_month=end_month,
            expenses=expenses,
            other_income=other_income,
            total_contributions=shareout_data[
                'total_contributions'
            ],
            shareout_fund=shareout_data[
                'shareout_fund'
            ],
            total_net_payable=shareout_data[
                'total_net_payable'
            ],
            eligible_members=shareout_data[
                'eligible_members'
            ],
            members_requiring_review=shareout_data[
                'members_requiring_review'
            ],
            readiness_score=shareout_data[
                'readiness_score'
            ],
            status='Draft',
            created_by=(
                session.get('user', {}).get('username')
                or session.get('username')
                or 'System'
            ),
        )

        db.session.add(cycle)
        db.session.commit()

    if cycle.status != 'Locked':
        cycle.expenses = expenses
        cycle.other_income = other_income
        cycle.total_contributions = shareout_data[
            'total_contributions'
        ]
        cycle.shareout_fund = shareout_data[
            'shareout_fund'
        ]
        cycle.total_net_payable = shareout_data[
            'total_net_payable'
        ]
        cycle.eligible_members = shareout_data[
            'eligible_members'
        ]
        cycle.members_requiring_review = shareout_data[
            'members_requiring_review'
        ]
        cycle.readiness_score = shareout_data[
            'readiness_score'
        ]

        db.session.commit()

    start_date = datetime.strptime(
        start_month + '-01',
        '%Y-%m-%d'
    ).date()

    end_year, end_mon = [
        int(value)
        for value in end_month.split('-')
    ]

    end_date = (
        date(end_year, end_mon, 28)
        + timedelta(days=4)
    ).replace(day=1) - timedelta(days=1)

    total_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .scalar()
    )

    outstanding_balance = money(
        shareout_data['total_net_payable']
        - total_paid
    )

    session_user = session.get('user') or {}

    operator_name = (
        session_user.get('full_name')
        or session_user.get('name')
        or session_user.get('username')
        or session.get('full_name')
        or session.get('name')
        or session.get('username')
        or session.get('user_name')
        or 'System User'
    )

    if request.method == 'POST':
        action = request.form.get('action')
        authorized_by = (
            request.form.get('authorized_by') or ''
        ).strip()

        comments = (
            request.form.get('comments') or ''
        ).strip()

        if not authorized_by:
            flash(
                'Please select the authorizing officer.',
                'error'
            )

            return redirect(
                url_for(
                    'shareout_approval',
                    start_month=start_month,
                    end_month=end_month,
                    expenses=expenses,
                    other_income=other_income,
                )
            )

        if action == 'approve':
            if cycle.status == 'Locked':
                flash(
                    'This Share-Out cycle is already locked.',
                    'error'
                )

            elif shareout_data['readiness_score'] < 75:
                flash(
                    (
                        'The Share-Out cycle is not ready for '
                        'approval. Resolve the outstanding '
                        'review matters first.'
                    ),
                    'error'
                )

            else:
                cycle.status = 'Approved'
                cycle.approved_by = authorized_by
                cycle.approved_on = datetime.utcnow()
                cycle.approval_comments = comments

                db.session.commit()

                log_audit(
                    'APPROVE_SHAREOUT',
                    'ShareOutCycle',
                    cycle.id,
                    (
                        f'Share-Out cycle {start_month} to {end_month} '
                        f'approved by {authorized_by}; '
                        f'recorded in the system by {operator_name}'
                    )
                )

                flash(
                    'Share-Out cycle approved successfully.',
                    'success'
                )

        elif action == 'lock':
            if cycle.status == 'Locked':
                flash(
                    'This Share-Out cycle is already locked.',
                    'error'
                )

            elif cycle.status != 'Approved':
                flash(
                    (
                        'The Share-Out cycle must be approved '
                        'before it can be locked.'
                    ),
                    'error'
                )

            else:
                cycle.status = 'Locked'
                cycle.locked_by = authorized_by
                cycle.locked_on = datetime.utcnow()
                cycle.lock_comments = comments

                db.session.commit()

                log_audit(
                    'LOCK_SHAREOUT',
                    'ShareOutCycle',
                    cycle.id,
                    (
                        f'Share-Out cycle {start_month} to {end_month} '
                        f'locked by {authorized_by}; '
                        f'recorded in the system by {operator_name}'
                    )
                         )

                flash(
                    (
                        'Share-Out cycle locked successfully. '
                        'The approved figures are now final.'
                    ),
                    'success'
                )

        return redirect(
            url_for(
                'shareout_approval',
                start_month=start_month,
                end_month=end_month,
                expenses=expenses,
                other_income=other_income,
            )
        )

    return render_template(
        'shareout/approval.html',
        cycle=cycle,
        shareout_data=shareout_data,
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
        total_paid=total_paid,
        outstanding_balance=outstanding_balance,
        operator_name=operator_name,
)
@app.route('/share-out-statement/<int:member_id>/pdf')
@login_required
@role_required('shareout')
def shareout_statement_pdf(member_id):
    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.args.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.args.get('end_month')
        or today_month
    )

    expenses = money(
        request.args.get('expenses') or 0
    )

    other_income = money(
        request.args.get('other_income') or 0
    )

    member = Member.query.get_or_404(member_id)

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    member_shareout = next(
        (
            row
            for row in shareout_data['rows']
            if row['member_id'] == member_id
        ),
        None
    )

    if not member_shareout:
        flash(
            'No Share-Out calculation was found for this member '
            'during the selected period.',
            'error'
        )

        return redirect(
            url_for(
                'shareout_schedule',
                start_month=start_month,
                end_month=end_month,
                expenses=expenses,
                other_income=other_income,
            )
        )

    start_date = datetime.strptime(
        start_month + '-01',
        '%Y-%m-%d'
    ).date()

    end_year, end_mon = [
        int(value)
        for value in end_month.split('-')
    ]

    end_date = (
        date(end_year, end_mon, 28)
        + timedelta(days=4)
    ).replace(day=1) - timedelta(days=1)

    payments = Distribution.query.filter(
        Distribution.member_id == member_id,
        Distribution.paid_on >= start_date,
        Distribution.paid_on <= end_date,
    ).order_by(
        Distribution.paid_on.asc(),
        Distribution.id.asc(),
    ).all()

    total_paid = money(
        sum(
            (
                money(payment.amount)
                for payment in payments
            ),
            Decimal('0.00')
        )
    )

    outstanding_balance = money(
        member_shareout['net_shareout'] - total_paid
    )

    cycle = get_shareout_cycle(
        start_month,
        end_month,
    )

    cycle_status = (
        cycle.status
        if cycle
        else 'Draft'
    )

    setting = SystemSetting.query.first()

    primary_colour = pdf_colour(
        setting.primary_color if setting else None,
        '#0D6EFD'
    )

    secondary_colour = pdf_colour(
        setting.secondary_color if setting else None,
        '#198754'
    )

    table_header_colour = pdf_colour(
        setting.table_header_color if setting else None,
        '#EAF2F8'
    )

    organization_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    organization_address = (
        setting.organization_address
        if setting and setting.organization_address
        else ''
    )

    organization_phone = (
        setting.organization_phone
        if setting and setting.organization_phone
        else ''
    )

    organization_email = (
        setting.organization_email
        if setting and setting.organization_email
        else ''
    )

    registration_number = (
        setting.registration_number
        if setting and setting.registration_number
        else ''
    )

    buffer = io.BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=(
            f'Share-Out Statement - '
            f'{member.member_no} - {member.full_name}'
        ),
        author=organization_name,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'StatementTitle',
        parent=styles['Title'],
        fontSize=17,
        leading=21,
        textColor=primary_colour,
        spaceAfter=5,
    )

    centre_style = ParagraphStyle(
        'StatementCentre',
        parent=styles['Normal'],
        fontSize=8.5,
        leading=11,
        alignment=1,
        textColor=primary_colour,
    )

    section_style = ParagraphStyle(
        'StatementSection',
        parent=styles['Heading2'],
        fontSize=11,
        leading=14,
        textColor=primary_colour,
        spaceBefore=8,
        spaceAfter=7,
    )

    small_style = ParagraphStyle(
        'StatementSmall',
        parent=styles['Normal'],
        fontSize=7.5,
        leading=9.5,
    )

    elements = []

    elements.append(
        build_pdf_branding(
            setting,
            styles
        )
    )

    elements.append(
        Spacer(1, 8)
    )

    

    elements.append(
        Paragraph(
            'INDIVIDUAL SHARE-OUT STATEMENT',
            section_style
        )
    )

    member_data = [
        [
            'Member Number',
            member.member_no,
            'Statement Period',
            f'{start_month} to {end_month}',
        ],
        [
            'Member Name',
            member.full_name,
            'Cycle Status',
            cycle_status,
        ],
        [
            'Group',
            member_shareout['group_name'],
            'Payment Status',
            (
                'Paid'
                if outstanding_balance == 0
                else 'Partially Paid'
                if total_paid > 0
                else 'Pending'
            ),
        ],
    ]

    member_table = Table(
        member_data,
        colWidths=[
            30 * mm,
            58 * mm,
            30 * mm,
            48 * mm,
        ],
    )

    member_table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (0, -1),table_header_colour),
            ('BACKGROUND', (2, 0), (2, -1),table_header_colour),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4,primary_colour),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ])
    )

    elements.append(member_table)
    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            'Share-Out Calculation',
            section_style
        )
    )

    calculation_data = [
        ['Description', 'Amount'],
        [
            'Total Contributions',
            kwacha(member_shareout['contributed']),
        ],
        [
            'Savings Interest',
            kwacha(member_shareout['savings_interest']),
        ],
        [
            'Gross Savings Value',
            kwacha(member_shareout['gross_savings_value']),
        ],
        [
            'Contribution Share',
            f"{member_shareout['percent']}%",
        ],
        [
            'Profit Share',
            kwacha(member_shareout['profit_share']),
        ],
        [
            'Gross Share-Out',
            kwacha(member_shareout['gross_shareout']),
        ],
    ]

    calculation_table = Table(
        calculation_data,
        colWidths=[105 * mm, 55 * mm],
    )

    calculation_table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_colour),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.4,primary_colour),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.white,
                colors.HexColor('#f7f9fb'),
            ]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ])
    )

    elements.append(calculation_table)
    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            'Deductions and Payment Position',
            section_style
        )
    )

    deduction_data = [
        ['Description', 'Amount'],
        [
            'Outstanding Loans',
            kwacha(member_shareout['outstanding_loans']),
        ],
        [
            'Outstanding Fines',
            kwacha(member_shareout['fine_balance']),
        ],
        [
            'Total Deductions',
            kwacha(member_shareout['total_deductions']),
        ],
        [
            'Net Share-Out',
            kwacha(member_shareout['net_shareout']),
        ],
        [
            'Total Paid',
            kwacha(total_paid),
        ],
        [
            'Outstanding Balance',
            kwacha(outstanding_balance),
        ],
    ]

    deduction_table = Table(
        deduction_data,
        colWidths=[105 * mm, 55 * mm],
    )

    deduction_table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_colour),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (0, 4), (-1, 6), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.4, primary_colour),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.white,
                table_header_colour,
            ]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ])
    )

    elements.append(deduction_table)
    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            'Payment History',
            section_style
        )
    )

    payment_data = [[
        'Date',
        'Method',
        'Reference',
        'Authorized By',
        'Amount',
    ]]

    for payment in payments:
        payment_data.append([
            payment.paid_on.strftime('%d %b %Y'),
            payment.method,
            payment.reference or '-',
            payment.authorized_by or '-',
            kwacha(payment.amount),
        ])

    if not payments:
        payment_data.append([
            'No payments recorded',
            '',
            '',
            '',
            '',
        ])

    payment_table = Table(
        payment_data,
        repeatRows=1,
        colWidths=[
            26 * mm,
            32 * mm,
            42 * mm,
            38 * mm,
            28 * mm,
        ],
    )

    payment_table.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary_colour),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.4, primary_colour),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                colors.white,
                table_header_colour,
            ]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ])
    )

    elements.append(payment_table)
    elements.append(Spacer(1, 10))

    signature_data = [
        [
            '____________________________',
            '____________________________',
            '____________________________',
        ],
        [
            'Member Signature',
            'Treasurer',
            'Chairperson',
        ],
    ]

    signature_table = Table(
        signature_data,
        colWidths=[
            55 * mm,
            55 * mm,
            55 * mm,
        ],
    )

    signature_table.setStyle(
        TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ])
    )

    elements.append(signature_table)
    elements.append(Spacer(1, 10))

    generated_on = datetime.now().strftime(
        '%d %B %Y at %H:%M'
    )

    elements.append(
        Paragraph(
            f'Generated on {generated_on}',
            small_style
        )
    )

    document.build(
    elements,
    onFirstPage=lambda canvas, doc: draw_pdf_footer(
        canvas,
        doc,
        setting
    ),
    onLaterPages=lambda canvas, doc: draw_pdf_footer(
        canvas,
        doc,
        setting
    )
)

    buffer.seek(0)

    log_audit(
        'EXPORT_SHAREOUT_STATEMENT_PDF',
        'Member',
        member.id,
        (
            f'Individual Share-Out statement PDF exported for '
            f'{member.full_name}; period '
            f'{start_month} to {end_month}'
        )
    )

    filename = (
            f'shareout_statement_'
            f'{member.member_no}_'
            f'{start_month}_to_{end_month}.pdf'
        )

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition':
                f'inline; filename="{filename}"'
        }
    )
@app.route('/share-out-dashboard')
@login_required
@role_required('shareout')
def shareout_dashboard():
    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.args.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.args.get('end_month')
        or today_month
    )

    expenses = money(
        request.args.get('expenses') or 0
    )

    other_income = money(
        request.args.get('other_income') or 0
    )

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    start_date = datetime.strptime(
        start_month + '-01',
        '%Y-%m-%d'
    ).date()

    end_year, end_mon = [
        int(value)
        for value in end_month.split('-')
    ]

    end_date = (
        date(end_year, end_mon, 28)
        + timedelta(days=4)
    ).replace(day=1) - timedelta(days=1)

    cycle = get_shareout_cycle(
        start_month,
        end_month,
    )

    cycle_status = (
        cycle.status
        if cycle
        else 'Draft'
    )

    payment_query = Distribution.query.filter(
        Distribution.paid_on >= start_date,
        Distribution.paid_on <= end_date,
    )

    total_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .scalar()
    )

    payments_by_member = dict(
        db.session.query(
            Distribution.member_id,
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .group_by(Distribution.member_id)
        .all()
    )

    payment_schedule = []

    members_paid = 0
    members_partially_paid = 0
    members_pending = 0
    members_overpaid = 0

    total_outstanding = Decimal('0.00')

    for row in shareout_data['rows']:
        net_shareout = money(
            row['net_shareout']
        )

        amount_paid = money(
            payments_by_member.get(
                row['member_id'],
                Decimal('0.00')
            )
        )

        outstanding_balance = money(
            net_shareout - amount_paid
        )

        if net_shareout <= 0:
            payment_status = 'Review'

        elif outstanding_balance < 0:
            payment_status = 'Overpaid'
            members_overpaid += 1

        elif amount_paid <= 0:
            payment_status = 'Pending'
            members_pending += 1

        elif outstanding_balance > 0:
            payment_status = 'Partially Paid'
            members_partially_paid += 1

        else:
            payment_status = 'Paid'
            members_paid += 1

        if outstanding_balance > 0:
            total_outstanding += outstanding_balance

        payment_schedule.append({
            **row,
            'amount_paid': amount_paid,
            'outstanding_balance': outstanding_balance,
            'payment_status': payment_status,
        })

    total_outstanding = money(
        total_outstanding
    )

    eligible_members = len([
        row
        for row in payment_schedule
        if money(row['net_shareout']) > 0
    ])

    payment_progress = (
        0
        if eligible_members == 0
        else round(
            (members_paid / eligible_members) * 100
        )
    )

    amount_progress = (
        0
        if shareout_data['total_net_payable'] <= 0
        else round(
            (
                total_paid
                / shareout_data['total_net_payable']
            ) * 100
        )
    )

    amount_progress = max(
        0,
        min(amount_progress, 100)
    )

    bank_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .filter(
            Distribution.method == 'Bank Transfer'
        )
        .scalar()
    )

    mobile_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .filter(
            Distribution.method.in_([
                'Mobile Money',
                'Airtel Money',
                'MTN Money',
            ])
        )
        .scalar()
    )

    cash_paid = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .filter(
            Distribution.method == 'Cash'
        )
        .scalar()
    )

    missing_references = payment_query.filter(
        db.or_(
            Distribution.reference.is_(None),
            Distribution.reference == '',
        )
    ).count()

    recent_payments = payment_query.order_by(
        Distribution.paid_on.desc(),
        Distribution.id.desc(),
    ).limit(8).all()

    highest_allocations = sorted(
        payment_schedule,
        key=lambda row: money(
            row['net_shareout']
        ),
        reverse=True,
    )[:5]

    largest_outstanding = sorted(
        [
            row
            for row in payment_schedule
            if money(
                row['outstanding_balance']
            ) > 0
        ],
        key=lambda row: money(
            row['outstanding_balance']
        ),
        reverse=True,
    )[:5]

    review_members = [
        row
        for row in payment_schedule
        if (
            row['payment_status'] in [
                'Review',
                'Overpaid',
            ]
            or money(row['net_shareout']) < 0
        )
    ]

    if cycle_status == 'Locked':
        executive_status = (
            'The Share-Out cycle is complete, approved '
            'and locked.'
        )

    elif cycle_status == 'Approved' and total_outstanding > 0:
        executive_status = (
            'The Share-Out cycle is approved. Continue '
            'processing outstanding member payments.'
        )

    elif cycle_status == 'Approved':
        executive_status = (
            'All payments appear complete. The cycle is '
            'ready for final locking.'
        )

    elif shareout_data['readiness_score'] >= 75:
        executive_status = (
            'The Share-Out calculation is ready for '
            'committee review and approval.'
        )

    else:
        executive_status = (
            'The Share-Out cycle requires review before '
            'approval or distribution.'
        )

    setting = SystemSetting.query.first()

    organization_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    return render_template(
        'shareout/dashboard.html',
        organization_name=organization_name,
        shareout_data=shareout_data,
        payment_schedule=payment_schedule,
        cycle=cycle,
        cycle_status=cycle_status,
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
        total_paid=total_paid,
        total_outstanding=total_outstanding,
        eligible_members=eligible_members,
        members_paid=members_paid,
        members_partially_paid=members_partially_paid,
        members_pending=members_pending,
        members_overpaid=members_overpaid,
        payment_progress=payment_progress,
        amount_progress=amount_progress,
        bank_paid=bank_paid,
        mobile_paid=mobile_paid,
        cash_paid=cash_paid,
        missing_references=missing_references,
        recent_payments=recent_payments,
        highest_allocations=highest_allocations,
        largest_outstanding=largest_outstanding,
        review_members=review_members,
        executive_status=executive_status,
    )

@app.route('/share-out-statement/<int:member_id>')
@login_required
@role_required('shareout')
def shareout_statement(member_id):
    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.args.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.args.get('end_month')
        or today_month
    )

    expenses = money(
        request.args.get('expenses') or 0
    )

    other_income = money(
        request.args.get('other_income') or 0
    )

    member = Member.query.get_or_404(member_id)

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    member_shareout = next(
        (
            row
            for row in shareout_data['rows']
            if row['member_id'] == member_id
        ),
        None
    )

    if not member_shareout:
        flash(
            'No Share-Out calculation was found for this member '
            'during the selected period.',
            'error'
        )

        return redirect(
            url_for(
                'shareout_schedule',
                start_month=start_month,
                end_month=end_month,
                expenses=expenses,
                other_income=other_income,
            )
        )

    start_date = datetime.strptime(
        start_month + '-01',
        '%Y-%m-%d'
    ).date()

    end_year, end_mon = [
        int(value)
        for value in end_month.split('-')
    ]

    end_date = (
        date(end_year, end_mon, 28)
        + timedelta(days=4)
    ).replace(day=1) - timedelta(days=1)

    payments = Distribution.query.filter(
        Distribution.member_id == member_id,
        Distribution.paid_on >= start_date,
        Distribution.paid_on <= end_date,
    ).order_by(
        Distribution.paid_on.asc(),
        Distribution.id.asc(),
    ).all()

    total_paid = money(
        sum(
            (
                money(payment.amount)
                for payment in payments
            ),
            Decimal('0.00')
        )
    )

    outstanding_balance = money(
        member_shareout['net_shareout'] - total_paid
    )

    if outstanding_balance < 0:
        payment_status = 'Overpaid'
    elif total_paid <= 0:
        payment_status = 'Pending'
    elif outstanding_balance > 0:
        payment_status = 'Partially Paid'
    else:
        payment_status = 'Paid'

    cycle = get_shareout_cycle(
        start_month,
        end_month,
    )

    cycle_status = (
        cycle.status
        if cycle
        else 'Draft'
    )

    setting = SystemSetting.query.first()

    organization_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    organization_address = (
        setting.organization_address
        if setting and setting.organization_address
        else ''
    )

    organization_phone = (
        setting.organization_phone
        if setting and setting.organization_phone
        else ''
    )

    organization_email = (
        setting.organization_email
        if setting and setting.organization_email
        else ''
    )

    registration_number = (
        setting.registration_number
        if setting and setting.registration_number
        else ''
    )

    return render_template(
        'shareout/statement.html',
        member=member,
        member_shareout=member_shareout,
        payments=payments,
        total_paid=total_paid,
        outstanding_balance=outstanding_balance,
        payment_status=payment_status,
        cycle=cycle,
        cycle_status=cycle_status,
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
        organization_name=organization_name,
        organization_address=organization_address,
        organization_phone=organization_phone,
        organization_email=organization_email,
        registration_number=registration_number,
    )    


@app.route('/fines', methods=['GET','POST'])
@login_required
@role_required('fines')
def fines():
    if request.method == 'POST':
        user = session.get('user') or {}
        fp = FinePenalty(
            member_id=int(request.form['member_id']),
            category=request.form['category'],
            amount=money(request.form['amount']),
            reason=request.form.get('reason'),
            fine_date=parse_date(request.form.get('fine_date')),
            recorded_by=user.get('full_name') or user.get('username')
        )
        db.session.add(fp)
        db.session.flush()

        post_to_cash_book(
            entry_date=fp.paid_on,
            entry_type='In',
            category='Fine Payment',
            amount=fp.amount,
            description=f'{fp.member.member_no} - {fp.member.full_name}',
            method=fp.method,
            reference=fp.reference,
            source_type='FinePayment',
            source_id=fp.id
        )

        db.session.commit()
        log_audit('RECORD_FINE', 'FinePenalty', f.id, f'{f.member.full_name} fined {kwacha(f.amount)} for {f.category}')
        flash('Fine / penalty recorded successfully.')
        return redirect(url_for('fines'))
    q = request.args.get('q','').strip()
    status = request.args.get('status','').strip()
    query = FinePenalty.query.join(Member)
    if q:
        like = f'%{q}%'
        query = query.filter(db.or_(Member.full_name.like(like), Member.member_no.like(like), FinePenalty.category.like(like), FinePenalty.reason.like(like)))
    if status:
        query = query.filter(FinePenalty.status == status)
    all_matching_fines = query.order_by(
    FinePenalty.fine_date.desc(),
    FinePenalty.id.desc()
        ).all()

    total_fines = money(sum((f.amount for f in all_matching_fines), Decimal('0.00')))
    total_paid = money(sum((f.total_paid for f in all_matching_fines), Decimal('0.00')))
    total_balance = money(sum((f.balance for f in all_matching_fines if f.status != 'Waived'), Decimal('0.00')))

    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = query.order_by(
        FinePenalty.fine_date.desc(),
        FinePenalty.id.desc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    return render_template(
        'fines.html',
        fines=pagination.items,
        pagination=pagination,
        members=Member.query.order_by(Member.full_name).all(),
        q=q,
        status=status,
        total_fines=total_fines,
        total_paid=total_paid,
        total_balance=total_balance
    )

@app.route('/fines/<int:fine_id>/pay', methods=['POST'])
@login_required
@role_required('fines')
def fine_payment(fine_id):
    fine = FinePenalty.query.get_or_404(fine_id)

    if fine.status == 'Waived':
        flash('Cannot pay a waived fine.', 'error')
        return redirect(url_for('fines'))

    payment = FinePayment(
        fine_id=fine.id,
        amount=money(request.form['amount']),
        method=request.form['method'],
        reference=request.form.get('reference'),
        paid_on=parse_date(request.form.get('paid_on'))
    )

    if payment.amount <= 0:
        flash('Payment amount must be greater than zero.', 'error')
        return redirect(url_for('fines'))

    if payment.amount > fine.balance:
        flash('Payment cannot exceed the outstanding fine balance.', 'error')
        return redirect(url_for('fines'))

    db.session.add(payment)
    db.session.flush()

    post_to_cash_book(
        entry_date=payment.paid_on,
        entry_type='In',
        category='Fine Payment',
        amount=payment.amount,
        description=f'{fine.member.member_no} - {fine.member.full_name}',
        method=payment.method,
        reference=payment.reference,
        source_type='FinePayment',
        source_id=payment.id
    )
    post_journal(
        entry_date=payment.paid_on,
        description=f'Fine payment - {fine.member.member_no} - {fine.member.full_name}',
        debit_account_code=cash_account(payment.method),
        credit_account_code='4010',
        amount=payment.amount,
        source_type='FinePayment',
        source_id=payment.id
    )

    db.session.commit()

    fine.status = 'Paid' if fine.balance <= 0 else 'Partially Paid'
    db.session.commit()

    log_audit(
        'RECORD_FINE_PAYMENT',
        'FinePayment',
        payment.id,
        f'{fine.member.full_name} paid {kwacha(payment.amount)} for fine #{fine.id} via {payment.method}'
    )

    flash('Fine payment recorded.')
    return redirect(url_for('fines'))

@app.route('/fines/<int:fine_id>/waive', methods=['POST'])
@login_required
@role_required('fines')
def fine_waive(fine_id):
    fine = FinePenalty.query.get_or_404(fine_id)
    user = session.get('user') or {}
    if fine.status == 'Paid':
        flash('Paid fines cannot be waived.', 'error')
    else:
        fine.status = 'Waived'
        fine.waived_by = user.get('full_name') or user.get('username')
        fine.waived_on = date.today()
        db.session.commit()
        log_audit('WAIVE_FINE', 'FinePenalty', fine.id, f'Fine #{fine.id} for {fine.member.full_name} waived by {fine.waived_by}. Reason: {request.form.get("waiver_reason") or "Not specified"}')
        flash('Fine waived successfully.')
    return redirect(url_for('fines'))

@app.route('/export/fines.csv')
@login_required
@role_required('fines')
def export_fines_csv():
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Date','Member No','Full Name','Category','Reason','Fine Amount','Paid','Balance','Status','Recorded By'])
    for f in FinePenalty.query.order_by(FinePenalty.fine_date.desc()).all():
        writer.writerow([f.fine_date, f.member.member_no, f.member.full_name, f.category, f.reason, f.amount, f.total_paid, f.balance, f.status, f.recorded_by])
    log_audit('EXPORT_FINES', 'FinePenalty', None, 'Fines and penalties exported to CSV')
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=fines_penalties.csv'})


@app.route('/welfare', methods=['GET','POST'])
@login_required
@role_required('welfare')
def welfare():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'contribution':
            c = WelfareContribution(
                member_id=int(request.form['member_id']),
                month=request.form['month'],
                amount=money(request.form['amount']),
                method=request.form['method'],
                reference=request.form.get('reference'),
                paid_on=parse_date(request.form.get('paid_on'))
            )
            db.session.add(c)
            db.session.flush()

            post_to_cash_book(
                entry_date=c.paid_on,
                entry_type='In',
                category='Welfare Contribution',
                amount=c.amount,
                description=f'{c.member.member_no} - {c.member.full_name}',
                method=c.method,
                reference=c.reference,
                source_type='WelfareContribution',
                source_id=c.id
            )
            post_journal(
                entry_date=c.paid_on,
                description=f'Welfare contribution - {c.member.member_no} - {c.member.full_name}',
                debit_account_code=cash_account(c.method),
                credit_account_code='2010',
                amount=c.amount,
                source_type='WelfareContribution',
                source_id=c.id
            )

            db.session.commit()

            log_audit(
                'RECORD_WELFARE_CONTRIBUTION',
                'WelfareContribution',
                c.id,
                f'{c.member.full_name} paid welfare contribution {kwacha(c.amount)} for {c.month}'
            )

            flash('Welfare contribution recorded.')
        elif action == 'claim':
            claim = WelfareClaim(
                member_id=int(request.form['member_id']),
                category=request.form['category'],
                amount_requested=money(request.form['amount_requested']),
                reason=request.form.get('reason'),
                requested_on=parse_date(request.form.get('requested_on'))
            )
            db.session.add(claim); db.session.commit()
            log_audit('CREATE_WELFARE_CLAIM', 'WelfareClaim', claim.id, f'{claim.member.full_name} requested {kwacha(claim.amount_requested)} for {claim.category}')
            flash('Welfare claim request recorded.')
        return redirect(url_for('welfare'))

    q = request.args.get('q','').strip()
    status = request.args.get('status','').strip()
    claim_query = WelfareClaim.query.join(Member)
    if q:
        like = f'%{q}%'
        claim_query = claim_query.filter(db.or_(Member.full_name.like(like), Member.member_no.like(like), WelfareClaim.category.like(like), WelfareClaim.reason.like(like)))
    if status:
        claim_query = claim_query.filter(WelfareClaim.status == status)
    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = claim_query.order_by(
        WelfareClaim.requested_on.desc(),
        WelfareClaim.id.desc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    claims = pagination.items

    contributions = WelfareContribution.query.order_by(
    WelfareContribution.paid_on.desc(),
    WelfareContribution.id.desc()
    ).limit(100).all()
    members = Member.query.order_by(Member.full_name).all()
    total_contributions = money(db.session.query(db.func.coalesce(db.func.sum(WelfareContribution.amount), 0)).scalar())
    total_paid = money(db.session.query(db.func.coalesce(db.func.sum(WelfareClaim.amount_approved), 0)).filter(WelfareClaim.status == 'Paid').scalar())
    approved_not_paid = money(db.session.query(db.func.coalesce(db.func.sum(WelfareClaim.amount_approved), 0)).filter(WelfareClaim.status == 'Approved').scalar())
    balance = money(total_contributions - total_paid)
    
    pending_claims = WelfareClaim.query.filter_by(status='Pending').count()
    approved_claims = WelfareClaim.query.filter_by(status='Approved').count()
    paid_claims = WelfareClaim.query.filter_by(status='Paid').count()

    this_month = date.today().strftime('%Y-%m')

    this_month_contributions = money(
        db.session.query(db.func.coalesce(db.func.sum(WelfareContribution.amount), 0))
        .filter(WelfareContribution.month == this_month)
        .scalar()
    )

    this_month_claims_paid = money(
        db.session.query(db.func.coalesce(db.func.sum(WelfareClaim.amount_approved), 0))
        .filter(WelfareClaim.status == 'Paid')
        .scalar()
    )
    
    welfare_service = WelfareIntelligenceService(
        balance=balance,
        total_contributions=total_contributions,
        total_paid=total_paid,
        pending_claims=pending_claims,
        approved_claims=approved_claims,
        paid_claims=paid_claims,
        this_month_contributions=this_month_contributions,
        this_month_claims_paid=this_month_claims_paid
    )

    welfare_data = welfare_service.build()
    
    return render_template(
    'welfare.html',
        claims=claims,
        contributions=contributions,
        members=members,
        pagination=pagination,
        q=q,
        status=status,
        total_contributions=total_contributions,
        total_paid=total_paid,
        approved_not_paid=approved_not_paid,
        balance=balance,
        pending_claims=pending_claims,
        approved_claims=approved_claims,
        paid_claims=paid_claims,
        this_month_contributions=this_month_contributions,
        this_month_claims_paid=this_month_claims_paid,

        **welfare_data
    )

@app.route('/welfare/claims/<int:claim_id>/review', methods=['POST'])
@login_required
@role_required('welfare')
def welfare_review(claim_id):
    claim = WelfareClaim.query.get_or_404(claim_id)
    if claim.status != 'Requested':
        flash('Only requested claims can be marked as reviewed.', 'error')
    else:
        user = session.get('user') or {}
        claim.status = 'Reviewed'
        claim.reviewed_by = user.get('full_name') or user.get('username')
        claim.reviewed_on = date.today()
        db.session.commit()
        log_audit('WELFARE_CLAIM_REVIEWED', 'WelfareClaim', claim.id, f'{claim.member.full_name} claim reviewed by {claim.reviewed_by}')
        flash('Welfare claim marked as reviewed.')
    return redirect(url_for('welfare'))

@app.route('/welfare/claims/<int:claim_id>/approve', methods=['POST'])
@login_required
@role_required('welfare')
def welfare_approve(claim_id):
    claim = WelfareClaim.query.get_or_404(claim_id)
    if claim.status not in ['Requested', 'Reviewed']:
        flash('Only requested or reviewed claims can be approved.', 'error')
    else:
        user = session.get('user') or {}
        claim.status = 'Approved'
        claim.amount_approved = money(request.form.get('amount_approved') or claim.amount_requested)
        claim.approved_by = request.form.get('approved_by') or user.get('full_name') or 'Management Committee'
        claim.approved_on = date.today()
        db.session.commit()
        log_audit('WELFARE_CLAIM_APPROVED', 'WelfareClaim', claim.id, f'{claim.member.full_name} welfare claim approved for {kwacha(claim.amount_approved)} by {claim.approved_by}')
        flash('Welfare claim approved. It is ready for payment.')
    return redirect(url_for('welfare'))

@app.route('/welfare/claims/<int:claim_id>/pay', methods=['POST'])
@login_required
@role_required('welfare')
def welfare_pay(claim_id):
    claim = WelfareClaim.query.get_or_404(claim_id)
    if claim.status != 'Approved':
        flash('Only approved welfare claims can be paid.', 'error')
    else:
        user = session.get('user') or {}
        claim.status = 'Paid'
        claim.paid_by = user.get('full_name') or user.get('username')
        claim.paid_on = parse_date(request.form.get('paid_on'))
        claim.payment_method = request.form.get('method')
        claim.reference = request.form.get('reference')

        post_to_cash_book(
            entry_date=claim.paid_on,
            entry_type='Out',
            category='Welfare Claim Payment',
            amount=claim.amount_approved,
            description=f'{claim.member.member_no} - {claim.member.full_name}',
            method=claim.payment_method,
            reference=claim.reference,
            source_type='WelfareClaim',
            source_id=claim.id
        )
        post_journal(
            entry_date=claim.paid_on,
            description=f'Welfare claim payment - {claim.member.member_no} - {claim.member.full_name}',
            debit_account_code='5030',
            credit_account_code=cash_account(claim.payment_method),
            amount=claim.amount_approved,
            source_type='WelfareClaim',
            source_id=claim.id
        )

        db.session.commit()
        log_audit('WELFARE_CLAIM_PAID', 'WelfareClaim', claim.id, f'{claim.member.full_name} welfare claim paid {kwacha(claim.amount_approved)} via {claim.payment_method}')
        flash('Welfare claim paid successfully.')
    return redirect(url_for('welfare'))

@app.route('/welfare/claims/<int:claim_id>/reject', methods=['POST'])
@login_required
@role_required('welfare')
def welfare_reject(claim_id):
    claim = WelfareClaim.query.get_or_404(claim_id)
    if claim.status in ['Paid']:
        flash('Paid welfare claims cannot be rejected.', 'error')
    else:
        claim.status = 'Rejected'
        claim.rejection_reason = request.form.get('reason') or 'Not specified'
        db.session.commit()
        log_audit('WELFARE_CLAIM_REJECTED', 'WelfareClaim', claim.id, f'{claim.member.full_name} welfare claim rejected. Reason: {claim.rejection_reason}')
        flash('Welfare claim rejected.')
    return redirect(url_for('welfare'))

@app.route('/export/welfare.csv')
@login_required
@role_required('welfare')
def export_welfare_csv():
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Type','Date','Member No','Full Name','Category/Month','Requested/Contribution','Approved','Status','Method','Reference'])
    for c in WelfareContribution.query.order_by(WelfareContribution.paid_on.desc()).all():
        writer.writerow(['Contribution', c.paid_on, c.member.member_no, c.member.full_name, c.month, c.amount, '', 'Paid', c.method, c.reference])
    for cl in WelfareClaim.query.order_by(WelfareClaim.requested_on.desc()).all():
        writer.writerow(['Claim', cl.requested_on, cl.member.member_no, cl.member.full_name, cl.category, cl.amount_requested, cl.amount_approved, cl.status, cl.payment_method or '', cl.reference or ''])
    log_audit('EXPORT_WELFARE', 'Welfare', None, 'Welfare fund records exported to CSV')
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=welfare_fund.csv'})

@app.route('/meetings', methods=['GET','POST'])
@login_required
@role_required('meetings')
def meetings():
    if request.method == 'POST':
        mtg = Meeting(meeting_type=request.form['meeting_type'], meeting_date=parse_date(request.form['meeting_date']), agenda=request.form.get('agenda'), resolutions=request.form.get('resolutions'), attendance_count=int(request.form.get('attendance_count') or 0))
        db.session.add(mtg); db.session.commit(); log_audit('SAVE_MEETING', 'Meeting', mtg.id, f'{mtg.meeting_type} meeting on {mtg.meeting_date}; attendance {mtg.attendance_count}'); flash('Meeting record saved.'); return redirect(url_for('meetings'))
    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = Meeting.query.order_by(
    Meeting.meeting_date.desc(),
    Meeting.id.desc()
    ).paginate(
    page=page,
    per_page=per_page,
    error_out=False
    )

    today = date.today()

    total_meetings = Meeting.query.count()

    next_meeting = Meeting.query.filter(
        Meeting.meeting_date >= today
    ).order_by(
        Meeting.meeting_date.asc()
    ).first()

    meetings_this_year = Meeting.query.filter(
        db.extract('year', Meeting.meeting_date) == today.year
    ).count()

    average_attendance = db.session.query(
        db.func.coalesce(db.func.avg(Meeting.attendance_count), 0)
    ).scalar()

    latest_meeting = Meeting.query.order_by(
        Meeting.meeting_date.desc(),
        Meeting.id.desc()
    ).first()

    today = date.today()

    next_meeting = Meeting.query.filter(
        Meeting.meeting_date >= today
    ).order_by(
        Meeting.meeting_date.asc()
    ).first()

    meetings_this_year = Meeting.query.filter(
        db.extract('year', Meeting.meeting_date) == today.year
    ).count()

    total_meetings = Meeting.query.count()

    average_attendance = db.session.query(
        db.func.coalesce(db.func.avg(Meeting.attendance_count), 0)
    ).scalar()

    pending_resolutions = Meeting.query.filter(
        Meeting.resolutions != None,
        Meeting.resolutions != ''
    ).count()

    days_to_next = None

    if next_meeting:
        days_to_next = (
            next_meeting.meeting_date - today
        ).days

    meeting_service = MeetingIntelligenceService(
    next_meeting=next_meeting,
    days_to_next=days_to_next,
    average_attendance=average_attendance,
    pending_resolutions=pending_resolutions,
    meetings_this_year=meetings_this_year
)

    meeting_data = meeting_service.build()

    return render_template(
    'meetings.html',
    meetings=pagination.items,
    pagination=pagination,

    next_meeting=next_meeting,
    days_to_next=days_to_next,
    meetings_this_year=meetings_this_year,
    total_meetings=total_meetings,
    average_attendance=round(average_attendance or 0),
    pending_resolutions=pending_resolutions
)


@app.route('/attendance')
@login_required
@role_required('attendance')
def attendance():
    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = Meeting.query.order_by(
        Meeting.meeting_date.desc(),
        Meeting.id.desc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    meetings_list = pagination.items
    summary = []
    total_members = Member.query.filter_by(status='Active').count()

    for mtg in meetings_list:
        records = MeetingAttendance.query.filter_by(meeting_id=mtg.id).all()
        present = sum(1 for r in records if r.status == 'Present')
        late = sum(1 for r in records if r.status == 'Late')
        absent = sum(1 for r in records if r.status == 'Absent')
        excused = sum(1 for r in records if r.status == 'Excused')

        summary.append({
            'meeting': mtg,
            'total': len(records),
            'present': present,
            'late': late,
            'absent': absent,
            'excused': excused,
            'expected': total_members
        })

    return render_template(
        'attendance.html',
        summary=summary,
        pagination=pagination
    )

@app.route('/attendance/<int:meeting_id>', methods=['GET', 'POST'])
@login_required
@role_required('attendance')
def attendance_register(meeting_id):
    meeting = Meeting.query.get_or_404(meeting_id)
    members = Member.query.filter_by(status='Active').order_by(Member.member_no).all()
    user = session.get('user') or {}
    if request.method == 'POST':
        absent_fine = Decimal('50.00')
        late_fine = Decimal('20.00')
        create_fines = True
        saved = 0
        fines_created = 0
        for member in members:
            status = request.form.get(f'status_{member.id}') or 'Present'
            remarks = request.form.get(f'remarks_{member.id}', '').strip()
            record = MeetingAttendance.query.filter_by(meeting_id=meeting.id, member_id=member.id).first()
            if not record:
                record = MeetingAttendance(meeting_id=meeting.id, member_id=member.id)
                db.session.add(record)
            record.status = status
            record.remarks = remarks
            record.recorded_by = user.get('full_name') or user.get('username')
            record.recorded_at = datetime.utcnow()
            if create_fines and not record.fine_generated:
                fine_amount = Decimal('0.00')
                category = None
                reason = None
                if status == 'Absent' and absent_fine > 0:
                    fine_amount = absent_fine
                    category = 'Absence from Meeting'
                    reason = f'Absent from {meeting.meeting_type} meeting on {meeting.meeting_date}'
                elif status == 'Late' and late_fine > 0:
                    fine_amount = late_fine
                    category = 'Late Meeting Attendance'
                    reason = f'Late for {meeting.meeting_type} meeting on {meeting.meeting_date}'
                if fine_amount > 0:
                    fine = FinePenalty(member_id=member.id, category=category, amount=fine_amount, reason=reason, fine_date=meeting.meeting_date, recorded_by=record.recorded_by)
                    db.session.add(fine)
                    record.fine_generated = True
                    fines_created += 1
            saved += 1
        meeting.attendance_count = sum(1 for member in members if (request.form.get(f'status_{member.id}') or 'Present') in ['Present', 'Late'])
        db.session.commit()
        log_audit('SAVE_ATTENDANCE', 'Meeting', meeting.id, f'Attendance saved for {meeting.meeting_type} on {meeting.meeting_date}. Records: {saved}; fines created: {fines_created}')
        flash(f'Attendance register saved. {saved} records updated. {fines_created} fines generated.')
        return redirect(url_for('attendance_register', meeting_id=meeting.id))

    existing = {r.member_id: r for r in MeetingAttendance.query.filter_by(meeting_id=meeting.id).all()}
    stats = {'Present':0, 'Absent':0, 'Late':0, 'Excused':0}
    for r in existing.values():
        stats[r.status] = stats.get(r.status, 0) + 1
    return render_template('attendance_register.html', meeting=meeting, members=members, existing=existing, stats=stats)

@app.route('/export/attendance.csv')
@login_required
@role_required('attendance')
def export_attendance_csv():
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Meeting Date','Meeting Type','Member No','Full Name','Status','Remarks','Recorded By','Recorded At','Fine Generated'])
    rows = MeetingAttendance.query.join(Meeting).join(Member).order_by(Meeting.meeting_date.desc(), Member.member_no).all()
    for r in rows:
        writer.writerow([r.meeting.meeting_date, r.meeting.meeting_type, r.member.member_no, r.member.full_name, r.status, r.remarks or '', r.recorded_by or '', r.recorded_at, 'Yes' if r.fine_generated else 'No'])
    log_audit('EXPORT_ATTENDANCE', 'MeetingAttendance', None, 'Attendance register exported to CSV')
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=meeting_attendance.csv'})

@app.route('/reports')
@login_required
@role_required('reports')
def reports():
    month = request.args.get('month') or date.today().strftime('%Y-%m')

    contribs = Contribution.query.filter_by(month=month).all()
    paid_member_ids = {c.member_id for c in contribs}

    arrears_query = Member.query.filter(
        Member.id.notin_(paid_member_ids)
    ) if paid_member_ids else Member.query

    page = request.args.get('page', 1, type=int)
    per_page = 25

    arrears_pagination = arrears_query.order_by(
        Member.member_no
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    open_loans = Loan.query.filter(
        Loan.status.in_(['Disbursed', 'Partially Paid'])
    ).order_by(
        Loan.issued_on.desc(),
        Loan.id.desc()
    ).all()

    return render_template(
        'reports.html',
        month=month,
        contribs=contribs,
        arrears=arrears_pagination.items,
        arrears_pagination=arrears_pagination,
        open_loans=open_loans
    )

@app.route('/opening-balances')
@login_required
@role_required('accounting')
def opening_balances():

    batches = OpeningBalanceBatch.query.order_by(
        OpeningBalanceBatch.imported_on.desc(),
        OpeningBalanceBatch.id.desc()
    ).all()

    total_batches = OpeningBalanceBatch.query.count()

    posted_batches = OpeningBalanceBatch.query.filter(
        OpeningBalanceBatch.status.in_([
            'Posted',
            'Locked'
        ])
    ).count()

    draft_batches = OpeningBalanceBatch.query.filter(
        OpeningBalanceBatch.status.in_([
            'Draft',
            'Validated',
            'Approved'
        ])
    ).count()

    reversed_batches = OpeningBalanceBatch.query.filter_by(
        status='Reversed'
    ).count()

    total_members_imported = db.session.query(
        db.func.coalesce(
            db.func.sum(OpeningBalanceBatch.total_members),
            0
        )
    ).filter(
        OpeningBalanceBatch.status.in_([
            'Posted',
            'Locked'
        ])
    ).scalar()

    return render_template(
        "opening_balances/control_centre.html",
        batches=batches,
        total_batches=total_batches,
        posted_batches=posted_batches,
        draft_batches=draft_batches,
        reversed_batches=reversed_batches,
        total_members_imported=total_members_imported
    )



@app.route("/opening-balances/import", methods=["GET", "POST"])
@role_required("accounting")
def opening_balance_import():

    if request.method == "POST":

        uploaded_file = request.files.get("opening_balance_file")
        effective_date_text = request.form.get(
            "effective_date",
            ""
        ).strip()
        description = request.form.get(
            "description",
            ""
        ).strip()

        if not uploaded_file or not uploaded_file.filename:
            flash(
                "Please select an Excel or CSV file.",
                "danger"
            )
            return redirect(
                url_for("opening_balance_import")
            )

        try:
            effective_date = datetime.strptime(
                effective_date_text,
                "%Y-%m-%d"
            ).date()
        except ValueError:
            flash(
                "Please enter a valid effective date.",
                "danger"
            )
            return redirect(
                url_for("opening_balance_import")
            )

        try:
            imported_rows = read_opening_balance_file(
                uploaded_file
            )

            required_columns = {
                "member_no",
                "savings_balance",
                "loan_principal",
                "loan_interest",
                "fine_balance",
                "welfare_balance"
            }

            imported_columns = {
                key
                for row in imported_rows
                for key in row.keys()
            }

            missing_columns = (
                required_columns - imported_columns
            )

            if missing_columns:
                raise ValueError(
                    "Missing required columns: "
                    + ", ".join(sorted(missing_columns))
                )
            
            batch = OpeningBalanceBatch(
                    batch_no=generate_opening_balance_batch_no(),
                    effective_date=effective_date,
                    description=description,
                    status="Draft",
                    total_savings=Decimal("0.00"),
                    total_loan_principal=Decimal("0.00"),
                    total_loan_interest=Decimal("0.00"),
                    total_fines=Decimal("0.00"),
                    total_welfare=Decimal("0.00"),
                    imported_by=(
                        (session.get('user') or {}).get('full_name')
                        or (session.get('user') or {}).get('username')
                        or 'System'
                    )
                )

            db.session.add(batch)
            db.session.flush()

            seen_member_numbers = set()
            imported_members = 0

            total_savings = Decimal("0.00")
            total_loan_principal = Decimal("0.00")
            total_loan_interest = Decimal("0.00")
            total_fines = Decimal("0.00")
            total_welfare = Decimal("0.00")

            validation_errors = []

            for row in imported_rows:

                row_number = row.get("_row_number")
                member_no = str(
                    row.get("member_no") or ""
                ).strip()

                if not member_no:
                    validation_errors.append(
                        f"Row {row_number}: Member number is missing."
                    )
                    continue

                if member_no in seen_member_numbers:
                    validation_errors.append(
                        f"Row {row_number}: Duplicate member "
                        f"number {member_no}."
                    )
                    continue

                seen_member_numbers.add(member_no)

                member = Member.query.filter_by(
                    member_no=member_no
                ).first()

                if not member:
                    validation_errors.append(
                        f"Row {row_number}: Member {member_no} "
                        f"does not exist."
                    )
                    continue

                try:
                    savings_balance = (
                        opening_balance_decimal(
                            row.get("savings_balance")
                        )
                    )

                    loan_principal = (
                        opening_balance_decimal(
                            row.get("loan_principal")
                        )
                    )

                    loan_interest = (
                        opening_balance_decimal(
                            row.get("loan_interest")
                        )
                    )

                    fine_balance = (
                        opening_balance_decimal(
                            row.get("fine_balance")
                        )
                    )

                    welfare_balance = (
                        opening_balance_decimal(
                            row.get("welfare_balance")
                        )
                    )

                except ValueError as error:
                    validation_errors.append(
                        f"Row {row_number}: {error}"
                    )
                    continue

                loan_due_date = None
                loan_due_date_value = row.get(
                    "loan_due_date"
                )

                if loan_due_date_value:
                    if isinstance(
                        loan_due_date_value,
                        datetime
                    ):
                        loan_due_date = (
                            loan_due_date_value.date()
                        )

                    elif isinstance(
                        loan_due_date_value,
                        date
                    ):
                        loan_due_date = loan_due_date_value

                    else:
                        try:
                            loan_due_date = datetime.strptime(
                                str(
                                    loan_due_date_value
                                ).strip(),
                                "%Y-%m-%d"
                            ).date()
                        except ValueError:
                            validation_errors.append(
                                f"Row {row_number}: Invalid loan "
                                f"due date. Use YYYY-MM-DD."
                            )
                            continue

                opening_balance = OpeningBalance(
                    batch_id=batch.id,
                    member_id=member.id,
                    savings_balance=savings_balance,
                    loan_principal=loan_principal,
                    loan_interest=loan_interest,
                    fine_balance=fine_balance,
                    welfare_balance=welfare_balance,
                    loan_due_date=loan_due_date,
                    remarks=str(
                        row.get("remarks") or ""
                    ).strip(),
                    validation_status="Valid",
                    validation_message="Imported successfully",
                    is_posted=False
                )

                db.session.add(opening_balance)

                imported_members += 1
                total_savings += savings_balance
                total_loan_principal += loan_principal
                total_loan_interest += loan_interest
                total_fines += fine_balance
                total_welfare += welfare_balance

            if imported_members == 0:
                error_details = "\n".join(validation_errors[:10])

                raise ValueError(
                "No valid member balances were imported.\n"
                    + (
                        error_details
                        if error_details
                        else "No specific validation error was recorded."
                    )
                )

            batch.total_members = imported_members
            batch.total_savings = total_savings
            batch.total_loan_principal = (
                total_loan_principal
            )
            batch.total_loan_interest = (
                total_loan_interest
            )
            batch.total_fines = total_fines
            batch.total_welfare = total_welfare

            if validation_errors:
                batch.status = "Draft"
                batch.description = (
                    f"{batch.description} — "
                    f"{len(validation_errors)} row(s) rejected"
                )
            else:
                batch.status = "Validated"

            db.session.commit()

            log_audit(
                "IMPORT_OPENING_BALANCES",
                (
                    f"Imported opening balance batch "
                    f"{batch.batch_no} for "
                    f"{imported_members} member(s)"
                )
            )

            if validation_errors:
                flash(
                    (
                        f"Batch {batch.batch_no} imported with "
                        f"{len(validation_errors)} rejected row(s)."
                    ),
                    "warning"
                )
            else:
                flash(
                    (
                        f"Batch {batch.batch_no} imported "
                        f"successfully."
                    ),
                    "success"
                )

            return redirect(
                url_for(
                    "opening_balance_preview",
                    batch_id=batch.id
                )
            )

        except Exception as error:
            db.session.rollback()

            flash(
                f"Opening balance import failed: {error}",
                "danger"
            )

            return redirect(
                url_for("opening_balance_import")
            )

    return render_template(
        "opening_balances/import.html"
    )

@app.route("/opening-balances/template")
@role_required("accounting")
def opening_balance_download_template():

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Opening Balances"

    # ---------------------------------------------------------
    # Workbook colours and styling
    # ---------------------------------------------------------
    dark_blue = "1F4E78"
    light_blue = "D9EAF7"
    light_yellow = "FFF2CC"
    light_green = "E2F0D9"
    white = "FFFFFF"
    grey = "666666"

    thin_border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7")
    )

    # ---------------------------------------------------------
    # Title
    # ---------------------------------------------------------
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "OPENING BALANCES IMPORT TEMPLATE"

    worksheet["A1"].font = Font(
        bold=True,
        size=16,
        color=white
    )

    worksheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=dark_blue
    )

    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet.row_dimensions[1].height = 28

    # ---------------------------------------------------------
    # Instructions
    # ---------------------------------------------------------
    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = (
        "Enter one member per row. Do not change the column headings."
    )

    worksheet["A2"].font = Font(
        italic=True,
        color=grey
    )

    worksheet["A2"].fill = PatternFill(
        fill_type="solid",
        fgColor=light_blue
    )

    worksheet["A2"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    worksheet.merge_cells("A3:H3")
    worksheet["A3"] = (
        "Member numbers must already exist in SL Village Banking Pro. "
        "Amounts must not be negative."
    )

    worksheet["A3"].font = Font(
        italic=True,
        color=grey
    )

    worksheet["A3"].fill = PatternFill(
        fill_type="solid",
        fgColor=light_blue
    )

    worksheet["A3"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # ---------------------------------------------------------
    # Required headings
    # ---------------------------------------------------------
    headings = [
        "member_no",
        "savings_balance",
        "loan_principal",
        "loan_interest",
        "fine_balance",
        "welfare_balance",
        "loan_due_date",
        "remarks"
    ]

    heading_row = 5

    for column_number, heading in enumerate(
        headings,
        start=1
    ):
        cell = worksheet.cell(
            row=heading_row,
            column=column_number,
            value=heading
        )

        cell.font = Font(
            bold=True,
            color=white
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=dark_blue
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = thin_border

    worksheet.row_dimensions[heading_row].height = 32

    # ---------------------------------------------------------
    # Example row
    # ---------------------------------------------------------
    example_values = [
        "M001",
        5000.00,
        1500.00,
        225.00,
        0.00,
        200.00,
        "2026-12-31",
        "Example only - delete before upload"
    ]

    example_row = 6

    for column_number, value in enumerate(
        example_values,
        start=1
    ):
        cell = worksheet.cell(
            row=example_row,
            column=column_number,
            value=value
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=light_yellow
        )

        cell.border = thin_border

    # ---------------------------------------------------------
    # Blank input rows
    # ---------------------------------------------------------
    for row_number in range(7, 207):

        for column_number in range(1, 9):
            cell = worksheet.cell(
                row=row_number,
                column=column_number
            )

            cell.border = thin_border

            if column_number in [2, 3, 4, 5, 6]:
                cell.number_format = '#,##0.00'

            if column_number == 7:
                cell.number_format = "yyyy-mm-dd"

        if row_number % 2 == 1:
            for column_number in range(1, 9):
                worksheet.cell(
                    row=row_number,
                    column=column_number
                ).fill = PatternFill(
                    fill_type="solid",
                    fgColor=light_green
                )

    # ---------------------------------------------------------
    # Column widths
    # ---------------------------------------------------------
    column_widths = {
        "A": 18,
        "B": 20,
        "C": 20,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 38
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width

    # Keep headings visible
    worksheet.freeze_panes = "A6"

    # Add filters
    worksheet.auto_filter.ref = "A5:H206"

    # Protect title and headings from accidental changes
    worksheet.protection.sheet = False

    # ---------------------------------------------------------
    # Instructions worksheet
    # ---------------------------------------------------------
    instructions = workbook.create_sheet(
        "Instructions"
    )

    instructions["A1"] = "OPENING BALANCES IMPORT INSTRUCTIONS"
    instructions["A1"].font = Font(
        bold=True,
        size=15,
        color=white
    )
    instructions["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor=dark_blue
    )
    instructions["A1"].alignment = Alignment(
        horizontal="center"
    )

    instructions.merge_cells("A1:D1")

    instruction_rows = [
        [
            "Column",
            "Required",
            "Description",
            "Example"
        ],
        [
            "member_no",
            "Yes",
            "Existing system member number",
            "M001"
        ],
        [
            "savings_balance",
            "Yes",
            "Member savings balance at the effective date",
            "5000.00"
        ],
        [
            "loan_principal",
            "Yes",
            "Outstanding loan principal",
            "1500.00"
        ],
        [
            "loan_interest",
            "Yes",
            "Outstanding accrued loan interest",
            "225.00"
        ],
        [
            "fine_balance",
            "Yes",
            "Outstanding fines",
            "0.00"
        ],
        [
            "welfare_balance",
            "Yes",
            "Member welfare balance",
            "200.00"
        ],
        [
            "loan_due_date",
            "No",
            "Use YYYY-MM-DD",
            "2026-12-31"
        ],
        [
            "remarks",
            "No",
            "Additional explanation",
            "Existing loan"
        ]
    ]

    for row_number, row_values in enumerate(
        instruction_rows,
        start=3
    ):
        for column_number, value in enumerate(
            row_values,
            start=1
        ):
            cell = instructions.cell(
                row=row_number,
                column=column_number,
                value=value
            )

            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            if row_number == 3:
                cell.font = Font(
                    bold=True,
                    color=white
                )
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=dark_blue
                )

    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 14
    instructions.column_dimensions["C"].width = 52
    instructions.column_dimensions["D"].width = 24

    instructions.freeze_panes = "A4"

    # ---------------------------------------------------------
    # Return workbook as download
    # ---------------------------------------------------------
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=(
            f"opening_balances_template_"
            f"{date.today().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )

@app.route("/opening-balances/<int:batch_id>/preview")
@role_required("accounting")
def opening_balance_preview(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    balances = (
        OpeningBalance.query
        .filter_by(batch_id=batch.id)
        .join(Member)
        .order_by(Member.member_no)
        .all()
    )

    validation_errors = (
        OpeningBalance.query
        .filter(
            OpeningBalance.batch_id == batch.id,
            OpeningBalance.validation_status != "Valid"
        )
        .count()
    )

    return render_template(
        "opening_balances/preview.html",
        batch=batch,
        balances=balances,
        validation_errors=validation_errors
    )

 
@app.route("/opening-balances/<int:batch_id>/validate", methods=["POST"])
@role_required("accounting")
def opening_balance_validate(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    balances = OpeningBalance.query.filter_by(batch_id=batch.id).all()

    errors = 0

    for balance in balances:

        message = []

        if balance.savings_balance is None:
            message.append("Savings balance missing.")

        if balance.loan_principal is None:
            message.append("Loan principal missing.")

        if balance.loan_interest is None:
            message.append("Loan interest missing.")

        if balance.fine_balance is None:
            message.append("Fine balance missing.")

        if balance.welfare_balance is None:
            message.append("Welfare balance missing.")

        if message:
            balance.validation_status = "Invalid"
            balance.validation_message = " ".join(message)
            errors += 1
        else:
            balance.validation_status = "Valid"
            balance.validation_message = ""

    if errors == 0:
        batch.status = "Validated"
        flash("Batch validated successfully.", "success")
    else:
        batch.status = "Draft"
        flash(f"{errors} validation error(s) found.", "warning")

    db.session.commit()

    log_audit(
        "VALIDATE_OPENING_BALANCES",
        f"Validated opening balance batch {batch.batch_no}"
    )

    return redirect(
        url_for(
            "opening_balance_preview",
            batch_id=batch.id
        )
    )

@app.route("/opening-balances/<int:batch_id>")
@role_required("accounting")
def opening_balance_batch(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    balances = (
        OpeningBalance.query
        .filter_by(batch_id=batch.id)
        .join(Member)
        .order_by(Member.member_no)
        .all()
    )

    validation_errors = (
        OpeningBalance.query
        .filter(
            OpeningBalance.batch_id == batch.id,
            OpeningBalance.validation_status != "Valid"
        )
        .count()
    )

    total_records = len(balances)

    valid_records = (
        OpeningBalance.query
        .filter_by(
            batch_id=batch.id,
            validation_status="Valid"
        )
        .count()
    )

    posted_records = (
        OpeningBalance.query
        .filter_by(
            batch_id=batch.id,
            is_posted=True
        )
        .count()
    )

    return render_template(
        "opening_balances/batch.html",
        batch=batch,
        balances=balances,
        validation_errors=validation_errors,
        total_records=total_records,
        valid_records=valid_records,
        posted_records=posted_records
    )

@app.route(
    "/opening-balances/<int:batch_id>/approve",
    methods=["POST"]
)
@role_required("accounting")
def opening_balance_approve(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    if batch.status == "Posted":
        flash(
            "This opening balance batch has already been posted.",
            "warning"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    validation_errors = (
        OpeningBalance.query
        .filter(
            OpeningBalance.batch_id == batch.id,
            OpeningBalance.validation_status != "Valid"
        )
        .count()
    )

    if validation_errors > 0:
        flash(
            (
                f"Batch cannot be approved because "
                f"{validation_errors} validation error(s) remain."
            ),
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    balance_count = (
        OpeningBalance.query
        .filter_by(batch_id=batch.id)
        .count()
    )

    if balance_count == 0:
        flash(
            "Batch cannot be approved because it contains no balances.",
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    batch.status = "Approved"

    db.session.commit()

    log_audit(
        "APPROVE_OPENING_BALANCES",
        (
            f"Approved opening balance batch "
            f"{batch.batch_no} containing "
            f"{balance_count} member balance(s)"
        )
    )

    flash(
        f"Opening balance batch {batch.batch_no} approved successfully.",
        "success"
    )

    return redirect(
        url_for(
            "opening_balance_batch",
            batch_id=batch.id
        )
    )


@app.route(
    "/accounting/financial-year/closing-preview"
)
@login_required
@role_required("accounting")
def financial_year_closing_preview():
    financial_year = current_financial_year()

    if not financial_year:
        flash(
            "No open financial year was found.",
            "danger"
        )
        return redirect(
            url_for("financial_year_control_centre")
        )

    preview = build_financial_year_closing_preview(
        financial_year
    )

    log_audit(
        "PREVIEW_YEAR_END",
        (
            f"Year-end closing preview generated for "
            f"financial year {financial_year.year}; "
            f"current surplus {kwacha(preview['current_surplus'])}."
        )
    )

    return render_template(
        "financial_year/closing_preview.html",
        financial_year=financial_year,
        preview=preview
    )

@app.route(
    "/opening-balances/<int:batch_id>/lock",
    methods=["POST"]
)
@role_required("accounting")
def opening_balance_lock(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    if batch.status != "Posted":
        flash(
            "Only posted opening balance batches can be locked.",
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    if batch.is_locked:
        flash(
            "This opening balance batch is already locked.",
            "warning"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    user = session.get("user") or {}

    locked_by = (
        user.get("full_name")
        or user.get("username")
        or "System"
    )

    batch.is_locked = True
    batch.locked_by = locked_by
    batch.locked_on = datetime.utcnow()

    db.session.commit()

    log_audit(
        "LOCK_OPENING_BALANCE_BATCH",
        (
            f"Locked opening balance batch "
            f"{batch.batch_no}"
        )
    )

    flash(
        (
            f"Opening balance batch "
            f"{batch.batch_no} locked successfully."
        ),
        "success"
    )

    return redirect(
        url_for(
            "opening_balance_batch",
            batch_id=batch.id
        )
    )


@app.route(
    "/accounting/financial-year/close",
    methods=["GET", "POST"]
)
@login_required
@role_required("accounting")
def close_financial_year():
    financial_year = current_financial_year()

    if not financial_year:
        flash("No open financial year was found.", "danger")
        return redirect(
            url_for("financial_year_control_centre")
        )

    preview = build_financial_year_closing_preview(
        financial_year
    )

    required_confirmation = f"CLOSE {financial_year.year}"

    if request.method == "GET":
        return render_template(
            "financial_year/close_confirmation.html",
            financial_year=financial_year,
            preview=preview,
            required_confirmation=required_confirmation
        )

    confirmation_text = (
        request.form.get("confirmation_text") or ""
    ).strip()

    if confirmation_text != required_confirmation:
        flash(
            f'Type "{required_confirmation}" exactly to confirm closure.',
            "danger"
        )

        return render_template(
            "financial_year/close_confirmation.html",
            financial_year=financial_year,
            preview=preview,
            required_confirmation=required_confirmation
        )

    try:
        # Re-read the year to reduce the risk of using stale information.
        financial_year = FinancialYear.query.filter_by(
            id=financial_year.id
        ).first()

        if not financial_year:
            raise ValueError(
                "The financial year could not be found."
            )

        if financial_year.status != "Open":
            raise ValueError(
                f"Financial Year {financial_year.year} is not open."
            )

        if financial_year.is_locked:
            raise ValueError(
                f"Financial Year {financial_year.year} is already locked."
            )

        # Always rerun the checks immediately before closing.
        validation = run_financial_year_preclose_checks(
            financial_year
        )

        if not validation["is_ready"]:
            raise ValueError(
                "The financial year can no longer be closed because "
                "one or more pre-close checks have failed."
            )

        duplicate_closing = JournalEntry.query.filter_by(
            source_type="FinancialYearClosing",
            source_id=str(financial_year.id)
        ).first()

        if duplicate_closing:
            raise ValueError(
                f"Financial Year {financial_year.year} already has "
                f"a closing journal: {duplicate_closing.reference}."
            )

        closing_data = build_financial_year_closing_lines(
            financial_year
        )

        if not closing_data["lines"]:
            raise ValueError(
                "There are no income or expense balances to close."
            )

        closing_reference = (
            generate_financial_year_closing_reference(
                financial_year
            )
        )

        username = (
            (session.get("user") or {}).get("username")
            or "System"
        )

        closing_entry = post_journal(
            entry_date=financial_year.end_date,
            description=(
                f"Financial Year {financial_year.year} "
                f"Closing Journal"
            ),
            reference=closing_reference,
            source_type="FinancialYearClosing",
            source_id=financial_year.id,
            lines=closing_data["lines"],
            commit=False
        )

        if not closing_entry:
            raise ValueError(
                "The closing journal was not created because "
                "a matching year-end journal already exists."
            )

        financial_year.status = "Closed"
        financial_year.is_locked = True
        financial_year.closed_on = datetime.utcnow()
        financial_year.closed_by = username
        financial_year.closing_reference = closing_reference
        financial_year.updated_on = datetime.utcnow()

        next_year_number = financial_year.year + 1

        next_financial_year = FinancialYear.query.filter_by(
            year=next_year_number
        ).first()

        if next_financial_year:
            if next_financial_year.status == "Closed":
                raise ValueError(
                    f"Financial Year {next_year_number} already exists "
                    f"and is closed."
                )

            next_financial_year.status = "Open"
            next_financial_year.is_locked = False

            if not next_financial_year.opened_on:
                next_financial_year.opened_on = datetime.utcnow()

            if not next_financial_year.opened_by:
                next_financial_year.opened_by = username

            next_financial_year.updated_on = datetime.utcnow()

        else:
            next_financial_year = FinancialYear(
                year=next_year_number,
                start_date=date(next_year_number, 1, 1),
                end_date=date(next_year_number, 12, 31),
                status="Open",
                opened_on=datetime.utcnow(),
                opened_by=username,
                is_locked=False,
                created_on=datetime.utcnow(),
                updated_on=datetime.utcnow()
            )

            db.session.add(next_financial_year)

        db.session.commit()

        log_audit(
            "CLOSE_FINANCIAL_YEAR",
            (
                f"Financial Year {financial_year.year} closed; "
                f"reference {closing_reference}; "
                f"surplus/(deficit) "
                f"{kwacha(closing_data['current_surplus'])}; "
                f"Financial Year {next_year_number} opened."
            )
        )

        flash(
            (
                f"Financial Year {financial_year.year} was closed "
                f"successfully. Closing reference: "
                f"{closing_reference}. Financial Year "
                f"{next_year_number} is now open."
            ),
            "success"
        )

        return redirect(
            url_for("financial_year_control_centre")
        )

    except Exception as exc:
        db.session.rollback()

        flash(
            f"Financial year closing failed: {exc}",
            "danger"
        )

        return redirect(
            url_for("close_financial_year")
        )

@app.route(
    "/opening-balances/<int:batch_id>/post",
    methods=["POST"]
)
@role_required("accounting")
def opening_balance_post(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    if batch.status == "Posted":
        flash(
            "This opening balance batch has already been posted.",
            "warning"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    if batch.status != "Approved":
        flash(
            "Only approved opening balance batches can be posted.",
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    balances = (
        OpeningBalance.query
        .filter_by(batch_id=batch.id)
        .all()
    )

    if not balances:
        flash(
            "No balances were found in this batch.",
            "warning"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    invalid_records = [
        balance
        for balance in balances
        if balance.validation_status != "Valid"
    ]

    if invalid_records:
        flash(
            f"{len(invalid_records)} record(s) are still invalid.",
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    already_posted = [
        balance
        for balance in balances
        if balance.is_posted
    ]

    if already_posted:
        flash(
            (
                f"{len(already_posted)} record(s) in this batch "
                f"have already been posted."
            ),
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    user = session.get("user") or {}

    posted_by = (
        user.get("full_name")
        or user.get("username")
        or "System"
    )

    posting_time = datetime.now(timezone.utc)

    posting_reference = (
        batch.posting_reference
        or f"OB-{batch.batch_no}"
    )

    try:
        total_savings = Decimal("0.00")
        total_loans = Decimal("0.00")
        total_interest = Decimal("0.00")
        total_fines = Decimal("0.00")
        total_welfare = Decimal("0.00")

        created_loans = 0
        created_fines = 0

        for balance in balances:
            savings = money(balance.savings_balance or 0)
            loan_principal = money(balance.loan_principal or 0)
            loan_interest = money(balance.loan_interest or 0)
            fine_balance = money(balance.fine_balance or 0)
            welfare = money(balance.welfare_balance or 0)

            total_savings += savings
            total_loans += loan_principal
            total_interest += loan_interest
            total_fines += fine_balance
            total_welfare += welfare

            # -------------------------------------------------
            # Create migrated opening loan
            # -------------------------------------------------
            if loan_principal > 0 or loan_interest > 0:

                duplicate_loan = Loan.query.filter_by(
                    source_type="OpeningBalance",
                    source_id=balance.id
                ).first()

                if duplicate_loan:
                    raise Exception(
                        (
                            f"An opening loan already exists for "
                            f"opening balance record {balance.id}."
                        )
                    )

                loan = Loan(
                    member_id=balance.member_id,
                    principal=loan_principal,
                    opening_interest=loan_interest,
                    interest_rate=Decimal("0.00"),
                    issued_on=batch.effective_date,
                    due_on=batch.effective_date,
                    purpose="Opening loan balance brought forward",
                    status="Disbursed",
                    approved_by="Opening Balance Migration",
                    reviewed_by="Opening Balance Migration",
                    disbursed_by="Opening Balance Migration",
                    reviewed_on=batch.effective_date,
                    approved_on=batch.effective_date,
                    disbursed_on=batch.effective_date,
                    disbursement_method="Opening Balance",
                    disbursement_reference=posting_reference,
                    source_type="OpeningBalance",
                    source_id=balance.id
                )

                db.session.add(loan)
                db.session.flush()

                loan.loan_no = f"OBLN{loan.id:06d}"

                balance.created_loan_id = loan.id
                created_loans += 1

            # -------------------------------------------------
            # Create migrated opening fine
            # -------------------------------------------------
            if fine_balance > 0:

                duplicate_fine = FinePenalty.query.filter_by(
                    source_type="OpeningBalance",
                    source_id=balance.id
                ).first()

                if duplicate_fine:
                    raise Exception(
                        (
                            f"An opening fine already exists for "
                            f"opening balance record {balance.id}."
                        )
                    )

                fine = FinePenalty(
                    member_id=balance.member_id,
                    category="Opening Balance",
                    amount=fine_balance,
                    reason="Outstanding fine brought forward",
                    fine_date=batch.effective_date,
                    status="Unpaid",
                    recorded_by=posted_by,
                    source_type="OpeningBalance",
                    source_id=balance.id
                )

                db.session.add(fine)
                db.session.flush()

                balance.created_fine_id = fine.id
                created_fines += 1

            balance.is_posted = True
            balance.posted_on = posting_time
            balance.posted_by = posted_by

        # -----------------------------------------------------
        # Build the opening journal
        # -----------------------------------------------------
        required_account_codes = [
            "1100",
            "1110",
            "1120",
            "2000",
            "2010",
            "3900"
        ]

        accounts = {
            account.code: account
            for account in Account.query.filter(
                Account.code.in_(required_account_codes)
            ).all()
        }

        missing_accounts = [
            code
            for code in required_account_codes
            if code not in accounts
        ]

        if missing_accounts:
            raise Exception(
                (
                    "Chart of Accounts missing: "
                    + ", ".join(missing_accounts)
                )
            )

        journal_lines = []

        if total_loans > 0:
            journal_lines.append({
                "account": accounts["1100"],
                "debit": money(total_loans),
                "memo": "Opening loan principal"
            })

        if total_interest > 0:
            journal_lines.append({
                "account": accounts["1110"],
                "debit": money(total_interest),
                "memo": "Opening loan interest receivable"
            })

        if total_fines > 0:
            journal_lines.append({
                "account": accounts["1120"],
                "debit": money(total_fines),
                "memo": "Opening fines receivable"
            })

        if total_savings > 0:
            journal_lines.append({
                "account": accounts["2000"],
                "credit": money(total_savings),
                "memo": "Opening member savings"
            })

        if total_welfare > 0:
            journal_lines.append({
                "account": accounts["2010"],
                "credit": money(total_welfare),
                "memo": "Opening welfare fund"
            })

        debit_total = money(
            total_loans
            + total_interest
            + total_fines
        )

        credit_total = money(
            total_savings
            + total_welfare
        )

        balancing_amount = money(
            credit_total - debit_total
        )

        if balancing_amount > 0:
            journal_lines.append({
                "account": accounts["3900"],
                "debit": balancing_amount,
                "memo": "Opening balance equity"
            })

        elif balancing_amount < 0:
            journal_lines.append({
                "account": accounts["3900"],
                "credit": money(abs(balancing_amount)),
                "memo": "Opening balance equity"
            })

        journal = post_journal(
            entry_date=batch.effective_date,
            description=(
                f"Opening balances posted from batch "
                f"{batch.batch_no}"
            ),
            reference=posting_reference,
            source_type="OpeningBalanceBatch",
            source_id=batch.id,
            lines=journal_lines
        )

        if journal is None:
            raise Exception(
                "A journal already exists for this opening balance batch."
            )

        # -----------------------------------------------------
        # Finalize batch
        # -----------------------------------------------------
        batch.status = "Posted"
        batch.posted_by = posted_by
        batch.posted_on = posting_time
        batch.posting_reference = posting_reference
        

        db.session.commit()

        log_audit(
            "POST_OPENING_BALANCES",
            (
                f"Posted opening balance batch {batch.batch_no}; "
                f"{len(balances)} member record(s), "
                f"{created_loans} opening loan(s), "
                f"{created_fines} opening fine(s), "
                f"reference {posting_reference}"
            )
        )

        flash(
            (
                f"Opening balance batch {batch.batch_no} "
                f"posted successfully."
            ),
            "success"
        )

    except Exception as e:
        db.session.rollback()

        app.logger.exception(
            "Opening balance posting failed for batch %s",
            batch.id
        )

        flash(
            f"Opening balance posting failed: {e}",
            "danger"
        )

    return redirect(
        url_for(
            "opening_balance_batch",
            batch_id=batch.id
        )
    )

@app.route(
    "/opening-balances/<int:batch_id>/reverse",
    methods=["POST"]
)
@role_required("accounting")
def opening_balance_reverse(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    if batch.status != "Posted":
        flash(
            "Only posted opening balance batches can be reversed.",
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    if batch.is_locked:
        flash(
            "This opening balance batch is locked and cannot be reversed.",
            "danger"
        )
        return redirect(
            url_for(
                "opening_balance_batch",
                batch_id=batch.id
            )
        )

    flash(
        (
            "The opening balance reversal engine has not yet been "
            "activated. No records were changed."
        ),
        "warning"
    )

    return redirect(
        url_for(
            "opening_balance_batch",
            batch_id=batch.id
        )
    )


@app.route("/opening-balances/<int:batch_id>/summary")
@role_required("accounting")
def opening_balance_summary_pdf(batch_id):

    batch = OpeningBalanceBatch.query.get_or_404(batch_id)

    balances = (
        OpeningBalance.query
        .filter_by(batch_id=batch.id)
        .join(Member)
        .order_by(Member.member_no)
        .all()
    )

    settings = SystemSetting.query.first()

    return render_template(
        "opening_balances/summary_pdf.html",
        batch=batch,
        balances=balances,
        settings=settings,
        client_name=CLIENT_NAME,
        generated_on=datetime.now()
    )

@app.route('/accounting/financial-year')
@login_required
@role_required('accounting')
def financial_year_control_centre():
    financial_year = current_financial_year()

    financial_years = (
        FinancialYear.query
        .order_by(FinancialYear.year.desc())
        .all()
    )

    return render_template(
        'financial_year/control_centre.html',
        financial_year=financial_year,
        financial_years=financial_years
    )

@app.route('/accounting/ledger-inquiry')
@login_required
@role_required('accounting')
def ledger_inquiry():

    account_id = request.args.get('account_id', type=int)
    source = request.args.get('source', '').strip()
    start = request.args.get('start')
    end = request.args.get('end')

    query = JournalLine.query.join(JournalEntry).join(Account)

    if account_id:
        query = query.filter(JournalLine.account_id == account_id)

    if source:
        query = query.filter(JournalEntry.source_type == source)

    if start:
        query = query.filter(
            JournalEntry.entry_date >= parse_date(start)
        )

    if end:
        query = query.filter(
            JournalEntry.entry_date <= parse_date(end)
        )

    entries = query.order_by(
        JournalEntry.entry_date.desc(),
        JournalEntry.id.desc(),
        JournalLine.id.desc()
    ).all()

    accounts = Account.query.filter_by(active=True).order_by(Account.code).all()

    total_debit = money(
        sum((e.debit for e in entries), Decimal('0.00'))
    )

    total_credit = money(
        sum((e.credit for e in entries), Decimal('0.00'))
    )

    return render_template(
        'ledger_inquiry.html',
        entries=entries,
        accounts=accounts,
        total_debit=total_debit,
        total_credit=total_credit,
        account_id=account_id,
        source=source,
        start=start,
        end=end
    )
@app.route('/accounting/general-ledger/<int:account_id>')
@login_required
@role_required('accounting')
def general_ledger(account_id):

    account = Account.query.get_or_404(account_id)

    lines = (
        JournalLine.query
        .join(JournalEntry)
        .filter(JournalLine.account_id == account.id)
        .order_by(
            JournalEntry.entry_date,
            JournalEntry.id,
            JournalLine.id
        )
        .all()
    )

    running_balance = Decimal('0.00')
    ledger = []

    for line in lines:

        debit = money(line.debit)
        credit = money(line.credit)

        if account.normal_balance == 'Debit':
            running_balance += debit
            running_balance -= credit
        else:
            running_balance += credit
            running_balance -= debit

        ledger.append({
            'date': line.entry.entry_date,
            'description': line.entry.description,
            'reference': line.entry.reference,
            'source': line.entry.source_type,
            'debit': debit,
            'credit': credit,
            'balance': money(running_balance),
            'journal_id': line.entry.id
        })

    return render_template(
        'general_ledger.html',
        account=account,
        ledger=ledger
    )

@app.route('/accounting/journal-adjustment/<int:journal_line_id>', methods=['GET', 'POST'])
@login_required
@role_required('accounting')
def journal_adjustment(journal_line_id):
    original_line = JournalLine.query.get_or_404(journal_line_id)
    original_entry = original_line.entry

    accounts = Account.query.filter_by(active=True).order_by(Account.code).all()

    if request.method == 'POST':
        correct_account = Account.query.get_or_404(
            int(request.form.get('correct_account_id'))
        )

        reason = request.form.get('reason', '').strip()

        if not reason:
            flash('Reason for adjustment is required.', 'error')
            return redirect(url_for('journal_adjustment', journal_line_id=journal_line_id))

        amount = money(original_line.debit or original_line.credit)
        wrong_account = original_line.account

        if original_line.debit and original_line.debit > 0:
            debit_account = correct_account
            credit_account = wrong_account
        else:
            debit_account = wrong_account
            credit_account = correct_account

        description = (
            f"Reclassification adjustment: "
            f"{wrong_account.code} {wrong_account.name} to "
            f"{correct_account.code} {correct_account.name}. "
            f"Reason: {reason}"
        )

        reference = f"ADJ-JL-{original_line.id}"

        adjustment = post_journal(
            date.today(),
            description,
            reference,
            'ManualAdjustment',
            f'{original_line.id}-{datetime.utcnow().timestamp()}',
            [
                {'account': debit_account, 'debit': amount},
                {'account': credit_account, 'credit': amount},
            ]
        )

        log_audit(
            'POST_JOURNAL_ADJUSTMENT',
            'JournalEntry',
            adjustment.id if adjustment else None,
            f'Adjusted journal line {original_line.id}: {wrong_account.code} to {correct_account.code}. Reason: {reason}'
        )

        flash('Journal adjustment posted successfully.')
        return redirect(url_for(
                'ledger_inquiry',
                account_id=wrong_account.id
            ))

    return render_template(
        'journal_adjustment.html',
        original_line=original_line,
        original_entry=original_entry,
        accounts=accounts
    )
@app.route('/accounting/control-centre')
@login_required
@role_required('accounting')
def accounting_control_centre():
    journal_entries_count = JournalEntry.query.count()
    journal_lines_count = JournalLine.query.count()
    cashbook_entries_count = CashBookEntry.query.count()
    active_accounts_count = Account.query.filter_by(active=True).count()

    manual_adjustments_count = JournalEntry.query.filter_by(
        source_type='ManualAdjustment'
    ).count()

    journals = JournalEntry.query.all()

    balanced_journals = 0
    journals_requiring_review = 0

    for journal in journals:
        debit_total = money(
            sum((line.debit for line in journal.lines), Decimal('0.00'))
        )

        credit_total = money(
            sum((line.credit for line in journal.lines), Decimal('0.00'))
        )

        if debit_total == credit_total:
            balanced_journals += 1
        else:
            journals_requiring_review += 1

    last_journal = JournalEntry.query.order_by(
        JournalEntry.created_at.desc()
    ).first()

    accounting_status = "Healthy"
    accounting_status_level = "good"

    if journals_requiring_review > 0:
        accounting_status = "Needs Review"
        accounting_status_level = "watch"

    if journal_entries_count == 0:
        accounting_status = "No Activity Yet"
        accounting_status_level = "watch"

    integrity_issues = []
    integrity_checks = []

    # 1. Unbalanced journals
    if journals_requiring_review > 0:
        integrity_issues.append(
            f'{journals_requiring_review} unbalanced journal(s) found.'
        )
    else:
        integrity_checks.append('All journals are balanced.')

    # 2. Cash Book entries without journals
    cashbook_without_journal = []

    for cb in CashBookEntry.query.all():
        exists = JournalEntry.query.filter_by(
            source_type='CashBook',
            source_id=str(cb.id)
        ).first()

        if not exists:
            cashbook_without_journal.append(cb)

    if cashbook_without_journal:
        integrity_issues.append(
            f'{len(cashbook_without_journal)} Cash Book entr(y/ies) have no journal.'
        )
    else:
        integrity_checks.append('All Cash Book entries have matching journals.')

    # 3. Journal entries without lines
    journals_without_lines = []

    for journal in journals:
        if not journal.lines:
            journals_without_lines.append(journal)

    if journals_without_lines:
        integrity_issues.append(
            f'{len(journals_without_lines)} journal entr(y/ies) have no lines.'
        )
    else:
        integrity_checks.append('All journals have posting lines.')

    balances = ledger_balances()

    total_debit = money(
        sum((b['debit'] for b in balances), Decimal('0.00'))
    )

    total_credit = money(
        sum((b['credit'] for b in balances), Decimal('0.00'))
    )

    # 4. Trial balance check
    if total_debit == total_credit:
        integrity_checks.append('Trial Balance agrees.')
    else:
        integrity_issues.append('Trial Balance does not agree.')

    # Accounting health score
    integrity_score = max(100 - (len(integrity_issues) * 20), 0)

    if integrity_score >= 90:
        integrity_status = 'Excellent'
    elif integrity_score >= 70:
        integrity_status = 'Good'
    elif integrity_score >= 50:
        integrity_status = 'Needs Attention'
    else:
        integrity_status = 'Critical'

    return render_template(
        'accounting_control_centre.html',
        journal_entries_count=journal_entries_count,
        journal_lines_count=journal_lines_count,
        cashbook_entries_count=cashbook_entries_count,
        active_accounts_count=active_accounts_count,
        manual_adjustments_count=manual_adjustments_count,
        balanced_journals=balanced_journals,
        journals_requiring_review=journals_requiring_review,
        last_journal=last_journal,
        accounting_status=accounting_status,
        accounting_status_level=accounting_status_level
    )
@app.route('/accounting/journal-voucher/<int:journal_id>')
@login_required
@role_required('accounting')
def journal_voucher(journal_id):
    journal = JournalEntry.query.get_or_404(journal_id)

    total_debit = money(
        sum((line.debit for line in journal.lines), Decimal('0.00'))
    )

    total_credit = money(
        sum((line.credit for line in journal.lines), Decimal('0.00'))
    )

    is_balanced = total_debit == total_credit

    return render_template(
        'journal_voucher.html',
        journal=journal,
        total_debit=total_debit,
        total_credit=total_credit,
        is_balanced=is_balanced
    )

@app.route('/accounting/bank-reconciliation', methods=['GET', 'POST'])
@login_required
@role_required('accounting')
def bank_reconciliation():
    if request.method == 'POST':
        user = session.get('user') or {}

        line = BankStatementLine(
            statement_date=parse_date(request.form.get('statement_date')),
            description=request.form.get('description'),
            reference=request.form.get('reference'),
            amount=money(request.form.get('amount')),
            entry_type=request.form.get('entry_type'),
            created_by=user.get('username')
        )

        db.session.add(line)
        db.session.commit()

        flash('Bank statement line added.')
        return redirect(url_for('bank_reconciliation'))

    bank_lines = BankStatementLine.query.order_by(
        BankStatementLine.statement_date.desc(),
        BankStatementLine.id.desc()
    ).all()

    cash_entries = CashBookEntry.query.order_by(
        CashBookEntry.entry_date.desc(),
        CashBookEntry.id.desc()
    ).all()

    import_batches = BankStatementImportBatch.query.order_by(
        BankStatementImportBatch.imported_on.desc()
    ).all()

    bank_statement_in = money(sum((l.amount for l in bank_lines if l.entry_type == 'In'), Decimal('0.00')))
    bank_statement_out = money(sum((l.amount for l in bank_lines if l.entry_type == 'Out'), Decimal('0.00')))
    bank_statement_balance = money(bank_statement_in - bank_statement_out)

    cash_book_in = money(sum((e.amount for e in cash_entries if e.entry_type == 'In'), Decimal('0.00')))
    cash_book_out = money(sum((e.amount for e in cash_entries if e.entry_type == 'Out'), Decimal('0.00')))
    cash_book_balance = money(cash_book_in - cash_book_out)

    reconciliation_difference = money(bank_statement_balance - cash_book_balance)

    matched_count = sum(1 for l in bank_lines if l.reconciled)
    outstanding_count = sum(1 for l in bank_lines if not l.reconciled)

    match_rate = 0
    if bank_lines:
        match_rate = round((matched_count / len(bank_lines)) * 100, 1)

    reconciled_cash_entry_ids = [
        l.cash_book_entry_id
        for l in bank_lines
        if l.reconciled and l.cash_book_entry_id
    ]

    unreconciled_cash_entries = [
        e for e in cash_entries
        if e.id not in reconciled_cash_entry_ids
    ]

    outstanding_deposits = [
        e for e in unreconciled_cash_entries
        if e.entry_type == 'In'
    ]

    unpresented_cheques = [
        e for e in unreconciled_cash_entries
        if e.entry_type == 'Out'
    ]

    outstanding_deposits_total = money(sum((e.amount for e in outstanding_deposits), Decimal('0.00')))
    unpresented_cheques_total = money(sum((e.amount for e in unpresented_cheques), Decimal('0.00')))

    return render_template(
        'bank_reconciliation.html',
        bank_lines=bank_lines,
        cash_entries=cash_entries,
        import_batches=import_batches,
        bank_statement_balance=bank_statement_balance,
        cash_book_balance=cash_book_balance,
        reconciliation_difference=reconciliation_difference,
        matched_count=matched_count,
        outstanding_count=outstanding_count,
        match_rate=match_rate,
        outstanding_deposits=outstanding_deposits,
        unpresented_cheques=unpresented_cheques,
        outstanding_deposits_total=outstanding_deposits_total,
        unpresented_cheques_total=unpresented_cheques_total
    )

@app.route('/accounting/bank-reconciliation/match', methods=['POST'])
@login_required
@role_required('accounting')
def match_bank_reconciliation():
    bank_line_id = request.form.get('bank_line_id', type=int)
    cash_entry_id = request.form.get('cash_entry_id', type=int)

    bank_line = BankStatementLine.query.get_or_404(bank_line_id)
    cash_entry = CashBookEntry.query.get_or_404(cash_entry_id)

    bank_line.reconciled = True
    bank_line.cash_book_entry_id = cash_entry.id

    db.session.commit()

    log_audit(
        'BANK_RECONCILIATION_MATCH',
        'BankStatementLine',
        bank_line.id,
        f'Bank statement line matched to Cash Book entry #{cash_entry.id}'
    )

    flash('Bank statement line matched to Cash Book entry.')
    return redirect(url_for('bank_reconciliation'))

@app.route('/accounting/bank-reconciliation/auto-match', methods=['POST'])
@login_required
@role_required('accounting')
def auto_match_bank_reconciliation():
    matched = 0

    unmatched_bank_lines = BankStatementLine.query.filter_by(
        reconciled=False
    ).all()

    for line in unmatched_bank_lines:
        cash_entry = CashBookEntry.query.filter(
            CashBookEntry.entry_date == line.statement_date,
            CashBookEntry.entry_type == line.entry_type,
            CashBookEntry.amount == line.amount
        ).first()

        if cash_entry:
            line.reconciled = True
            line.cash_book_entry_id = cash_entry.id
            matched += 1

    db.session.commit()

    log_audit(
        'BANK_RECONCILIATION_AUTO_MATCH',
        'BankStatementLine',
        None,
        f'{matched} bank statement line(s) automatically matched.'
    )

    flash(f'{matched} bank statement line(s) automatically matched.')
    return redirect(url_for('bank_reconciliation'))

import csv
from io import StringIO

@app.route('/accounting/bank-reconciliation/import', methods=['POST'])
@login_required
@role_required('accounting')
def import_bank_statement_csv():
    file = request.files.get('bank_statement')

    if not file:
        flash('Please select a bank statement CSV file.', 'error')
        return redirect(url_for('bank_reconciliation'))

    content = file.stream.read().decode('utf-8-sig')
    reader = csv.DictReader(StringIO(content))

    imported = 0
    user = session.get('user') or {}

    for row in reader:
        amount = money(row.get('amount') or row.get('Amount') or 0)

        line = BankStatementLine(
            statement_date=parse_date(row.get('date') or row.get('Date')),
            description=row.get('description') or row.get('Description'),
            reference=row.get('reference') or row.get('Reference'),
            amount=abs(amount),
            entry_type='In' if amount > 0 else 'Out',
            created_by=user.get('username')
        )

        db.session.add(line)
        imported += 1

    db.session.commit()

    flash(f'{imported} bank statement line(s) imported successfully.')
    return redirect(url_for('bank_reconciliation'))


@app.route('/accounting/bank-reconciliation/import-pdf', methods=['POST'])
@login_required
@role_required('accounting')
def import_bank_statement_pdf():
    file = request.files.get('bank_statement_pdf')

    if not file:
        flash('Please select a PDF bank statement.', 'error')
        return redirect(url_for('bank_reconciliation'))

    user = session.get('user') or {}

    try:
        transactions = import_bank_statement(file, bank_name="FNB")

        # Create one import batch
        batch = BankStatementImportBatch(
            file_name=file.filename,
            bank_name="FNB",
            imported_by=user.get("username"),
            line_count=len(transactions)
        )

        db.session.add(batch)
        db.session.flush()      # Gets batch.id without committing

    # Save all imported transactions
        for tx in transactions:
            line = BankStatementLine(
                statement_date=tx["statement_date"],
                description=tx["description"],
                reference=tx["reference"],
                amount=money(tx["amount"]),
                entry_type=tx["entry_type"],
                created_by=user.get("username"),
                import_batch_id=batch.id      # Link every line to this batch
            )

            db.session.add(line)

        db.session.commit()

        flash(f'{len(transactions)} bank statement line(s) imported successfully.')

    except Exception as e:
        db.session.rollback()
        flash(f'PDF import failed: {str(e)}', 'error')

    return redirect(url_for('bank_reconciliation'))

@app.route('/accounting/bank-reconciliation/delete-batch/<int:batch_id>', methods=['POST'])
@login_required
@role_required('accounting')
def delete_bank_statement_batch(batch_id):
    batch = BankStatementImportBatch.query.get_or_404(batch_id)

    matched_count = BankStatementLine.query.filter_by(
        import_batch_id=batch.id,
        reconciled=True
    ).count()

    if matched_count > 0:
        flash('Cannot delete this batch because some lines have already been reconciled.', 'error')
        return redirect(url_for('bank_reconciliation'))

    BankStatementLine.query.filter_by(
        import_batch_id=batch.id
    ).delete()

    db.session.delete(batch)
    db.session.commit()

    flash('Bank statement import batch deleted successfully.')
    return redirect(url_for('bank_reconciliation'))

@app.route('/accounting/bank-reconciliation/statement')
@login_required
@role_required('accounting')
def bank_reconciliation_statement():
    bank_lines = BankStatementLine.query.all()
    cash_entries = CashBookEntry.query.all()

    bank_statement_in = money(sum((l.amount for l in bank_lines if l.entry_type == 'In'), Decimal('0.00')))
    bank_statement_out = money(sum((l.amount for l in bank_lines if l.entry_type == 'Out'), Decimal('0.00')))
    bank_statement_balance = money(bank_statement_in - bank_statement_out)

    cash_book_in = money(sum((e.amount for e in cash_entries if e.entry_type == 'In'), Decimal('0.00')))
    cash_book_out = money(sum((e.amount for e in cash_entries if e.entry_type == 'Out'), Decimal('0.00')))
    cash_book_balance = money(cash_book_in - cash_book_out)

    reconciled_cash_entry_ids = [
        l.cash_book_entry_id
        for l in bank_lines
        if l.reconciled and l.cash_book_entry_id
    ]

    unreconciled_cash_entries = [
        e for e in cash_entries
        if e.id not in reconciled_cash_entry_ids
    ]

    outstanding_deposits = [
        e for e in unreconciled_cash_entries
        if e.entry_type == 'In'
    ]

    unpresented_cheques = [
        e for e in unreconciled_cash_entries
        if e.entry_type == 'Out'
    ]

    outstanding_deposits_total = money(sum((e.amount for e in outstanding_deposits), Decimal('0.00')))
    unpresented_cheques_total = money(sum((e.amount for e in unpresented_cheques), Decimal('0.00')))

    adjusted_bank_balance = money(
        bank_statement_balance
        + outstanding_deposits_total
        - unpresented_cheques_total
    )

    difference = money(adjusted_bank_balance - cash_book_balance)

    return render_template(
        'bank_reconciliation_statement.html',
        bank_statement_balance=bank_statement_balance,
        outstanding_deposits=outstanding_deposits,
        outstanding_deposits_total=outstanding_deposits_total,
        unpresented_cheques=unpresented_cheques,
        unpresented_cheques_total=unpresented_cheques_total,
        adjusted_bank_balance=adjusted_bank_balance,
        cash_book_balance=cash_book_balance,
        difference=difference,
        current_date=date.today()
    )
@app.route('/accounting/bank-reconciliation/complete', methods=['POST'])
@login_required
@role_required('accounting')
def complete_bank_reconciliation():
    bank_lines = BankStatementLine.query.all()
    cash_entries = CashBookEntry.query.all()

    bank_statement_in = money(sum((l.amount for l in bank_lines if l.entry_type == 'In'), Decimal('0.00')))
    bank_statement_out = money(sum((l.amount for l in bank_lines if l.entry_type == 'Out'), Decimal('0.00')))
    bank_statement_balance = money(bank_statement_in - bank_statement_out)

    cash_book_in = money(sum((e.amount for e in cash_entries if e.entry_type == 'In'), Decimal('0.00')))
    cash_book_out = money(sum((e.amount for e in cash_entries if e.entry_type == 'Out'), Decimal('0.00')))
    cash_book_balance = money(cash_book_in - cash_book_out)

    reconciled_cash_entry_ids = [
        l.cash_book_entry_id
        for l in bank_lines
        if l.reconciled and l.cash_book_entry_id
    ]

    unreconciled_cash_entries = [
        e for e in cash_entries
        if e.id not in reconciled_cash_entry_ids
    ]

    outstanding_deposits = [
        e for e in unreconciled_cash_entries
        if e.entry_type == 'In'
    ]

    unpresented_cheques = [
        e for e in unreconciled_cash_entries
        if e.entry_type == 'Out'
    ]

    outstanding_deposits_total = money(sum((e.amount for e in outstanding_deposits), Decimal('0.00')))
    unpresented_cheques_total = money(sum((e.amount for e in unpresented_cheques), Decimal('0.00')))

    adjusted_bank_balance = money(
        bank_statement_balance
        + outstanding_deposits_total
        - unpresented_cheques_total
    )

    difference = money(adjusted_bank_balance - cash_book_balance)

    if difference != 0:
        flash('Bank reconciliation cannot be completed because the difference is not zero.', 'error')
        return redirect(url_for('bank_reconciliation_statement'))

    user = session.get('user') or {}

    reconciliation = BankReconciliation(
        reconciliation_date=date.today(),
        bank_statement_balance=bank_statement_balance,
        outstanding_deposits=outstanding_deposits_total,
        unpresented_cheques=unpresented_cheques_total,
        adjusted_bank_balance=adjusted_bank_balance,
        cash_book_balance=cash_book_balance,
        difference=difference,
        status='Completed',
        prepared_by=user.get('full_name') or user.get('username')
    )

    db.session.add(reconciliation)
    db.session.commit()

    log_audit(
        'COMPLETE_BANK_RECONCILIATION',
        'BankReconciliation',
        reconciliation.id,
        f'Bank reconciliation completed as at {reconciliation.reconciliation_date}'
    )

    flash('Bank reconciliation marked as completed.')
    return redirect(url_for('bank_reconciliation_statement'))

@app.route(
    "/accounting/financial-year/pre-close-checks"
)
@login_required
@role_required("accounting")
def financial_year_preclose_checks():
    financial_year = current_financial_year()

    validation = run_financial_year_preclose_checks(
        financial_year
    )

    log_audit(
        "RUN_PRE_CLOSE_CHECKS",
        (
            f"Pre-close checks run for financial year "
            f"{financial_year.year if financial_year else 'Unknown'}; "
            f"{validation['passed_count']} passed, "
            f"{validation['failed_count']} failed."
        )
    )

    return render_template(
        "financial_year/pre_close_checks.html",
        financial_year=financial_year,
        validation=validation
    )

@app.route('/accounting/year-end')
@login_required
@role_required('accounting')
def year_end_closing():
    start = request.args.get('start')
    end = request.args.get('end')

    start_date = parse_date(start) if start else date(date.today().year, 1, 1)
    end_date = parse_date(end) if end else date(date.today().year, 12, 31)

    balances = ledger_balances(start_date, end_date)

    total_debits = money(sum((b['debit'] for b in balances), Decimal('0.00')))
    total_credits = money(sum((b['credit'] for b in balances), Decimal('0.00')))
    trial_balance_ok = total_debits == total_credits

    income_accounts = []
    expense_accounts = []

    for b in balances:
        account = b['account']

        if account.account_type == 'Income' and b['balance'] != 0:
            income_accounts.append(b)

        if account.account_type == 'Expense' and b['balance'] != 0:
            expense_accounts.append(b)

    total_income = money(sum((b['balance'] for b in income_accounts), Decimal('0.00')))
    total_expenses = money(sum((b['balance'] for b in expense_accounts), Decimal('0.00')))
    net_surplus = money(total_income - total_expenses)

    return render_template(
        'year_end_closing.html',
        start=start_date,
        end=end_date,
        balances=balances,
        total_debits=total_debits,
        total_credits=total_credits,
        trial_balance_ok=trial_balance_ok,
        income_accounts=income_accounts,
        expense_accounts=expense_accounts,
        total_income=total_income,
        total_expenses=total_expenses,
        net_surplus=net_surplus
    )
@app.route('/accounting/year-end/close', methods=['POST'])
@login_required
@role_required('accounting')
def year_end_close():

    start = parse_date(request.form['start'])
    end = parse_date(request.form['end'])

    if not start or not end:
        flash('Invalid financial year.', 'error')
        return redirect(url_for('year_end_closing'))

    fy = FinancialYear.query.filter_by(
        start_date=start,
        end_date=end
    ).first()

    if fy and fy.status == 'Closed':
        flash('This financial year has already been closed.', 'error')
        return redirect(url_for('year_end_closing'))

    balances = ledger_balances(start, end)

    total_debits = money(sum((b['debit'] for b in balances), Decimal('0.00')))
    total_credits = money(sum((b['credit'] for b in balances), Decimal('0.00')))

    if total_debits != total_credits:
        flash('Cannot close year because the Trial Balance is not balanced.', 'error')
        return redirect(url_for('year_end_closing'))

    flash('Trial Balance verified. Ready to generate closing journals.', 'success')
    
    journal = JournalEntry(
    entry_date=end,
    description=f'Year-End Closing {end.year}',
    source_type='YearEndClosing'
)

    db.session.add(journal)
    db.session.flush()

    closing_amount = Decimal('0.00')

    for b in balances:
        account = b['account']

        if account.account_type == 'Income' and b['balance'] != 0:

            db.session.add(
                JournalLine(
                    journal_entry_id=journal.id,
                    account_id=account.id,
                    debit=b['balance'],
                    credit=Decimal('0.00')
                )
            )

            closing_amount += b['balance']

    surplus_account = Account.query.filter_by(code='3000').first()

    if not surplus_account:
        raise Exception('Account 3000 - Accumulated Surplus is missing.')

    if closing_amount > 0:
        db.session.add(
            JournalLine(
                journal_entry_id=journal.id,
                account_id=surplus_account.id,
                debit=Decimal('0.00'),
                credit=closing_amount
            )
        )

        expense_closing_amount = Decimal('0.00')

    for b in balances:
        account = b['account']

        if account.account_type == 'Expense' and b['balance'] != 0:

            db.session.add(
                JournalLine(
                    journal_entry_id=journal.id,
                    account_id=account.id,
                    debit=Decimal('0.00'),
                    credit=b['balance']
                )
            )

            expense_closing_amount += b['balance']

    if expense_closing_amount > 0:
        db.session.add(
            JournalLine(
                journal_entry_id=journal.id,
                account_id=surplus_account.id,
                debit=expense_closing_amount,
                credit=Decimal('0.00')
            )
        )
    if not fy:
        fy = FinancialYear(
            start_date=start,
            end_date=end
        )
        db.session.add(fy)

    user = session.get('user') or {}

    fy.status = 'Closed'
    fy.closed_on = datetime.utcnow()
    fy.closed_by = user.get('full_name') or user.get('username')

    db.session.commit()

    log_audit(
        'YEAR_END_CLOSE',
        'FinancialYear',
        fy.id,
        f'Financial year {start} to {end} closed. Net surplus transferred to Accumulated Surplus.'
    )

    flash('Financial year closed successfully.')
    return redirect(url_for('year_end_closing', start=start, end=end))       

@app.route('/accounting', methods=['GET', 'POST'])
@login_required
@role_required('accounting')
def accounting():
    seed_chart_of_accounts()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'sync':
            created = sync_operational_transactions_to_gl()
            log_audit('SYNC_ACCOUNTING', 'JournalEntry', None, f'{created} operational transaction(s) posted to the general ledger')
            flash(f'{created} operational transaction(s) posted to the General Ledger.')
            return redirect(url_for('accounting'))
        if action == 'account':
            code = request.form.get('code','').strip()
            name = request.form.get('name','').strip()
            account_type = request.form.get('account_type','Asset')
            normal_balance = 'Debit' if account_type in ['Asset','Expense'] else 'Credit'
            if not code or not name:
                flash('Account code and name are required.', 'error')
            elif Account.query.filter_by(code=code).first():
                flash('That account code already exists.', 'error')
            else:
                db.session.add(Account(code=code, name=name, account_type=account_type, normal_balance=normal_balance))
                db.session.commit()
                log_audit('CREATE_ACCOUNT', 'Account', code, f'Account {code} - {name} created')
                flash('Account created successfully.')
            return redirect(url_for('accounting'))
       
        if action == 'journal':
            entry_date = parse_date(request.form.get('entry_date'))
            description = request.form.get('description', '').strip()
            reference = request.form.get('reference', '').strip()

            debit_account = Account.query.get(int(request.form.get('debit_account_id') or 0))
            credit_account = Account.query.get(int(request.form.get('credit_account_id') or 0))

            amount = money(request.form.get('amount') or 0)

            if not description or not debit_account or not credit_account or amount <= 0:
                flash('Description, debit account, credit account and amount are required.', 'error')

            else:
                entry = post_journal(
                    entry_date=entry_date,
                    description=description,
                    debit_account_code=debit_account.code,
                    credit_account_code=credit_account.code,
                    amount=amount,
                    reference=reference,
                    source_type='Manual',
                    source_id=f'{datetime.utcnow().timestamp()}'
                )

                log_audit(
                    'POST_MANUAL_JOURNAL',
                    'JournalEntry',
                    entry.id if entry else None,
                    f'Manual journal posted: {description}'
                )

                flash('Manual journal entry posted successfully.')

        return redirect(url_for('accounting'))

    start = request.args.get('start')
    end = request.args.get('end')
    start_date = parse_date(start) if start else None
    end_date = parse_date(end) if end else None
    accounts = Account.query.order_by(Account.code).all()

    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = JournalEntry.query.order_by(
        JournalEntry.entry_date.desc(),
        JournalEntry.id.desc()
    ).paginate(
    page=page,
    per_page=per_page,
    error_out=False
    )
    entries = pagination.items
    balances = ledger_balances(start_date, end_date)
    total_debits = money(sum((b['debit'] for b in balances), Decimal('0.00')))
    total_credits = money(sum((b['credit'] for b in balances), Decimal('0.00')))
    income_total = money(sum((b['balance'] for b in balances if b['account'].account_type == 'Income'), Decimal('0.00')))
    expense_total = money(sum((b['balance'] for b in balances if b['account'].account_type == 'Expense'), Decimal('0.00')))
    net_surplus = money(income_total - expense_total)
    asset_total = money(sum((b['balance'] for b in balances if b['account'].account_type == 'Asset'), Decimal('0.00')))
    liability_total = money(sum((b['balance'] for b in balances if b['account'].account_type == 'Liability'), Decimal('0.00')))
    equity_total = money(sum((b['balance'] for b in balances if b['account'].account_type == 'Equity'), Decimal('0.00')))
    cashbook = [b for b in balances if b['account'].code in ['1000','1010','1020']]
    cash_total = money(
    sum(
        (
            b['balance']
            for b in cashbook
        ),
        Decimal("0.00")
    )
)
    log_audit('VIEW_ACCOUNTING', 'Accounting', None, 'General Ledger and Accounting viewed')

    financial_service = FinancialIntelligenceService(
    cash_total=cash_total,
    income_total=income_total,
    expense_total=expense_total,
    asset_total=asset_total,
    liability_total=liability_total,
    equity_total=equity_total,
    total_debits=total_debits,
    total_credits=total_credits
)

    financial_data = financial_service.build()

    return render_template('accounting.html', 
                           **locals(),
                           **financial_data)

@app.route('/accounting/cash-book', methods=['GET', 'POST'])
@login_required
@role_required('accounting')
def cash_book():
    if request.method == 'POST':
        user = session.get('user') or {}

        entry = CashBookEntry(
            entry_date=parse_date(request.form.get('entry_date')),
            entry_type=request.form.get('entry_type'),
            category=request.form.get('category'),
            description=request.form.get('description'),
            amount=money(request.form.get('amount')),
            method=request.form.get('method'),
            reference=request.form.get('reference'),
            created_by=user.get('username'),
        )

        db.session.add(entry)
        db.session.flush()

        selected_account = request.form.get('account_code')

        if not selected_account:
            db.session.rollback()
            flash('Please select a ledger account.', 'error')
            return redirect(url_for('cash_book'))

        if entry.entry_type == 'In':
            debit_account = cash_account(entry.method)
            credit_account = selected_account
        else:
            debit_account = selected_account
            credit_account = cash_account(entry.method)

        post_journal(
            entry_date=entry.entry_date,
            description=entry.description or entry.category,
            debit_account_code=debit_account,
            credit_account_code=credit_account,
            amount=entry.amount,
            source_type='CashBook',
            source_id=entry.id
        )

        db.session.commit()

        log_audit(
            'CREATE_CASH_BOOK_ENTRY',
            'CashBookEntry',
            entry.id,
            f'{entry.entry_type} - {entry.category} - {kwacha(entry.amount)}'
        )

        flash('Cash book entry recorded.')
        return redirect(url_for('cash_book'))

    entries = CashBookEntry.query.order_by(
        CashBookEntry.entry_date.desc(),
        CashBookEntry.id.desc()
    ).all()

    total_in = money(
        sum((e.amount for e in entries if e.entry_type == 'In'), Decimal('0.00'))
    )

    total_out = money(
        sum((e.amount for e in entries if e.entry_type == 'Out'), Decimal('0.00'))
    )

    bank_balance = money(total_in - total_out)

    accounts = Account.query.filter(
        ~Account.code.in_(['1000', '1010', '1020'])
    ).order_by(Account.code).all()
    
    
    return render_template(
        'cash_book.html',
        entries=entries,
        total_in=total_in,
        total_out=total_out,
        bank_balance=bank_balance,
        payment_methods=PAYMENT_METHODS,
        accounts=accounts
    )

@app.route('/accounting/trial-balance')
@login_required
@role_required('accounting')
def trial_balance():
    start = request.args.get('start')
    end = request.args.get('end')

    start_date = parse_date(start) if start else None
    end_date = parse_date(end) if end else None

    balances = ledger_balances(start_date, end_date)

    total_debits = money(sum((b['debit'] for b in balances), Decimal('0.00')))
    total_credits = money(sum((b['credit'] for b in balances), Decimal('0.00')))

    return render_template(
        'trial_balance.html',
        balances=balances,
        total_debits=total_debits,
        total_credits=total_credits,
        start=start,
        end=end
    )

@app.route('/accounting/income-statement')
@login_required
@role_required('accounting')
def income_statement():
    start = request.args.get('start')
    end = request.args.get('end')

    start_date = parse_date(start) if start else None
    end_date = parse_date(end) if end else None

    balances = ledger_balances(start_date, end_date)

    income_accounts = []
    expense_accounts = []

    for b in balances:
        account = b['account']

        if account.account_type == 'Income' and b['balance'] != 0:
            income_accounts.append(b)

        if account.account_type == 'Expense' and b['balance'] != 0:
            expense_accounts.append(b)

    total_income = money(sum((b['balance'] for b in income_accounts), Decimal('0.00')))
    total_expenses = money(sum((b['balance'] for b in expense_accounts), Decimal('0.00')))
    net_surplus = money(total_income - total_expenses)

    return render_template(
        'income_statement.html',
        start=start,
        end=end,
        income_accounts=income_accounts,
        expense_accounts=expense_accounts,
        total_income=total_income,
        total_expenses=total_expenses,
        net_surplus=net_surplus
    )

@app.route('/accounting/balance-sheet')
@login_required
@role_required('accounting')
def balance_sheet():
    balances = ledger_balances()

    income = Decimal('0.00')
    expenses = Decimal('0.00')

    for b in balances:
        account = b['account']

        if account.account_type == 'Income':
            income += b['balance']

        elif account.account_type == 'Expense':
            expenses += b['balance']

    current_surplus = money(income - expenses)

    balance_map = {
        b['account'].code: b['balance']
        for b in balances
    }

    cash_on_hand = money(
    balance_map.get('1000', Decimal('0.00'))
    )

    bank_account = money(
        balance_map.get('1010', Decimal('0.00'))
    )

    mobile_money = money(
        balance_map.get('1020', Decimal('0.00'))
    )

    loans_receivable = money(
        balance_map.get('1100', Decimal('0.00'))
    )

    loan_interest_receivable = money(
        balance_map.get('1110', Decimal('0.00'))
    )

    fines_receivable = money(
        balance_map.get('1120', Decimal('0.00'))
    )

    member_savings = money(
        balance_map.get('2000', Decimal('0.00'))
    )

    welfare_fund = money(
        balance_map.get('2010', Decimal('0.00'))
    )

    accumulated_surplus = money(
        balance_map.get('3000', Decimal('0.00'))
    )

    opening_balance_equity = money(
        balance_map.get('3900', Decimal('0.00'))
    )

    total_cash = money(
        cash_on_hand
        + bank_account
        + mobile_money
    )

    total_receivables = money(
        loans_receivable
        + loan_interest_receivable
        + fines_receivable
    )

    total_assets = money(
        total_cash
        + total_receivables
    )

    total_liabilities = money(
        member_savings
        + welfare_fund
    )

    total_equity = money(
        accumulated_surplus
        + opening_balance_equity
        + current_surplus
    )
    total_liabilities_equity = money(total_liabilities + total_equity)

    difference = money(total_assets - total_liabilities_equity)

    return render_template(
        'balance_sheet.html',
        cash_on_hand=cash_on_hand,
        bank_account=bank_account,
        mobile_money=mobile_money,
        total_cash=total_cash,
        loans_receivable=loans_receivable,
        total_assets=total_assets,
        member_savings=member_savings,
        welfare_fund=welfare_fund,
        total_liabilities=total_liabilities,
        accumulated_surplus=accumulated_surplus,
        current_surplus=current_surplus,
        total_equity=total_equity,
        total_liabilities_equity=total_liabilities_equity,
        difference=difference,
        loan_interest_receivable=loan_interest_receivable,
        fines_receivable=fines_receivable,
        total_receivables=total_receivables,
        opening_balance_equity=opening_balance_equity,
    )

@app.route('/accounting/cash-flow')
@login_required
@role_required('accounting')
def cash_flow_statement():
    entries = CashBookEntry.query.order_by(
        CashBookEntry.entry_date.asc(),
        CashBookEntry.id.asc()
    ).all()

    cash_in_entries = [e for e in entries if e.entry_type == 'In']
    cash_out_entries = [e for e in entries if e.entry_type == 'Out']

    total_cash_in = money(
        sum((e.amount for e in cash_in_entries), Decimal('0.00'))
    )

    total_cash_out = money(
        sum((e.amount for e in cash_out_entries), Decimal('0.00'))
    )

    net_cash_flow = money(total_cash_in - total_cash_out)

    cash_in_by_category = {}
    cash_out_by_category = {}

    for e in cash_in_entries:
        category = e.category or 'Unclassified Income'
        cash_in_by_category[category] = money(
            cash_in_by_category.get(category, Decimal('0.00')) + e.amount
        )

    for e in cash_out_entries:
        category = e.category or 'Unclassified Payments'
        cash_out_by_category[category] = money(
            cash_out_by_category.get(category, Decimal('0.00')) + e.amount
        )

    opening_cash_balance = Decimal("0.00")

    closing_cash_balance = net_cash_flow

    cash_reconciliation_difference = money(
        closing_cash_balance - net_cash_flow
    )

    return render_template(
        'cash_flow.html',
        cash_in_by_category=cash_in_by_category,
        cash_out_by_category=cash_out_by_category,
        total_cash_in=total_cash_in,
        total_cash_out=total_cash_out,
        net_cash_flow=net_cash_flow,
        opening_cash_balance=opening_cash_balance,
        closing_cash_balance=closing_cash_balance,
        cash_reconciliation_difference=cash_reconciliation_difference
    )

@app.route('/export/accounting.csv')
@login_required
@role_required('accounting')
def export_accounting_csv():
    seed_chart_of_accounts()
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Entry Date','Reference','Description','Account Code','Account Name','Debit','Credit','Posted By'])
    for entry in JournalEntry.query.order_by(JournalEntry.entry_date.desc(), JournalEntry.id.desc()).all():
        for line in entry.lines:
            writer.writerow([entry.entry_date, entry.reference or '', entry.description, line.account.code, line.account.name, line.debit, line.credit, entry.posted_by or ''])
    log_audit('EXPORT_ACCOUNTING', 'JournalEntry', None, 'General ledger exported to CSV')
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=general_ledger.csv'})

@app.route('/notifications', methods=['GET', 'POST'])
@login_required
@role_required('notifications')
def notifications():
    members = Member.query.order_by(Member.member_no).all()
    q = request.args.get('q', '').strip()

    setting = SystemSetting.query.first()
    organization_name = setting.organisation_name if setting and setting.organisation_name else CLIENT_NAME

    templates = {
        'Contribution Reminder': f'Dear {{name}}, your monthly contribution is due. Please pay through bank transfer, mobile money, or cash. {organization_name}.',
        'Loan Repayment Reminder': f'Dear {{name}}, this is a reminder to make your loan repayment by the due date. {organization_name}.',
        'Meeting Reminder': f'Dear {{name}}, please remember the upcoming Village Banking meeting. Your attendance is important. {organization_name}.',
        'Welfare Notification': f'Dear {{name}}, your welfare fund update has been recorded. Contact the committee for details. {organization_name}.',
        'Share-Out Notification': f'Dear {{name}}, your share-out/dividend information is ready. Contact the treasurer for your statement. {organization_name}.',
        'General Notice': f'Dear {{name}}, this is a notice from {organization_name}.'
    }

    if request.method == 'POST':
        notification_type = request.form.get('notification_type') or 'General Notice'
        channel = request.form.get('channel') or 'SMS'
        subject = request.form.get('subject') or notification_type
        message = request.form.get('message') or ''
        recipient_mode = request.form.get('recipient_mode') or 'selected'
        selected_ids = request.form.getlist('member_ids')

        if recipient_mode == 'all':
            recipients = members
        else:
            ids = [int(x) for x in selected_ids if x.isdigit()]
            recipients = Member.query.filter(Member.id.in_(ids)).all() if ids else []

        if not recipients:
            flash('Please select at least one recipient or choose all members.', 'error')
            return redirect(url_for('notifications'))

        created = 0
        sent = 0
        failed = 0
        user = session.get('user') or {}

        for m in recipients:
            personalized = (
                message
                .replace('{name}', m.full_name)
                .replace('{member_no}', m.member_no)
                .replace('{phone}', m.phone or '')
            )

            status = 'Prepared'

            if channel == 'SMS':
                ok, provider_response = send_sms_via_africas_talking(
                    m.phone,
                    personalized
                )

                if ok:
                    status = 'Sent'
                    sent += 1
                else:
                    status = 'Failed'
                    failed += 1
            flash(f'SMS failed for {m.full_name}: {provider_response}', 'error')

            n = NotificationLog(
                channel=channel,
                notification_type=notification_type,
                recipient_type='All Members' if recipient_mode == 'all' else 'Selected Members',
                member_id=m.id,
                phone=m.phone,
                subject=subject,
                message=personalized,
                status=status,
                created_by=user.get('username')
            )

            db.session.add(n)
            created += 1

        db.session.commit()

        log_audit(
            'SEND_NOTIFICATIONS',
            'NotificationLog',
            None,
            f'{created} {channel} notification(s). Sent: {sent}, Failed: {failed}'
        )

        flash(f'{created} notification(s) processed. Sent: {sent}, Failed: {failed}.')
        return redirect(url_for('notifications'))

    query = NotificationLog.query

    if q:
        query = query.outerjoin(Member).filter(
            (Member.full_name.contains(q)) |
            (Member.member_no.contains(q)) |
            (NotificationLog.phone.contains(q)) |
            (NotificationLog.message.contains(q)) |
            (NotificationLog.notification_type.contains(q))
        )

    page = request.args.get('page', 1, type=int)
    per_page = 10

    pagination = query.order_by(
        NotificationLog.created_at.desc(),
        NotificationLog.id.desc()
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    logs = pagination.items

    return render_template(
        'notifications.html',
        members=members,
        logs=logs,
        pagination=pagination,
        q=q,
        templates=templates
    )

@app.route('/sms-test', methods=['GET', 'POST'])
@login_required
@role_required('settings')
def sms_test():
    if request.method == 'POST':
        phone = request.form.get('phone')
        message = request.form.get('message')

        ok, response = send_sms_via_africas_talking(phone, message)

        if ok:
            flash(f'SMS sent successfully. {response}')
        else:
            flash(f'SMS failed: {response}', 'error')

    return render_template('sms_test.html')

@app.route('/export/notifications.csv')
@login_required
@role_required('notifications')
def export_notifications_csv():
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Created At','Channel','Type','Member No','Member','Phone','Subject','Status','Message'])
    for n in NotificationLog.query.order_by(NotificationLog.created_at.desc()).all():
        writer.writerow([n.created_at, n.channel, n.notification_type, n.member.member_no if n.member else '', n.member.full_name if n.member else '', n.phone, n.subject, n.status, n.message])
    log_audit('EXPORT_NOTIFICATIONS', 'NotificationLog', None, 'Notification log exported to CSV')
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=notifications.csv'})

@app.route('/users')
@login_required
@role_required('users')
def users():
    return render_template('users.html', users=User.query.order_by(User.full_name).all())

@app.route('/users/new', methods=['GET', 'POST'])
@login_required
@role_required('users')
def user_new():
    if request.method == 'POST':
        username = request.form.get('username','').strip().lower()
        password = request.form.get('password','')
        if not username or not password:
            flash('Username and password are required.', 'error')
            return redirect(url_for('user_new'))
        if User.query.filter_by(username=username).first():
            flash('That username already exists.', 'error')
            return redirect(url_for('user_new'))
        user = User(username=username, full_name=request.form.get('full_name','').strip(), role=request.form.get('role'), active=bool(request.form.get('active')), password_hash=generate_password_hash(password))
        db.session.add(user); db.session.commit(); log_audit('CREATE_USER', 'User', user.id, f'{user.username} created with role {user.role}'); flash('User account created successfully.'); return redirect(url_for('users'))
    return render_template('user_form.html', user=None)

@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('users')
def user_edit(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.full_name = request.form.get('full_name','').strip()
        user.role = request.form.get('role')
        user.active = bool(request.form.get('active'))
        new_password = request.form.get('password','').strip()
        if new_password:
            user.password_hash = generate_password_hash(new_password)
        db.session.commit(); log_audit('UPDATE_USER', 'User', user.id, f'{user.username} updated; role {user.role}; active {user.active}'); flash('User account updated successfully.'); return redirect(url_for('users'))
    return render_template('user_form.html', user=user)

@app.route('/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@role_required('users')
def user_toggle(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == session.get('user', {}).get('id'):
        flash('You cannot deactivate your own account while logged in.', 'error')
    else:
        user.active = not user.active
        db.session.commit()
        log_audit('TOGGLE_USER_STATUS', 'User', user.id, f'{user.username} active={user.active}')
        flash('User status updated.')
    return redirect(url_for('users'))



@app.route('/backups')
@login_required
@role_required('backups')
def backups():
    folder = backups_folder()
    disk_files = []
    for f in sorted(Path(folder).glob('*.db'), key=lambda x: x.stat().st_mtime, reverse=True):
        disk_files.append({'filename': f.name, 'size': f.stat().st_size, 'modified': datetime.fromtimestamp(f.stat().st_mtime)})
    records = BackupRecord.query.order_by(BackupRecord.created_at.desc()).limit(100).all()
    return render_template('backups.html', disk_files=disk_files, records=records)

@app.route('/backups/create', methods=['POST'])
@login_required
@role_required('backups')
def backup_create():
    try:
        notes = request.form.get('notes') or 'Manual backup created from Administration menu'
        record = create_database_backup('backup', notes=notes)
        log_audit('CREATE_BACKUP', 'BackupRecord', record.id, f'Backup created: {record.filename}')
        flash(f'Backup created successfully: {record.filename}')
    except Exception as exc:
        flash(f'Backup failed: {exc}', 'error')
    return redirect(url_for('backups'))

@app.route('/backups/download/<path:filename>')
@login_required
@role_required('backups')
def backup_download(filename):
    safe_name = secure_filename(filename)
    path = os.path.join(backups_folder(), safe_name)
    if not os.path.exists(path):
        flash('Backup file not found.', 'error')
        return redirect(url_for('backups'))
    log_audit('DOWNLOAD_BACKUP', 'Backup', safe_name, f'Backup downloaded: {safe_name}')
    return send_file(path, as_attachment=True, download_name=safe_name)

@app.route('/backups/restore', methods=['POST'])
@login_required
@role_required('backups')
def backup_restore():
    uploaded = request.files.get('backup_file')
    confirm = request.form.get('confirm_restore') == 'YES'
    if not uploaded or not uploaded.filename:
        flash('Please choose a backup file to restore.', 'error')
        return redirect(url_for('backups'))
    if not confirm:
        flash('Restore cancelled. Type YES in the confirmation box before restoring.', 'error')
        return redirect(url_for('backups'))
    filename = secure_filename(uploaded.filename)
    if not filename.lower().endswith('.db'):
        flash('Only .db backup files are allowed.', 'error')
        return redirect(url_for('backups'))
    try:
        # Safety backup before restore.
        safety = create_database_backup('pre_restore', notes=f'Safety backup before restoring {filename}')
        temp_path = os.path.join(backups_folder(), f'upload_restore_{datetime.now().strftime("%Y_%m_%d_%H%M%S")}_{filename}')
        uploaded.save(temp_path)
        db.session.remove()
        db.engine.dispose()
        shutil.copy2(temp_path, database_file_path())
        log_audit('RESTORE_BACKUP', 'Backup', filename, f'Database restored from {filename}. Safety backup: {safety.filename}')
        flash(f'Database restored from {filename}. Please stop the server with Ctrl+C and run python app.py again.')
    except Exception as exc:
        flash(f'Restore failed: {exc}', 'error')
    return redirect(url_for('backups'))

@app.route('/backups/export-json')
@login_required
@role_required('backups')
def backup_export_json():
    setting = SystemSetting.query.first()
    organization_name = setting.organisation_name if setting and setting.organisation_name else CLIENT_NAME

    data = {
        'created_at': datetime.utcnow().isoformat(),
        'organization': organization_name,

        'members': [
            {
                'id': m.id,
                'member_no': m.member_no,
                'full_name': m.full_name,
                'phone': m.phone,
                'national_id': m.national_id,
                'group_name': m.group_name,
                'status': m.status,
            }
            for m in Member.query.order_by(Member.id).all()
        ],

        'contributions': [
            {
                'id': c.id,
                'member_id': c.member_id,
                'month': c.month,
                'amount': str(c.amount),
                'method': c.method,
                'reference': c.reference,
                'paid_on': str(c.paid_on),
            }
            for c in Contribution.query.order_by(Contribution.id).all()
        ],

        'loans': [
            {
                'id': l.id,
                'member_id': l.member_id,
                'principal': str(l.principal),
                'interest_amount': str(l.interest_amount),
                'total_due': str(l.total_due),
                'balance': str(l.balance),
                'status': l.status,
                'issued_on': str(l.issued_on),
                'due_on': str(l.due_on),
            }
            for l in Loan.query.order_by(Loan.id).all()
        ],

        'repayments': [
            {
                'id': r.id,
                'loan_id': r.loan_id,
                'amount': str(r.amount),
                'method': r.method,
                'reference': r.reference,
                'paid_on': str(r.paid_on),
            }
            for r in Repayment.query.order_by(Repayment.id).all()
        ],

        'distributions': [
            {
                'id': d.id,
                'member_id': d.member_id,
                'amount': str(d.amount),
                'method': d.method,
                'reference': d.reference,
                'paid_on': str(d.paid_on),
            }
            for d in Distribution.query.order_by(Distribution.id).all()
        ],

        'fines': [
            {
                'id': f.id,
                'member_id': f.member_id,
                'category': f.category,
                'amount': str(f.amount),
                'balance': str(f.balance),
                'status': f.status,
            }
            for f in FinePenalty.query.order_by(FinePenalty.id).all()
        ],

        'fine_payments': [
            {
                'id': fp.id,
                'fine_id': fp.fine_id,
                'amount': str(fp.amount),
                'method': fp.method,
                'reference': fp.reference,
                'paid_on': str(fp.paid_on),
            }
            for fp in FinePayment.query.order_by(FinePayment.id).all()
        ],

        'welfare_contributions': [
            {
                'id': w.id,
                'member_id': w.member_id,
                'month': w.month,
                'amount': str(w.amount),
                'method': w.method,
                'reference': w.reference,
                'paid_on': str(w.paid_on),
            }
            for w in WelfareContribution.query.order_by(WelfareContribution.id).all()
        ],

        'welfare_claims': [
            {
                'id': wc.id,
                'member_id': wc.member_id,
                'category': wc.category,
                'amount_requested': str(wc.amount_requested),
                'amount_approved': str(wc.amount_approved),
                'status': wc.status,
                'reason': wc.reason,
                'requested_on': str(wc.requested_on),
                'paid_on': str(wc.paid_on),
            }
            for wc in WelfareClaim.query.order_by(WelfareClaim.id).all()
        ],

        'savings_interest': [
            {
                'id': s.id,
                'member_id': s.member_id,
                'month': s.month,
                'opening_balance': str(s.opening_balance),
                'interest_rate': str(s.interest_rate),
                'interest_amount': str(s.interest_amount),
                'closing_balance': str(s.closing_balance),
            }
            for s in SavingsInterest.query.order_by(SavingsInterest.id).all()
        ],

        'loan_interest': [
            {
                'id': li.id,
                'loan_id': li.loan_id,
                'member_id': li.member_id,
                'month': li.month,
                'opening_balance': str(li.opening_balance),
                'interest_rate': str(li.interest_rate),
                'interest_amount': str(li.interest_amount),
                'closing_balance': str(li.closing_balance),
            }
            for li in LoanInterest.query.order_by(LoanInterest.id).all()
        ],
    }

    buffer = io.BytesIO()
    buffer.write(json.dumps(data, indent=2).encode('utf-8'))
    buffer.seek(0)

    filename = f'cloud_backup_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.json'

    log_audit(
        'EXPORT_JSON_BACKUP',
        'Backup',
        None,
        f'JSON backup exported: {filename}'
    )

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/json'
    )

@app.route('/audit')
@login_required
@role_required('audit')
def audit_trail():
    q = request.args.get('q', '').strip()
    action = request.args.get('action', '').strip()

    query = AuditLog.query

    if q:
        like = f'%{q}%'
        query = query.filter(
            db.or_(
                AuditLog.username.like(like),
                AuditLog.full_name.like(like),
                AuditLog.entity.like(like),
                AuditLog.details.like(like)
            )
        )

    if action:
        query = query.filter(AuditLog.action == action)

    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = query.order_by(AuditLog.created_at.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    logs = pagination.items

    actions = [
        row[0]
        for row in db.session.query(AuditLog.action)
        .distinct()
        .order_by(AuditLog.action)
        .all()
    ]

    return render_template(
        'audit.html',
        logs=logs,
        pagination=pagination,
        q=q,
        action=action,
        actions=actions
    )
@app.route('/export/audit.csv')
@login_required
@role_required('audit')
def export_audit_csv():
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow(['Date/Time','User','Role','Action','Entity','Entity ID','Details','IP Address'])
    for a in AuditLog.query.order_by(AuditLog.created_at.desc()).limit(2000).all():
        writer.writerow([a.created_at, a.full_name or a.username, a.role, a.action, a.entity, a.entity_id, a.details, a.ip_address])
    log_audit('EXPORT_AUDIT_TRAIL', 'AuditLog', None, 'Audit trail exported to CSV')
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': 'attachment; filename=audit_trail.csv'})


@app.route('/statements')
@login_required
@role_required('statements')
def statements():
    q = request.args.get('q', '').strip()

    query = Member.query

    if q:
        query = query.filter(
            Member.full_name.contains(q) |
            Member.member_no.contains(q) |
            Member.phone.contains(q)
        )

    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = query.order_by(
        Member.member_no
    ).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    return render_template(
        'statements.html',
        members=pagination.items,
        pagination=pagination,
        q=q
    )

@app.route('/statements/member/<int:member_id>.pdf')
@login_required
@role_required('statements')
def member_statement_pdf(member_id):
    member = Member.query.get_or_404(member_id)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=16*mm, leftMargin=16*mm, topMargin=16*mm, bottomMargin=16*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('StatementTitle', parent=styles['Title'], fontSize=18, leading=22, textColor=colors.HexColor('#1f4f68'))
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.HexColor('#555555'))
    normal = styles['Normal']
    story = []

    setting = SystemSetting.query.first()

    organization_name = setting.organisation_name if setting and setting.organisation_name else CLIENT_NAME
    organization_address = setting.organization_address if setting and setting.organization_address else ''
    organization_phone = setting.organization_phone if setting and setting.organization_phone else ''
    organization_email = setting.organization_email if setting and setting.organization_email else ''
    registration_number = setting.registration_number if setting and setting.registration_number else ''

    story.append(Paragraph(organization_name, title_style))

    if registration_number:
        story.append(Paragraph(f'Registration No: {registration_number}', small_style))

    contact_line = ' | '.join(
        x for x in [organization_address, organization_phone, organization_email]
        if x
    )

    if contact_line:
        story.append(Paragraph(contact_line, small_style))

    story.append(Paragraph(f'Member Statement | Produced by {PRODUCER_NAME}', small_style))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f'<b>Member:</b> {member.full_name} &nbsp;&nbsp; <b>Member No:</b> {member.member_no}', normal))
    story.append(Paragraph(f'<b>Phone:</b> {member.phone or "-"} &nbsp;&nbsp; <b>NRC/ID:</b> {member.national_id or "-"} &nbsp;&nbsp; <b>Group:</b> {member.group_name or "-"}', normal))
    story.append(Paragraph(f'<b>Status:</b> {member.status} &nbsp;&nbsp; <b>Generated:</b> {datetime.now().strftime("%d-%b-%Y %H:%M")}', normal))
    story.append(Spacer(1, 10))

    

    contributions = Contribution.query.filter_by(member_id=member.id).order_by(Contribution.paid_on.desc()).all()
    loans = Loan.query.filter_by(member_id=member.id).order_by(Loan.issued_on.desc()).all()
    distributions = Distribution.query.filter_by(member_id=member.id).order_by(Distribution.paid_on.desc()).all()
    fines = FinePenalty.query.filter_by(member_id=member.id).order_by(FinePenalty.fine_date.desc()).all()
    welfare_contribs = WelfareContribution.query.filter_by(member_id=member.id).order_by(WelfareContribution.paid_on.desc()).all()
    welfare_claims = WelfareClaim.query.filter_by(member_id=member.id).order_by(WelfareClaim.requested_on.desc()).all()
    total_contrib = money(sum((c.amount for c in contributions), Decimal('0.00')))
    total_principal = money(sum((l.principal for l in loans), Decimal('0.00')))
    total_interest = money(sum((l.interest_amount for l in loans), Decimal('0.00')))
    total_repaid = money(sum((l.total_paid for l in loans), Decimal('0.00')))
    total_balance = money(sum((l.balance for l in loans), Decimal('0.00')))
    total_distrib = money(sum((d.amount for d in distributions), Decimal('0.00')))
    total_fines = money(sum((f.amount for f in fines if f.status != 'Waived'), Decimal('0.00')))
    total_fine_paid = money(sum((f.total_paid for f in fines), Decimal('0.00')))
    total_fine_balance = money(sum((f.balance for f in fines if f.status != 'Waived'), Decimal('0.00')))
    total_welfare_contrib = money(sum((w.amount for w in welfare_contribs), Decimal('0.00')))
    total_welfare_paid = money(
    sum((w.amount_approved for w in welfare_claims if w.status == 'Paid'),
    Decimal('0.00'))
        )

    total_savings_interest = money(
        db.session.query(
            db.func.coalesce(db.func.sum(SavingsInterest.interest_amount), 0)
        )
        .filter(SavingsInterest.member_id == member.id)
        .scalar()
    )

    gross_savings_value = money(
        total_contrib + total_savings_interest
    )

    total_loan_interest_charged = money(
        db.session.query(
            db.func.coalesce(db.func.sum(LoanInterest.interest_amount), 0)
        )
        .filter(LoanInterest.member_id == member.id)
        .scalar()
    )

    adjusted_loan_balance = money(
        total_balance + total_loan_interest_charged
    )

    member_equity = money(
        gross_savings_value
        + total_distrib
        + total_welfare_paid
        - adjusted_loan_balance
        - total_fine_balance
    )

    equity_label = (
        'Member Equity / Surplus'
        if member_equity >= 0
        else 'Member Equity / Deficit'
    )

    equity_label = 'Member Equity / Surplus' if member_equity >= 0 else 'Member Equity / Deficit'

    summary = [
    ['Total Contributions', kwacha(total_contrib)],
    ['Savings Interest Earned', kwacha(total_savings_interest)],
    ['Gross Savings Value', kwacha(gross_savings_value)],

    ['Loan Principal', kwacha(total_principal)],
    ['Initial Loan Charged', kwacha(total_interest)],
    ['Loan Interest Charged', kwacha(total_loan_interest_charged)],
    ['Loan Repayments', kwacha(total_repaid)],
    ['Outstanding Loan Balance', kwacha(total_balance)],
    ['Adjusted Loan Balance', kwacha(adjusted_loan_balance)],

    ['Distributions Received', kwacha(total_distrib)],

    ['Fines Charged', kwacha(total_fines)],
    ['Fines Paid', kwacha(total_fine_paid)],
    ['Outstanding Fines', kwacha(total_fine_balance)],

    ['Welfare Contributions', kwacha(total_welfare_contrib)],
    ['Welfare Support Paid', kwacha(total_welfare_paid)],

    [equity_label, kwacha(member_equity)],
]
    equity_color = colors.green if member_equity >= 0 else colors.red
    
    summary_table = Table(summary, colWidths=[95*mm, 65*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e8f2f6')),
        ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#cccccc')),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 6),
        ('TEXTCOLOR', (0, len(summary)-1), (-1, len(summary)-1), equity_color),
    ]))
    story.append(Paragraph('<b>Account Summary</b>', normal))
    story.append(summary_table)
    story.append(Spacer(1, 12))

    equity_breakdown = [
    ['Gross Savings Value', kwacha(gross_savings_value)],
    ['+ Distributions Received', kwacha(total_distrib)],
    ['+ Welfare Support Paid', kwacha(total_welfare_paid)],
    ['- Adjusted Loan Balance', f'({kwacha(adjusted_loan_balance)})'],
    ['- Outstanding Fines', f'({kwacha(total_fine_balance)})'],
    [equity_label, kwacha(member_equity)],
]

    equity_table = Table(equity_breakdown, colWidths=[95*mm, 65*mm])

    equity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f2f6')),
        ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#f3edf8')),
        ('TEXTCOLOR', (0, 5), (-1, 5), equity_color),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    story.append(Paragraph('<b>Member Equity Breakdown</b>', normal))
    story.append(equity_table)
    story.append(Spacer(1, 12))

    def add_table(title, headers, rows, widths):
            story.append(Paragraph(f'<b>{title}</b>', normal))
            data = [headers] + (rows if rows else [['No records found'] + ['']*(len(headers)-1)])
            table = Table(data, colWidths=widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1f4f68')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#dddddd')),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('ALIGN', (-1,1), (-1,-1), 'RIGHT'),
                ('PADDING', (0,0), (-1,-1), 5),
            ]))
            story.append(table)
            story.append(Spacer(1, 10))

    add_table('Contributions', ['Month','Date','Method','Reference','Amount'], [[c.month, c.paid_on.strftime('%d-%b-%Y'), c.method, c.reference or '-', kwacha(c.amount)] for c in contributions], [24*mm, 28*mm, 30*mm, 48*mm, 30*mm])
    loan_rows = [[l.issued_on.strftime('%d-%b-%Y'), l.due_on.strftime('%d-%b-%Y'), kwacha(l.principal), kwacha(l.interest_amount), kwacha(l.total_due), kwacha(l.total_paid), kwacha(l.balance), l.status] for l in loans]
    add_table('Loans', ['Issued','Due','Principal','Interest','Total','Paid','Balance','Status'], loan_rows, [22*mm,22*mm,22*mm,22*mm,22*mm,22*mm,22*mm,20*mm])
    add_table('Distributions', ['Date','Method','Reference','Authorized By','Amount'], [[d.paid_on.strftime('%d-%b-%Y'), d.method, d.reference or '-', d.authorized_by or '-', kwacha(d.amount)] for d in distributions], [28*mm, 30*mm, 45*mm, 37*mm, 25*mm])
    add_table('Fines & Penalties', ['Date','Category','Amount','Paid','Balance','Status'], [[f.fine_date.strftime('%d-%b-%Y'), f.category, kwacha(f.amount), kwacha(f.total_paid), kwacha(f.balance), f.status] for f in fines], [28*mm, 45*mm, 25*mm, 25*mm, 25*mm, 25*mm])
    add_table('Welfare Contributions', ['Month','Date','Method','Reference','Amount'], [[w.month, w.paid_on.strftime('%d-%b-%Y'), w.method, w.reference or '-', kwacha(w.amount)] for w in welfare_contribs], [24*mm, 28*mm, 30*mm, 48*mm, 30*mm])
    add_table('Welfare Claims', ['Date','Category','Requested','Approved','Status','Paid On'], [[w.requested_on.strftime('%d-%b-%Y'), w.category, kwacha(w.amount_requested), kwacha(w.amount_approved), w.status, w.paid_on.strftime('%d-%b-%Y') if w.paid_on else '-'] for w in welfare_claims], [28*mm, 40*mm, 25*mm, 25*mm, 25*mm, 30*mm])

    story.append(Spacer(1, 10))
    story.append(
    Paragraph(
        f'This statement was generated from the {organization_name} system.',
        small_style
    )
        )
    doc.build(story)
    buffer.seek(0)
    log_audit('GENERATE_MEMBER_STATEMENT', 'Member', member.id, f'Member statement PDF generated for {member.member_no} - {member.full_name}')
    filename = f'{member.member_no}_{member.full_name.replace(" ", "_")}_statement.pdf'
    return send_file(buffer, as_attachment=True, download_name=secure_filename(filename), mimetype='application/pdf')

def get_shareout_cycle(start_month, end_month):
    return ShareOutCycle.query.filter_by(
        start_month=start_month,
        end_month=end_month,
    ).first()

def calculate_shareout_data(
    start_month,
    end_month,
    expenses=Decimal('0.00'),
    other_income=Decimal('0.00')
):
    """
    Calculate the complete Share-Out schedule and summary.

    This function is shared by:
    - Share-Out Calculator
    - Members Share-Out Schedule
    - Future PDF statements
    - Payment Register
    - Executive Share-Out Dashboard
    """

    expenses = money(expenses or 0)
    other_income = money(other_income or 0)

    start_date = datetime.strptime(
        start_month + '-01',
        '%Y-%m-%d'
    ).date()

    end_year, end_mon = [
        int(value)
        for value in end_month.split('-')
    ]

    end_date = (
        date(end_year, end_mon, 28)
        + timedelta(days=4)
    ).replace(day=1) - timedelta(days=1)

    contrib_rows = db.session.query(
        Member.id,
        Member.member_no,
        Member.full_name,
        Member.group_name,
        db.func.coalesce(
            db.func.sum(Contribution.amount),
            0
        )
    ).join(
        Contribution,
        Contribution.member_id == Member.id
    ).filter(
        Contribution.month >= start_month,
        Contribution.month <= end_month
    ).group_by(
        Member.id,
        Member.member_no,
        Member.full_name,
        Member.group_name
    ).order_by(
        Member.member_no
    ).all()

    total_contributions = money(
        sum(
            (money(row[4]) for row in contrib_rows),
            Decimal('0.00')
        )
    )

    total_savings_interest = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(SavingsInterest.interest_amount),
                0
            )
        )
        .filter(SavingsInterest.month >= start_month)
        .filter(SavingsInterest.month <= end_month)
        .scalar()
    )

    fines_paid_total = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(FinePayment.amount),
                0
            )
        )
        .filter(FinePayment.paid_on >= start_date)
        .filter(FinePayment.paid_on <= end_date)
        .scalar()
    )

    distributions_total = money(
        db.session.query(
            db.func.coalesce(
                db.func.sum(Distribution.amount),
                0
            )
        )
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .scalar()
    )

    shareout_fund = money(
        total_contributions
        + total_savings_interest
        + fines_paid_total
        + other_income
        - expenses
        - distributions_total
    )

    rows = []

    eligible_members = 0
    members_requiring_review = 0

    highest_shareout = Decimal('0.00')
    lowest_shareout = None

    total_gross_shareout = Decimal('0.00')
    total_deductions = Decimal('0.00')
    total_net_payable = Decimal('0.00')

    for (
        member_id,
        member_no,
        full_name,
        group_name,
        contributed
    ) in contrib_rows:

        contributed = money(contributed)

        savings_interest = money(
            db.session.query(
                db.func.coalesce(
                    db.func.sum(
                        SavingsInterest.interest_amount
                    ),
                    0
                )
            )
            .filter(SavingsInterest.member_id == member_id)
            .filter(SavingsInterest.month >= start_month)
            .filter(SavingsInterest.month <= end_month)
            .scalar()
        )

        gross_savings_value = money(
            contributed + savings_interest
        )

        percent = (
            Decimal('0.00')
            if total_contributions == 0
            else (
                contributed
                / total_contributions
                * Decimal('100')
            )
        )

        gross_shareout = (
            Decimal('0.00')
            if total_contributions == 0
            else money(
                contributed
                / total_contributions
                * shareout_fund
            )
        )

        loan_principal_interest = money(
            sum(
                (
                    loan.balance
                    for loan in Loan.query.filter_by(
                        member_id=member_id
                    ).all()
                    if loan.status != 'Rejected'
                ),
                Decimal('0.00')
            )
        )

        compounded_loan_interest = money(
            db.session.query(
                db.func.coalesce(
                    db.func.sum(
                        LoanInterest.interest_amount
                    ),
                    0
                )
            )
            .filter(LoanInterest.member_id == member_id)
            .filter(LoanInterest.month >= start_month)
            .filter(LoanInterest.month <= end_month)
            .scalar()
        )

        outstanding_loans = money(
            loan_principal_interest
            + compounded_loan_interest
        )

        fine_balance = money(
            sum(
                (
                    fine.balance
                    for fine in FinePenalty.query.filter_by(
                        member_id=member_id
                    ).all()
                    if fine.status != 'Waived'
                ),
                Decimal('0.00')
            )
        )

        member_deductions = money(
            outstanding_loans + fine_balance
        )

        net_payable = money(
            gross_shareout - member_deductions
        )

        if net_payable >= 0:
            eligible_members += 1
        else:
            members_requiring_review += 1

        highest_shareout = max(
            highest_shareout,
            gross_shareout
        )

        if (
            lowest_shareout is None
            or gross_shareout < lowest_shareout
        ):
            lowest_shareout = gross_shareout

        total_gross_shareout += gross_shareout
        total_deductions += member_deductions
        total_net_payable += net_payable

        rows.append({
            'member_id': member_id,
            'member_no': member_no,
            'full_name': full_name,
            'member_name': full_name,
            'member_number': member_no,
            'group_name': group_name or '-',

            'contributed': contributed,
            'contributions': contributed,

            'savings_interest': savings_interest,
            'gross_savings_value': gross_savings_value,

            'percent': percent.quantize(
                Decimal('0.01')
            ),
            'percentage_share': percent.quantize(
                Decimal('0.01')
            ),

            'profit_share': money(
                gross_shareout - contributed
            ),

            'gross_shareout': gross_shareout,

            'outstanding_loans': outstanding_loans,
            'loan_deduction': outstanding_loans,

            'fine_balance': fine_balance,
            'other_deductions': fine_balance,

            'total_deductions': member_deductions,

            'net_payable': net_payable,
            'net_shareout': net_payable,

            'net_status': (
                'surplus'
                if net_payable >= 0
                else 'loss'
            ),

            'payment_status': 'Pending',
        })

    total_gross_shareout = money(
        total_gross_shareout
    )

    total_deductions = money(
        total_deductions
    )

    total_net_payable = money(
        total_net_payable
    )

    average_shareout = money(
        Decimal('0.00')
        if eligible_members == 0
        else total_net_payable / eligible_members
    )

    average_contribution = money(
        Decimal('0.00')
        if len(contrib_rows) == 0
        else total_contributions / len(contrib_rows)
    )

    readiness_score = 100
    readiness_messages = []

    if members_requiring_review > 0:
        readiness_score -= 25
        readiness_messages.append(
            f'{members_requiring_review} member(s) require review.'
        )

    if shareout_fund <= 0:
        readiness_score = 0
        readiness_messages.append(
            'No distributable fund available.'
        )

    if eligible_members == 0:
        readiness_score = 0
        readiness_messages.append(
            'No eligible members.'
        )

    readiness_score = max(0, readiness_score)

    distribution_ready = readiness_score >= 75

    recommendation = (
        'Proceed with Distribution'
        if distribution_ready
        else 'Do Not Distribute Yet'
    )

    surplus = money(
        total_savings_interest
        + fines_paid_total
        + other_income
        - expenses
        - distributions_total
    )

    return {
        'rows': rows,
        'total_contributions': total_contributions,
        'total_savings_interest': total_savings_interest,
        'fines_paid_total': fines_paid_total,
        'distributions_total': distributions_total,
        'expenses': expenses,
        'other_income': other_income,
        'surplus': surplus,
        'shareout_fund': shareout_fund,
        'eligible_members': eligible_members,
        'members_requiring_review': members_requiring_review,
        'highest_shareout': money(highest_shareout),
        'lowest_shareout': money(
            lowest_shareout or Decimal('0.00')
        ),
        'total_gross_shareout': total_gross_shareout,
        'total_deductions': total_deductions,
        'total_net_payable': total_net_payable,
        'average_shareout': average_shareout,
        'average_contribution': average_contribution,
        'readiness_score': readiness_score,
        'readiness_messages': readiness_messages,
        'distribution_ready': distribution_ready,
        'recommendation': recommendation,
    }
    


@app.route('/shareout', methods=['GET', 'POST'])
@login_required
@role_required('shareout')
def shareout():
    setting = SystemSetting.query.first()
    organization_name = setting.organisation_name if setting and setting.organisation_name else CLIENT_NAME

    today_month = date.today().strftime('%Y-%m')
    start_month = request.values.get('start_month') or f'{date.today().year}-01'
    end_month = request.values.get('end_month') or today_month

    expenses = money(request.values.get('expenses') or 0)
    other_income = money(request.values.get('other_income') or 0)

    cycle = get_shareout_cycle(
        start_month,
        end_month,
    )

    cycle_locked = bool(
        cycle and cycle.status == 'Locked'
    )

    start_date = datetime.strptime(start_month + '-01', '%Y-%m-%d').date()
    end_year, end_mon = [int(x) for x in end_month.split('-')]
    end_date = (date(end_year, end_mon, 28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    contrib_rows = db.session.query(
        Member.id,
        Member.member_no,
        Member.full_name,
        Member.group_name,
        db.func.coalesce(db.func.sum(Contribution.amount), 0)
    ).join(
        Contribution, Contribution.member_id == Member.id
    ).filter(
        Contribution.month >= start_month,
        Contribution.month <= end_month
    ).group_by(
        Member.id,
        Member.member_no,
        Member.full_name,
        Member.group_name
    ).order_by(
        Member.member_no
    ).all()

    total_contributions = money(
        sum((money(row[4]) for row in contrib_rows), Decimal('0.00'))
    )

    total_savings_interest = money(
        db.session.query(db.func.coalesce(db.func.sum(SavingsInterest.interest_amount), 0))
        .filter(SavingsInterest.month >= start_month)
        .filter(SavingsInterest.month <= end_month)
        .scalar()
    )

    fines_paid_total = money(
        db.session.query(db.func.coalesce(db.func.sum(FinePayment.amount), 0))
        .filter(FinePayment.paid_on >= start_date)
        .filter(FinePayment.paid_on <= end_date)
        .scalar()
    )

    distributions_total = money(
        db.session.query(db.func.coalesce(db.func.sum(Distribution.amount), 0))
        .filter(Distribution.paid_on >= start_date)
        .filter(Distribution.paid_on <= end_date)
        .scalar()
    )

    shareout_fund = money(
        total_contributions
        + total_savings_interest
        + fines_paid_total
        + other_income
        - expenses
        - distributions_total
    )
    eligible_members = 0
    members_requiring_review = 0

    highest_shareout = Decimal("0.00")
    lowest_shareout = None

    total_net_payable = Decimal("0.00")

    readiness_score = 100
    readiness_messages = []

    rows = []

    for member_id, member_no, full_name, group_name, contributed in contrib_rows:
        contributed = money(contributed)

        savings_interest = money(
            db.session.query(db.func.coalesce(db.func.sum(SavingsInterest.interest_amount), 0))
            .filter(SavingsInterest.member_id == member_id)
            .filter(SavingsInterest.month >= start_month)
            .filter(SavingsInterest.month <= end_month)
            .scalar()
        )

        gross_savings_value = money(contributed + savings_interest)

        percent = Decimal('0.00') if total_contributions == 0 else (
            contributed / total_contributions * Decimal('100')
        )

        gross_shareout = Decimal('0.00') if total_contributions == 0 else money(
            contributed / total_contributions * shareout_fund
        )

        loan_principal_interest = money(
            sum(
                (
                    l.balance
                    for l in Loan.query.filter_by(member_id=member_id).all()
                    if l.status != 'Rejected'
                ),
                Decimal('0.00')
            )
        )

        compounded_loan_interest = money(
            db.session.query(db.func.coalesce(db.func.sum(LoanInterest.interest_amount), 0))
            .filter(LoanInterest.member_id == member_id)
            .filter(LoanInterest.month >= start_month)
            .filter(LoanInterest.month <= end_month)
            .scalar()
        )

        outstanding_loans = money(loan_principal_interest + compounded_loan_interest)

        fine_balance = money(
            sum(
                (
                    f.balance
                    for f in FinePenalty.query.filter_by(member_id=member_id).all()
                    if f.status != 'Waived'
                ),
                Decimal('0.00')
            )
        )

        total_deductions = money(outstanding_loans + fine_balance)

        net_payable = money(gross_shareout - total_deductions)
        total_net_payable += net_payable

        if net_payable >= 0:
            eligible_members += 1
        else:
            members_requiring_review += 1

        if gross_shareout > highest_shareout:
            highest_shareout = gross_shareout

        if lowest_shareout is None or gross_shareout < lowest_shareout:
            lowest_shareout = gross_shareout

        net_status = 'surplus' if net_payable >= 0 else 'loss'

        rows.append({
            'member_no': member_no,
            'full_name': full_name,
            'group_name': group_name or '-',
            'contributed': contributed,
            'savings_interest': savings_interest,
            'gross_savings_value': gross_savings_value,
            'percent': percent.quantize(Decimal('0.01')),
            'gross_shareout': gross_shareout,
            'outstanding_loans': outstanding_loans,
            'fine_balance': fine_balance,
            'total_deductions': total_deductions,
            'net_payable': net_payable,
            'net_status': net_status,
        })
        average_shareout = money(
        Decimal('0.00')
        if eligible_members == 0
        else total_net_payable / eligible_members
    )

    average_contribution = money(
        Decimal('0.00')
        if len(contrib_rows) == 0
        else total_contributions / len(contrib_rows)
    )

    if members_requiring_review > 0:
        readiness_score -= 25
        readiness_messages.append(
            f'{members_requiring_review} member(s) require review.'
        )

    if shareout_fund <= 0:
        readiness_score = 0
        readiness_messages.append(
            'No distributable fund available.'
        )

    if eligible_members == 0:
        readiness_score = 0
        readiness_messages.append(
            'No eligible members.'
        )

    readiness_score = max(0, readiness_score)
    distribution_ready = readiness_score >= 75

    recommendation = (
        'Proceed with Distribution'
        if distribution_ready
        else 'Do Not Distribute Yet'
    )    
    page = request.args.get('page', 1, type=int)
    per_page = 25

    total_rows = len(rows)
    total_pages = (total_rows + per_page - 1) // per_page if total_rows else 1

    if page < 1:
        page = 1
    elif page > total_pages:
        page = total_pages

    start = (page - 1) * per_page
    end = start + per_page

    paged_rows = rows[start:end] 

    if request.method == 'POST' and cycle_locked:
        flash(
            'This Share-Out cycle is locked and cannot be recalculated.',
            'error'
        )

        return redirect(
            url_for(
                'shareout',
                start_month=start_month,
                end_month=end_month,
                expenses=expenses,
                other_income=other_income,
            )
        )   
   
    if request.method == 'POST':
        log_audit(
            'CALCULATE_SHAREOUT',
            'ShareOut',
            None,
            f'Share-out calculated for {start_month} to {end_month}; fund {kwacha(shareout_fund)}'
        )

    return render_template('shareout.html', **locals())

@app.route('/share-out-control-centre')
@login_required
def shareout_control_centre():
    """
    Central navigation and monitoring hub for all Share-Out functions.
    """

    shareout_summary = {
        'total_shareout': 0,
        'eligible_members': 0,
        'members_paid': 0,
        'members_unpaid': 0,
        'amount_paid': 0,
        'balance_remaining': 0,
        'payment_progress': 0,
        'approval_status': 'Not Started',
        'is_locked': False,
    }

    return render_template(
        'shareout/control_centre.html',
        shareout_summary=shareout_summary
    )
@app.route('/share-out-schedule')
@login_required
@role_required('shareout')
def shareout_schedule():
    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.args.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.args.get('end_month')
        or today_month
    )

    expenses = money(
        request.args.get('expenses') or 0
    )

    other_income = money(
        request.args.get('other_income') or 0
    )

    search = request.args.get(
        'search',
        ''
    ).strip()

    payment_status = request.args.get(
        'payment_status',
        ''
    ).strip()

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    schedule_rows = shareout_data['rows']

    if search:
        search_lower = search.lower()

        schedule_rows = [
            row
            for row in schedule_rows
            if (
                search_lower
                in row['member_no'].lower()
                or search_lower
                in row['full_name'].lower()
            )
        ]

    if payment_status:
        schedule_rows = [
            row
            for row in schedule_rows
            if row['payment_status'] == payment_status
        ]

    members_paid = sum(
        1
        for row in schedule_rows
        if row['payment_status'] == 'Paid'
    )

    members_pending = sum(
        1
        for row in schedule_rows
        if row['payment_status'] == 'Pending'
    )

    schedule_summary = {
        'eligible_members':
            shareout_data['eligible_members'],

        'total_contributions':
            shareout_data['total_contributions'],

        'total_profit_share': money(
            shareout_data['total_gross_shareout']
            - shareout_data['total_contributions']
        ),

        'total_gross_shareout':
            shareout_data['total_gross_shareout'],

        'total_deductions':
            shareout_data['total_deductions'],

        'total_net_shareout':
            shareout_data['total_net_payable'],

        'average_shareout':
            shareout_data['average_shareout'],

        'members_paid': members_paid,
        'members_pending': members_pending,
    }

    return render_template(
        'shareout/members_schedule.html',
        schedule_rows=schedule_rows,
        schedule_summary=schedule_summary,
        search=search,
        payment_status=payment_status,
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

@app.route('/shareout.pdf')
@login_required
@role_required('shareout')
def shareout_pdf():
    setting = SystemSetting.query.first()
    organization_name = setting.organisation_name if setting and setting.organisation_name else CLIENT_NAME
    organization_address = setting.organization_address if setting and setting.organization_address else ''
    organization_phone = setting.organization_phone if setting and setting.organization_phone else ''
    organization_email = setting.organization_email if setting and setting.organization_email else ''
    registration_number = setting.registration_number if setting and setting.registration_number else ''

    today_month = date.today().strftime('%Y-%m')
    start_month = request.args.get('start_month') or f'{date.today().year}-01'
    end_month = request.args.get('end_month') or today_month
    expenses = money(request.args.get('expenses') or 0)
    other_income = money(request.args.get('other_income') or 0)

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    rows = shareout_data['rows']
    total_contributions = shareout_data[
        'total_contributions'
    ]
    total_savings_interest = shareout_data[
        'total_savings_interest'
    ]
    fines_paid_total = shareout_data[
        'fines_paid_total'
    ]
    distributions_total = shareout_data[
        'distributions_total'
    ]
    surplus = shareout_data['surplus']
    shareout_fund = shareout_data['shareout_fund']

    eligible_members = shareout_data[
        'eligible_members'
    ]

    members_requiring_review = shareout_data[
        'members_requiring_review'
    ]

    highest_shareout = shareout_data[
        'highest_shareout'
    ]

    lowest_shareout = shareout_data[
        'lowest_shareout'
    ]

    total_net_payable = shareout_data[
        'total_net_payable'
    ]

    average_shareout = shareout_data[
        'average_shareout'
    ]

    average_contribution = shareout_data[
        'average_contribution'
    ]

    readiness_score = shareout_data[
        'readiness_score'
    ]

    readiness_messages = shareout_data[
        'readiness_messages'
    ]

    distribution_ready = shareout_data[
        'distribution_ready'
    ]

    recommendation = shareout_data[
        'recommendation'
    ]
    primary_colour = pdf_colour(
        setting.primary_color if setting else None,
        '#0D6EFD'
    )

    secondary_colour = pdf_colour(
        setting.secondary_color if setting else None,
        '#198754'
    )

    table_header_colour = pdf_colour(
        setting.table_header_color if setting else None,
        '#EAF2F8'
    )

    success_colour = pdf_colour(
        setting.success_color if setting else None,
        '#198754'
    )

    warning_colour = pdf_colour(
        setting.warning_color if setting else None,
        '#FFC107'
    )

    danger_colour = pdf_colour(
        setting.danger_color if setting else None,
        '#DC3545'
    )

    buffer = io.BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=17 * mm,
        title=(
            f'Members Share-Out Schedule '
            f'{start_month} to {end_month}'
        ),
        author=organization_name,
    )

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'ShareOutScheduleTitle',
        parent=styles['Heading1'],
        fontSize=15,
        leading=18,
        textColor=primary_colour,
        spaceBefore=3,
        spaceAfter=7,
    )

    section_style = ParagraphStyle(
        'ShareOutScheduleSection',
        parent=styles['Heading2'],
        fontSize=10,
        leading=12,
        textColor=primary_colour,
        spaceBefore=6,
        spaceAfter=6,
    )

    small_style = ParagraphStyle(
        'ShareOutScheduleSmall',
        parent=styles['Normal'],
        fontSize=6.5,
        leading=8,
    )

    small_bold_style = ParagraphStyle(
        'ShareOutScheduleSmallBold',
        parent=small_style,
        fontName='Helvetica-Bold',
    )

    centred_small_style = ParagraphStyle(
        'ShareOutScheduleCentredSmall',
        parent=small_style,
        alignment=1,
    )

    elements.append(
        build_pdf_branding(
            setting,
            styles
        )
    )

    elements.append(Spacer(1, 6))

    elements.append(
        Paragraph(
            'MEMBERS SHARE-OUT SCHEDULE',
            title_style
        )
    )

    period_data = [
        [
            'Share-Out Period',
            f'{start_month} to {end_month}',
            'Cycle Status',
            (
                get_shareout_cycle(
                    start_month,
                    end_month
                ).status
                if get_shareout_cycle(
                    start_month,
                    end_month
                )
                else 'Draft'
            ),
        ]
    ]

    period_table = Table(
        period_data,
        colWidths=[
            32 * mm,
            55 * mm,
            29 * mm,
            42 * mm,
        ],
        hAlign='LEFT',
    )

    period_table.setStyle(
        TableStyle([
            (
                'BACKGROUND',
                (0, 0),
                (0, -1),
                table_header_colour
            ),
            (
                'BACKGROUND',
                (2, 0),
                (2, -1),
                table_header_colour
            ),
            (
                'FONTNAME',
                (0, 0),
                (0, -1),
                'Helvetica-Bold'
            ),
            (
                'FONTNAME',
                (2, 0),
                (2, -1),
                'Helvetica-Bold'
            ),
            (
                'FONTSIZE',
                (0, 0),
                (-1, -1),
                7.5
            ),
            (
                'GRID',
                (0, 0),
                (-1, -1),
                0.4,
                primary_colour
            ),
            (
                'VALIGN',
                (0, 0),
                (-1, -1),
                'MIDDLE'
            ),
            (
                'TOPPADDING',
                (0, 0),
                (-1, -1),
                5
            ),
            (
                'BOTTOMPADDING',
                (0, 0),
                (-1, -1),
                5
            ),
        ])
    )

    elements.append(period_table)
    elements.append(Spacer(1, 8))

    summary_data = [
        [
            'Total Contributions',
            kwacha(total_contributions),
            'Savings Interest',
            kwacha(total_savings_interest),
            'Fines Paid',
            kwacha(fines_paid_total),
        ],
        [
            'Other Income',
            kwacha(other_income),
            'Expenses',
            kwacha(expenses),
            'Distributions',
            kwacha(distributions_total),
        ],
        [
            'Surplus / Income',
            kwacha(surplus),
            'Share-Out Fund',
            kwacha(shareout_fund),
            'Net Payable',
            kwacha(total_net_payable),
        ],
        [
            'Eligible Members',
            str(eligible_members),
            'Require Review',
            str(members_requiring_review),
            'Readiness Score',
            f'{readiness_score}%',
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            34 * mm,
            31 * mm,
            30 * mm,
            31 * mm,
            29 * mm,
            31 * mm,
        ],
        hAlign='LEFT',
    )

    summary_table.setStyle(
        TableStyle([
            (
                'BACKGROUND',
                (0, 0),
                (0, -1),
                table_header_colour
            ),
            (
                'BACKGROUND',
                (2, 0),
                (2, -1),
                table_header_colour
            ),
            (
                'BACKGROUND',
                (4, 0),
                (4, -1),
                table_header_colour
            ),
            (
                'FONTNAME',
                (0, 0),
                (0, -1),
                'Helvetica-Bold'
            ),
            (
                'FONTNAME',
                (2, 0),
                (2, -1),
                'Helvetica-Bold'
            ),
            (
                'FONTNAME',
                (4, 0),
                (4, -1),
                'Helvetica-Bold'
            ),
            (
                'FONTSIZE',
                (0, 0),
                (-1, -1),
                7
            ),
            (
                'ALIGN',
                (1, 0),
                (1, -1),
                'RIGHT'
            ),
            (
                'ALIGN',
                (3, 0),
                (3, -1),
                'RIGHT'
            ),
            (
                'ALIGN',
                (5, 0),
                (5, -1),
                'RIGHT'
            ),
            (
                'GRID',
                (0, 0),
                (-1, -1),
                0.35,
                colors.HexColor('#AAB7C4')
            ),
            (
                'VALIGN',
                (0, 0),
                (-1, -1),
                'MIDDLE'
            ),
            (
                'TOPPADDING',
                (0, 0),
                (-1, -1),
                4
            ),
            (
                'BOTTOMPADDING',
                (0, 0),
                (-1, -1),
                4
            ),
        ])
    )

    elements.append(summary_table)
    elements.append(Spacer(1, 7))

    readiness_background = (
        success_colour
        if distribution_ready
        else warning_colour
    )

    readiness_text_colour = (
        colors.white
        if distribution_ready
        else colors.black
    )

    readiness_table = Table(
        [[
            'Executive Recommendation',
            recommendation,
            'Readiness',
            f'{readiness_score}%',
        ]],
        colWidths=[
            40 * mm,
            95 * mm,
            24 * mm,
            27 * mm,
        ],
    )

    readiness_table.setStyle(
        TableStyle([
            (
                'BACKGROUND',
                (0, 0),
                (-1, -1),
                readiness_background
            ),
            (
                'TEXTCOLOR',
                (0, 0),
                (-1, -1),
                readiness_text_colour
            ),
            (
                'FONTNAME',
                (0, 0),
                (0, 0),
                'Helvetica-Bold'
            ),
            (
                'FONTNAME',
                (2, 0),
                (2, 0),
                'Helvetica-Bold'
            ),
            (
                'FONTNAME',
                (1, 0),
                (1, 0),
                'Helvetica-Bold'
            ),
            (
                'FONTSIZE',
                (0, 0),
                (-1, -1),
                7.5
            ),
            (
                'GRID',
                (0, 0),
                (-1, -1),
                0.4,
                primary_colour
            ),
            (
                'VALIGN',
                (0, 0),
                (-1, -1),
                'MIDDLE'
            ),
            (
                'TOPPADDING',
                (0, 0),
                (-1, -1),
                5
            ),
            (
                'BOTTOMPADDING',
                (0, 0),
                (-1, -1),
                5
            ),
        ])
    )

    elements.append(readiness_table)

    if readiness_messages:
        elements.append(Spacer(1, 4))

        for message in readiness_messages:
            elements.append(
                Paragraph(
                    f'• {message}',
                    small_style
                )
            )

    elements.append(Spacer(1, 7))

    schedule_data = [[
        'Member',
        'Contributions',
        'Savings\nInterest',
        'Gross\nSavings',
        'Share %',
        'Gross\nShare-Out',
        'Loans +\nInterest',
        'Fines',
        'Total\nDeductions',
        'Net\nPayable',
    ]]

    for row in rows:
        schedule_data.append([
            Paragraph(
                (
                    f"<b>{row['member_no']}</b><br/>"
                    f"{row['full_name']}"
                ),
                small_style
            ),
            kwacha(row['contributed']),
            kwacha(row['savings_interest']),
            kwacha(row['gross_savings_value']),
            f"{row['percent']}%",
            kwacha(row['gross_shareout']),
            kwacha(row['outstanding_loans']),
            kwacha(row['fine_balance']),
            kwacha(row['total_deductions']),
            kwacha(row['net_shareout']),
        ])

    if not rows:
        schedule_data.append([
            Paragraph(
                'No Share-Out records found for this period.',
                small_style
            ),
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
        ])

    schedule_table = Table(
        schedule_data,
        repeatRows=1,
        colWidths=[
            36 * mm,
            23 * mm,
            22 * mm,
            23 * mm,
            17 * mm,
            24 * mm,
            25 * mm,
            19 * mm,
            24 * mm,
            25 * mm,
        ],
        hAlign='LEFT',
    )

    schedule_style = [
        (
            'BACKGROUND',
            (0, 0),
            (-1, 0),
            primary_colour
        ),
        (
            'TEXTCOLOR',
            (0, 0),
            (-1, 0),
            colors.white
        ),
        (
            'FONTNAME',
            (0, 0),
            (-1, 0),
            'Helvetica-Bold'
        ),
        (
            'FONTSIZE',
            (0, 0),
            (-1, 0),
            6
        ),
        (
            'FONTSIZE',
            (0, 1),
            (-1, -1),
            5.8
        ),
        (
            'ALIGN',
            (1, 1),
            (-1, -1),
            'RIGHT'
        ),
        (
            'VALIGN',
            (0, 0),
            (-1, -1),
            'MIDDLE'
        ),
        (
            'GRID',
            (0, 0),
            (-1, -1),
            0.3,
            colors.HexColor('#AAB7C4')
        ),
        (
            'ROWBACKGROUNDS',
            (0, 1),
            (-1, -1),
            [
                colors.white,
                table_header_colour,
            ]
        ),
        (
            'TOPPADDING',
            (0, 0),
            (-1, -1),
            3
        ),
        (
            'BOTTOMPADDING',
            (0, 0),
            (-1, -1),
            3
        ),
    ]

    for row_number, row in enumerate(
        rows,
        start=1
    ):
        net_payable = money(
            row['net_shareout']
        )

        if net_payable < 0:
            schedule_style.append(
                (
                    'TEXTCOLOR',
                    (9, row_number),
                    (9, row_number),
                    danger_colour
                )
            )

        else:
            schedule_style.append(
                (
                    'TEXTCOLOR',
                    (9, row_number),
                    (9, row_number),
                    success_colour
                )
            )

        schedule_style.append(
            (
                'FONTNAME',
                (9, row_number),
                (9, row_number),
                'Helvetica-Bold'
            )
        )

    schedule_table.setStyle(
        TableStyle(schedule_style)
    )

    elements.append(schedule_table)
    elements.append(Spacer(1, 8))

    totals_data = [[
        'Grand Totals',
        kwacha(total_contributions),
        kwacha(total_savings_interest),
        '',
        '',
        kwacha(
            sum(
                (
                    money(row['gross_shareout'])
                    for row in rows
                ),
                Decimal('0.00')
            )
        ),
        kwacha(
            sum(
                (
                    money(row['outstanding_loans'])
                    for row in rows
                ),
                Decimal('0.00')
            )
        ),
        kwacha(
            sum(
                (
                    money(row['fine_balance'])
                    for row in rows
                ),
                Decimal('0.00')
            )
        ),
        kwacha(
            sum(
                (
                    money(row['total_deductions'])
                    for row in rows
                ),
                Decimal('0.00')
            )
        ),
        kwacha(total_net_payable),
    ]]

    totals_table = Table(
        totals_data,
        colWidths=[
            36 * mm,
            23 * mm,
            22 * mm,
            23 * mm,
            17 * mm,
            24 * mm,
            25 * mm,
            19 * mm,
            24 * mm,
            25 * mm,
        ],
    )

    totals_table.setStyle(
        TableStyle([
            (
                'BACKGROUND',
                (0, 0),
                (-1, -1),
                secondary_colour
            ),
            (
                'TEXTCOLOR',
                (0, 0),
                (-1, -1),
                colors.white
            ),
            (
                'FONTNAME',
                (0, 0),
                (-1, -1),
                'Helvetica-Bold'
            ),
            (
                'FONTSIZE',
                (0, 0),
                (-1, -1),
                6
            ),
            (
                'ALIGN',
                (1, 0),
                (-1, -1),
                'RIGHT'
            ),
            (
                'GRID',
                (0, 0),
                (-1, -1),
                0.35,
                colors.white
            ),
            (
                'TOPPADDING',
                (0, 0),
                (-1, -1),
                4
            ),
            (
                'BOTTOMPADDING',
                (0, 0),
                (-1, -1),
                4
            ),
        ])
    )

    elements.append(totals_table)
    elements.append(Spacer(1, 13))

    elements.append(
        Paragraph(
            'Certification and Approval',
            section_style
        )
    )

    certification_data = [
        [
            '________________________',
            '________________________',
            '________________________',
            '________________________',
        ],
        [
            'Prepared By',
            'Verified By',
            'Treasurer',
            'Chairperson',
        ],
        [
            '',
            '',
            '',
            '',
        ],
        [
            'Date: __________________',
            'Date: __________________',
            'Date: __________________',
            'Date: __________________',
        ],
    ]

    certification_table = Table(
        certification_data,
        colWidths=[
            47 * mm,
            47 * mm,
            47 * mm,
            47 * mm,
        ],
    )

    certification_table.setStyle(
        TableStyle([
            (
                'ALIGN',
                (0, 0),
                (-1, -1),
                'CENTER'
            ),
            (
                'FONTSIZE',
                (0, 0),
                (-1, -1),
                7
            ),
            (
                'TOPPADDING',
                (0, 0),
                (-1, -1),
                3
            ),
            (
                'BOTTOMPADDING',
                (0, 0),
                (-1, -1),
                3
            ),
        ])
    )

    elements.append(certification_table)

    generated_on = datetime.now().strftime(
        '%d %B %Y at %H:%M'
    )

    elements.append(Spacer(1, 7))

    elements.append(
        Paragraph(
            f'Generated on {generated_on}',
            small_style
        )
    )

    document.build(
        elements,
        onFirstPage=lambda canvas, doc: draw_pdf_footer(
            canvas,
            doc,
            setting
        ),
        onLaterPages=lambda canvas, doc: draw_pdf_footer(
            canvas,
            doc,
            setting
        ),
    )

    buffer.seek(0)

    log_audit(
        'EXPORT_SHAREOUT_PDF',
        'ShareOut',
        None,
        (
            f'Branded Members Share-Out Schedule PDF '
            f'exported for {start_month} to {end_month}'
        )
    )

    filename = (
        f'members_shareout_schedule_'
        f'{start_month}_to_{end_month}.pdf'
    )

    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={
            'Content-Disposition':
                f'inline; filename="{filename}"'
        }
    )   
@app.route('/shareout.csv')
@login_required
@role_required('shareout')
def shareout_csv():
    setting = SystemSetting.query.first()

    organization_name = (
        setting.organisation_name
        if setting and setting.organisation_name
        else CLIENT_NAME
    )

    today_month = date.today().strftime('%Y-%m')

    start_month = (
        request.args.get('start_month')
        or f'{date.today().year}-01'
    )

    end_month = (
        request.args.get('end_month')
        or today_month
    )

    expenses = money(
        request.args.get('expenses') or 0
    )

    other_income = money(
        request.args.get('other_income') or 0
    )

    shareout_data = calculate_shareout_data(
        start_month=start_month,
        end_month=end_month,
        expenses=expenses,
        other_income=other_income,
    )

    rows = shareout_data['rows']

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([organization_name])
    writer.writerow([
        f'Share-Out Report: {start_month} to {end_month}'
    ])
    writer.writerow([])

    writer.writerow([
        'Member No',
        'Full Name',
        'Group',
        'Contributions',
        'Savings Interest',
        'Gross Savings Value',
        'Contribution %',
        'Gross Share-Out',
        'Outstanding Loans',
        'Outstanding Fines',
        'Total Deductions',
        'Net Payable',
    ])

    for row in rows:
        writer.writerow([
            row['member_no'],
            row['full_name'],
            row['group_name'],
            row['contributed'],
            row['savings_interest'],
            row['gross_savings_value'],
            row['percent'],
            row['gross_shareout'],
            row['outstanding_loans'],
            row['fine_balance'],
            row['total_deductions'],
            row['net_payable'],
        ])

    log_audit(
        'EXPORT_SHAREOUT',
        'ShareOut',
        None,
        (
            f'Share-out CSV exported for '
            f'{start_month} to {end_month}'
        )
    )

    filename = (
        f'shareout_{start_month}_to_{end_month}.csv'
    )

    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition':
                f'attachment; filename="{filename}"'
        }
    )



@app.route('/month-end', methods=['GET', 'POST'])
@login_required
@role_required('accounting')
def month_end():
    selected_month = request.values.get('month') or date.today().strftime('%Y-%m')
    setting = SystemSetting.query.first()

    if not setting:
        setting = SystemSetting()
        db.session.add(setting)
        db.session.commit()

    rate = Decimal(setting.savings_interest_rate or 15) / Decimal('100')
    loan_rate = Decimal(setting.loan_interest_rate or 15) / Decimal('100')

    if request.method == 'POST':
        existing = MonthEndProcess.query.filter_by(month=selected_month).first()

        if existing:
            flash(f'Month-end interest has already been processed for {selected_month}. Reverse it first if you want to reprocess.', 'error')
            return redirect(url_for('month_end', month=selected_month))

        savings_total = Decimal('0.00')
        loan_interest_total = Decimal('0.00')
        members_processed = 0
        loans_processed = 0

        SavingsInterest.query.filter_by(month=selected_month).delete()
        LoanInterest.query.filter_by(month=selected_month).delete()
        db.session.commit()

        members = Member.query.filter_by(status='Active').all()

        for member in members:
            total_contributions = money(
                db.session.query(db.func.coalesce(db.func.sum(Contribution.amount), 0))
                .filter(Contribution.member_id == member.id)
                .scalar()
            )

            previous_interest = money(
                db.session.query(db.func.coalesce(db.func.sum(SavingsInterest.interest_amount), 0))
                .filter(SavingsInterest.member_id == member.id)
                .filter(SavingsInterest.month < selected_month)
                .scalar()
            )

            distributions_total = money(
                db.session.query(db.func.coalesce(db.func.sum(Distribution.amount), 0))
                .filter(Distribution.member_id == member.id)
                .scalar()
            )

            opening_balance = money(total_contributions + previous_interest - distributions_total)

            if opening_balance > 0:
                interest_amount = money(opening_balance * rate)
                closing_balance = money(opening_balance + interest_amount)

                db.session.add(SavingsInterest(
                    member_id=member.id,
                    month=selected_month,
                    opening_balance=opening_balance,
                    interest_rate=setting.savings_interest_rate,
                    interest_amount=interest_amount,
                    closing_balance=closing_balance
                ))

                savings_total += interest_amount
                members_processed += 1

        loans = Loan.query.filter(
            Loan.status.in_(['Disbursed', 'Partially Paid'])
        ).all()

        for loan in loans:
            previous_interest = money(
                db.session.query(db.func.coalesce(db.func.sum(LoanInterest.interest_amount), 0))
                .filter(LoanInterest.loan_id == loan.id)
                .filter(LoanInterest.month < selected_month)
                .scalar()
            )

            opening_balance = money(loan.balance + previous_interest)

            if opening_balance > 0:
                interest_amount = money(opening_balance * loan_rate)
                closing_balance = money(opening_balance + interest_amount)

                db.session.add(LoanInterest(
                    loan_id=loan.id,
                    member_id=loan.member_id,
                    month=selected_month,
                    opening_balance=opening_balance,
                    interest_rate=setting.loan_interest_rate,
                    interest_amount=interest_amount,
                    closing_balance=closing_balance
                ))

                loan_interest_total += interest_amount
                loans_processed += 1

        user = session.get('user') or {}

        process = MonthEndProcess(
            month=selected_month,
            savings_interest_total=savings_total,
            loan_interest_total=loan_interest_total,
            members_processed=members_processed,
            loans_processed=loans_processed,
            reversed=False,
            processed_by=user.get('full_name') or user.get('username')
        )

        db.session.add(process)
        db.session.commit()

        log_audit(
            'MONTH_END_PROCESS',
            'MonthEndProcess',
            process.id,
            f'Month-end processed for {selected_month}. Savings interest: {kwacha(savings_total)}, Loan interest: {kwacha(loan_interest_total)}, Members processed: {members_processed}, Loans processed: {loans_processed}'
        )

        flash(f'Month-end processed for {selected_month}. Members: {members_processed}, Loans: {loans_processed}.')
        return redirect(url_for('month_end', month=selected_month))

    page = request.args.get('page', 1, type=int)
    per_page = 25

    pagination = MonthEndProcess.query.order_by(
    MonthEndProcess.month.desc(),
    MonthEndProcess.id.desc()
    ).paginate(
    page=page,
    per_page=per_page,
    error_out=False
    )

    processes = pagination.items

    reversal_page = request.args.get('reversal_page', 1, type=int)

    reversal_pagination = AuditLog.query.filter_by(
        action='REVERSE_MONTH_END'
    ).order_by(
        AuditLog.created_at.desc()
    ).paginate(
        page=reversal_page,
        per_page=10,
        error_out=False
    )

    reversal_logs = reversal_pagination.items

    return render_template(
        'month_end.html',
        selected_month=selected_month,
        processes=processes,
        pagination=pagination,
        reversal_logs=reversal_logs,
        reversal_pagination=reversal_pagination
    )


@app.route('/month-end/<month>')
@login_required
@role_required('accounting')
def month_end_details(month):
    process = MonthEndProcess.query.filter_by(month=month).first_or_404()

    savings_page = request.args.get('savings_page', 1, type=int)
    loans_page = request.args.get('loans_page', 1, type=int)
    per_page = 25

    savings_pagination = SavingsInterest.query.filter_by(month=month).order_by(
        SavingsInterest.id.desc()
    ).paginate(
        page=savings_page,
        per_page=per_page,
        error_out=False
    )

    loan_pagination = LoanInterest.query.filter_by(month=month).order_by(
        LoanInterest.id.desc()
    ).paginate(
        page=loans_page,
        per_page=per_page,
        error_out=False
    )

    return render_template(
        'month_end_details.html',
        process=process,
        savings_entries=savings_pagination.items,
        loan_entries=loan_pagination.items,
        savings_pagination=savings_pagination,
        loan_pagination=loan_pagination
    )

@app.route('/month-end/<month>/reverse', methods=['POST'])
@login_required
@role_required('accounting')
def month_end_reverse(month):
    reason = request.form.get('reason', '').strip()

    if not reason:
        flash('Please provide a reason for reversing month-end.', 'error')
        return redirect(url_for('month_end_details', month=month))

    process = MonthEndProcess.query.filter_by(month=month).first_or_404()

    SavingsInterest.query.filter_by(month=month).delete()
    LoanInterest.query.filter_by(month=month).delete()
    db.session.delete(process)

    db.session.commit()

    log_audit(
        'REVERSE_MONTH_END',
        'MonthEndProcess',
        None,
        f'Month-end reversed for {month}. Reason: {reason}'
    )

    flash(f'Month-end processing for {month} has been reversed. You can now process it again.')
    return redirect(url_for('month_end'))

@app.route('/committee')
@login_required
@role_required('members')
def committee_register():
    committee_members = Member.query.filter(
        Member.member_type == 'Committee Member'
    ).order_by(
        Member.committee_position,
        Member.member_no
    ).all()

    return render_template(
        'committee.html',
        committee_members=committee_members
    )



@app.route('/export/<kind>.csv')
@login_required
@role_required('exports')
def export_csv(kind):
    output = io.StringIO(); writer = csv.writer(output)
    if kind == 'members':
        writer.writerow([
    'Member No',
    'Full Name',
    'Phone',
    'NRC/ID',
    'Group',
    'Member Type',
    'Committee Position',
    'Status'
        ])
        for m in Member.query.order_by(Member.member_no): writer.writerow([
    m.member_no,
    m.full_name,
    m.phone or '',
    m.national_id or '',
    m.group_name or '',
    m.member_type or 'Ordinary Member',
    m.committee_position or '',
    m.status
        ])
    elif kind == 'loans':
        writer.writerow([
            'Member No',
            'Member Name',
            'Principal',
            'Interest',
            'Total Due',
            'Total Paid',
            'Balance',
            'Status',
            'Purpose',
            'Issued On',
            'Due On',
            'Reviewed By',
            'Reviewed On',
            'Approved By',
            'Approved On',
            'Disbursed By',
            'Disbursed On',
            'Rejected On',
            'Rejection Reason',
            'Disbursement Method',
            'Disbursement Reference',])
        for l in Loan.query.order_by(Loan.issued_on.desc(), Loan.id.desc()).all():
            writer.writerow([
            l.member.member_no if l.member else '',
            l.member.full_name if l.member else '',
            l.principal,
            l.interest_amount,
            l.total_due,
            l.total_paid,
            l.balance,
            l.status,
            l.purpose or '',
            l.issued_on or '',
            l.due_on or '',
            l.reviewed_by or '',
            l.reviewed_on or '',
            l.approved_by or '',
            l.approved_on or '',
            l.disbursed_by or '',
            l.disbursed_on or '',
            l.rejected_on or '',
            l.rejection_reason or '',
            l.disbursement_method or '',
            l.disbursement_reference or '',
        ])
    else:
        writer.writerow(['Unsupported export'])
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={kind}.csv'})

@app.route('/admin/reset-transactions', methods=['POST'])
@login_required
@role_required('settings')
def reset_transactions():
    confirm = request.form.get('confirm')

    if confirm != 'RESET':
        flash('Reset not confirmed. Type RESET to proceed.', 'error')
        return redirect(url_for('settings'))

    try:
        # Delete child/detail records first
        CashBookEntry.query.delete()
        NotificationLog.query.delete()

        Repayment.query.delete()
        LoanGuarantor.query.delete()
        LoanInterest.query.delete()
        Loan.query.delete()

        FinePayment.query.delete()
        FinePenalty.query.delete()

        WelfareContribution.query.delete()
        WelfareClaim.query.delete()

        Distribution.query.delete()
        SavingsInterest.query.delete()
        Contribution.query.delete()

        MonthEndProcess.query.delete()

        JournalLine.query.delete()
        JournalEntry.query.delete()
       

        db.session.commit()

        log_audit(
            'RESET_TRANSACTIONS',
            'System',
            None,
            'All transactional records were reset. Members, users and settings were retained.'
        )

        flash('All transactional records have been reset successfully.')

    except Exception as e:
        db.session.rollback()
        flash(f'Reset failed: {str(e)}', 'error')

    return redirect(url_for('settings'))

def parse_date(value):
    if not value: return date.today()
    return datetime.strptime(value, '%Y-%m-%d').date()

def next_month_end(d):
    return (d.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

def next_quarter_meeting(d):
    for m in [3,6,9,12]:
        if d.month <= m:
            return date(d.year, m, 28)
    return date(d.year + 1, 3, 28)


def ensure_schema():
    # Add workflow columns when upgrading an existing SQLite database.
    if db.engine.url.drivername != 'sqlite':
        return
    columns = {row[1] for row in db.session.execute(db.text('PRAGMA table_info(loan)')).fetchall()}
    additions = {
        'reviewed_by': 'VARCHAR(120)',
        'disbursed_by': 'VARCHAR(120)',
        'reviewed_on': 'DATE',
        'approved_on': 'DATE',
        'disbursed_on': 'DATE',
        'rejected_on': 'DATE',
        'rejection_reason': 'VARCHAR(250)',
    }
    for name, sqltype in additions.items():
        if name not in columns:
            db.session.execute(db.text(f'ALTER TABLE loan ADD COLUMN {name} {sqltype}'))
    db.session.commit()

def ensure_admin():
    if not User.query.filter_by(username='admin').first():
        db.session.add(User(username='admin', full_name='System Administrator', role='Administrator', password_hash=generate_password_hash('admin123')))
        db.session.commit()

@app.cli.command('init-db')
def init_db():
    """Initialize a clean production database. No demo members are created."""
    db.drop_all(); db.create_all()
    admin = User(username='admin', full_name='System Administrator', role='Administrator', password_hash=generate_password_hash('admin123'))
    db.session.add(admin)
    db.session.commit()
    db.session.add(AuditLog(user_id=admin.id, username='admin', full_name='System Administrator', role='Administrator', action='INIT_DB', entity='Database', details='Clean production database initialized with default admin user'))
    db.session.commit()
    print('Clean database initialized with default admin user: admin / admin123')

@app.cli.command('init-demo-db')
def init_demo_db():
    """Initialize a demo database with 250 sample members. Use for testing only."""
    db.drop_all(); db.create_all()
    for i in range(1, 251):
        db.session.add(Member(member_no=f'M{i:03d}', full_name=f'Member {i:03d}', phone=f'+260 97 {i:06d}', group_name=f'Group {((i-1)//25)+1}'))
    admin = User(username='admin', full_name='System Administrator', role='Administrator', password_hash=generate_password_hash('admin123'))
    db.session.add(admin)
    db.session.commit()
    db.session.add(AuditLog(user_id=admin.id, username='admin', full_name='System Administrator', role='Administrator', action='INIT_DEMO_DB', entity='Database', details='Demo database initialized with 250 sample members and default admin user'))
    db.session.commit()
    print('Demo database initialized with 250 members and default admin user: admin / admin123')

def ensure_loan_columns():
    columns = {
        'loan_no': 'VARCHAR(30)',
        'disbursement_method': 'VARCHAR(50)',
        'disbursement_reference': 'VARCHAR(100)',
    }

    for column, definition in columns.items():
        try:
            db.session.execute(
                db.text(
                    f'ALTER TABLE loan ADD COLUMN IF NOT EXISTS {column} {definition}'
                )
            )
            db.session.commit()
        except Exception:
            db.session.rollback()

def ensure_member_columns():
    columns = {
        'member_type': 'VARCHAR(50)',
        'committee_position': 'VARCHAR(100)',
        }

    for column, definition in columns.items():
        try:
            db.session.execute(
                db.text(
                    f'ALTER TABLE member ADD COLUMN IF NOT EXISTS {column} {definition}'
                )
            )
            db.session.commit()
        except Exception:
            db.session.rollback()

""" @app.route('/fix-loan-numbers')
@login_required
def fix_loan_numbers():
    loans = Loan.query.filter(
        (Loan.loan_no == None) | (Loan.loan_no == '')
    ).all()

    for loan in loans:
        loan.loan_no = f'LN{loan.id:04d}'

    db.session.commit()

    return f'Updated {len(loans)} loans.'
 """

def ensure_chart_of_accounts():
    """
    Creates mandatory system accounts if they do not already exist.
    Safe to run whenever the application starts.
    """

    required_accounts = [
        {
            "code": "1110",
            "name": "Loan Interest Receivable",
            "account_type": "Asset",
            "normal_balance": "Debit"
        },
        {
            "code": "1120",
            "name": "Fines Receivable",
            "account_type": "Asset",
            "normal_balance": "Debit"
        },
        {
            "code": "3900",
            "name": "Opening Balance Equity",
            "account_type": "Equity",
            "normal_balance": "Credit"
        }
    ]

    for item in required_accounts:
        account = Account.query.filter_by(
            code=item["code"]
        ).first()

        if account:
            continue

        db.session.add(
            Account(
                code=item["code"],
                name=item["name"],
                account_type=item["account_type"],
                normal_balance=item["normal_balance"],
                active=True
            )
        )

    db.session.commit()


def initialize_database():
    with app.app_context():
        db.create_all()

        ensure_month_end_columns()
        ensure_settings_columns()
        ensure_member_columns()
        ensure_loan_columns()
        ensure_opening_balance_columns()
        ensure_financial_year_columns()
        ensure_schema()
        ensure_chart_of_accounts()
        ensure_admin()
        

# Initialize tables when the application is imported by Gunicorn on Render.
initialize_database()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=os.environ.get('FLASK_DEBUG') == '1')
